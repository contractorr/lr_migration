##############################################################################################################
from datetime import date
import log 
logger = log.get_logger('root')
import glob
import os 
import mig_functions as mig
import openpyxl as oxl

# Destination of key files
os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrate These')
#os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrate - All Files')
#os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Test')
input_files = glob.glob('*.xlsx')            
dest_directory = r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrated'
log.update_handler(logger,'Compilation')
##############################################################################################################

# Replace backslashes in dest_directory
dest_directory = dest_directory.replace('\\','/')
dest_file = dest_directory + f'/Compilation_In_Out_Commitment.xlsx' 
wb = oxl.Workbook()
ws = wb.active
all_rows = []

dst_row_num = 2


ws['A1'].font = oxl.styles.Font(size=15)   
ws['B1'].font = oxl.styles.Font(size=15)   
ws['C1'].font = oxl.styles.Font(size=15)
ws['D1'].font = oxl.styles.Font(size=15)
ws['E1'].font = oxl.styles.Font(size=15)
ws['F1'].font = oxl.styles.Font(size=15)
ws['G1'].font = oxl.styles.Font(size=15)
ws['H1'].font = oxl.styles.Font(size=15)
ws['I1'].font = oxl.styles.Font(size=15)
ws[f'A1'] = 'Share'
ws[f'B1'] = 'Investee Fund'
ws[f'C1'] = 'Date'
ws[f'D1'] = 'Description'
ws[f'E1'] = 'Investment Amount'
ws[f'F1'] = 'Fee Amount'
ws[f'G1'] = 'Return of Cap Amount'
ws[f'H1'] = 'Capital Gains Amount'
ws[f'I1'] = 'Undrawn Delta'

for input_file in input_files:
    #if input_file == 'LRI_O shares@2021 Q1.xlsx':
    print(f'{input_file}-----------------------------------------------------------------')
    # Extract the file name and format it 
    if 'LRI_' in input_file:
        file_name = input_file.split("LRI_")[1].split("@")[0]
        file_name = file_name.replace('shares','Shares')
        prefix = file_name.split(' ')[0]
    else:
        file_name = 'Y Shares'   
        prefix = 'Y' 


    fund_details, fund_sheets, fund_sheets_euros, all_fund_sheets = mig.extract_funds(input_file,prefix) 
    """  
    for fund_name, sheet_names in all_fund_sheets.items():
        for sheet_name in sheet_names:
            if sheet_name in ['CHF I (2)','AstIV (2)']:
                print(f"{fund_name}: {sheet_name}, {input_file}")
    """   
    dst_rows = mig.compile_data(input_file, file_name)
    all_rows.append(dst_rows)

    current_investee_fund = None

    # Insert
    for i, dst_row in dst_rows.iterrows():
        if current_investee_fund == None or dst_row['fund'] != current_investee_fund:
            current_investee_fund = dst_row['fund']
        elif dst_row['fund'] == current_investee_fund:
            current_investee_fund 

        if dst_row['fees_fund_ccy'] > 0 and (dst_row['roc_fund_ccy'] > 0 or dst_row['capital_gains_fund_ccy'] > 0):# and dst_row['change_undrawn_fund_ccy'] > 0:
        
            # Insert the values in the destination row 
            ws[f'A{dst_row_num}'] = file_name
            ws[f'B{dst_row_num}'] = dst_row['fund_name']
            ws[f'C{dst_row_num}'] = dst_row['date']
            ws[f'D{dst_row_num}'] = dst_row['description']
            ws[f'E{dst_row_num}'] = dst_row['investments_fund_ccy']            
            ws[f'F{dst_row_num}'] = dst_row['fees_fund_ccy']
            ws[f'G{dst_row_num}'] = dst_row['roc_fund_ccy']
            ws[f'H{dst_row_num}'] = dst_row['capital_gains_fund_ccy']
            ws[f'I{dst_row_num}'] = dst_row['change_undrawn_fund_ccy']
        
            dst_row_num = dst_row_num + 1

    wb.save(dest_file)   


