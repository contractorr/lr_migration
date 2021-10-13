import pandas as pd
import numpy as np
import openpyxl as oxl
import logging
import datetime as dt
from fuzzywuzzy import fuzz, process
from re import search
import time

logger = logging.getLogger('root')
######################################################################################
#                                       Misc                                         #
######################################################################################
def extract_funds(file_name,share,env='UAT'):
    """ Inputs:
            file_name: the absolute path to the source file
        Extract the sheets that contain fund operations, along with details about the corresponding fund
    """
    # Define dictionaries we will populate
    fund_details = pd.DataFrame(columns=['Name','Vintage','Commitment','Commitment_Euros','Currency'])
    fund_sheets = {}
    fund_sheets_euros = {}
    all_fund_sheets = {}
    
    xl = pd.ExcelFile(file_name)

    # Get a list of funds under this share class
    if env == 'UAT':
        fund_import_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/05 Funds.xlsx'
    else:
        fund_import_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/05 DEV Funds.xlsx'
    fund_import = pd.read_excel(fund_import_file, index_col=None, header=2, usecols='D:Q',sheet_name=share)[2:] # First fund is the management fund 
    fund_import = fund_import['Fund'].to_list()
    for fund in fund_import:
        if search('.*- LR$',fund):
            fund_import.remove(fund)
    logging.info(f'\t\t\t---Fund Name Replacements---')
    print(f'\t\t\t---Fund Name Replacements---')
    # Loop through sheets
    
    for sheet in xl.book.worksheets:
        sheet_name = sheet.title
        if sheet.sheet_state == 'hidden':
            # Ignore hidden sheets
            pass
        elif (share == 'N' 
        # and sheet_name in n_funds_to_skip # According to email on 29/07/2021 we shouldn't import all the funds in N and O shares
        and sheet_name not in [# There should be 7 funds we care about in N Shares
                           'PeI4 (2)','PeI3 (2)','PeI4','PeI3'
                          ,'MEIF (2)','MEIF'
                          ,'GSCP (2)','GSCP (2) (€)','GSCP','GSCP (€)'
                          ,'GSIP (2)','GSIP (2) (€)','GSIP','GSIP (€)'
                          ,'Avista (2)','Avista (2) (€)','Avista','Avista (€)'
                          ,'GSMP V (2)','GSMP V (2) (€)','GSMP V','GSMP V (€)'
                          ,'Latour (2)','Latour']):
            
            pass 
        elif (share == 'O' 
        # and sheet_name in o_funds_to_skip)
        and sheet_name not in [# There should be 15 funds we care about in O Shares
                           'Ath (2)','Ath'
                          ,'Rho (2)','Rho (2) (€)','Rho','Rho (€)'
                          ,'RhoIII (2)','RhoIII'
                          ,'CHF I (2)','CHF I (2) (€)','CHF I','CHF I (€)'
                          ,'21CP III (2)','21CP III'
                          ,'AstIV (2)','AstIV'
                          ,'Mand (2)','Mand'
                          ,'EqxII (2)','EqxII'
                          ,'DGPA (2)','DGPA'
                          ,'CHF II (2)','CHF II (2) (€)','CHF II','CHF II (€)'
                          ,'RIH (2)','RIH'
                          ,'21 CP IV (2)','21 CP IV'
                          ,'Quercus (2)','Quercus'
                          ,'InvI V (2)','InvI V'
                          ,'InvI']):
            # N and O shares have duplicate sheets that we need to ignore
            pass 
        else:
            # Sheet is not hidden so read it in
            
            try:
                df = xl.parse(sheet_name,usecols='A:B',nrows=6)
            except OverflowError:
                logging.warning(f'Date error on {sheet_name}')
            
            if df.keys()[0] == 'Name' and 'Unnamed' not in df.keys()[1]:
                # We've found a fund sheet. Now store it's details
                fund_name_report = df.keys()[1].strip()
                fund_name_compare = replace_fund_name(share,fund_name_report)
                # Match what's in the sheet against what's in the official fund import file and find the right match 
                fund_name = process.extractOne(fund_name_compare,fund_import)[0]
                if fund_name != fund_name_report:
                    logging.info(f'\t\t\t{share}, {sheet_name}: From {fund_name_report} to {fund_name}')
                    print(f'\t\t\t{sheet_name}: From {fund_name_report} to {fund_name}')

                fund_vintage = df.iloc[0,1]

                if ((share == 'N' and sheet_name not in ['PeI4','PeI3','MEIF','GSCP','GSCP (€)','GSIP','GSIP (€)','Avista','Avista (€)','GSMP V','GSMP V (€)','Latour'])
                or (share == 'O' and sheet_name not in ['Ath','Rho','Rho (€)','RhoIII','CHF I','CHF I (€)','21CP III','AstIV','Mand','EqxII','DGPA','CHF II','CHF II (€)','RIH','21 CP IV','Quercus','InvI V'])):
                    # These are the funds we need to stick a - LR in front of
                    fund_name += ' - LR' 

                if fund_name in all_fund_sheets:
                    all_fund_sheets[fund_name].append(sheet_name)
                else:
                    all_fund_sheets[fund_name] = []
                    all_fund_sheets[fund_name].append(sheet_name)

                fund_currency = df.iloc[5,1]    

                # Store the € Sheets in a different dictionary to the others
                if fund_currency == 'EUR' or '€' in sheet_name: 
                    # We have found a fund where the fund currency != investor currency (euros)
                    if fund_name in fund_sheets_euros:
                        fund_sheets_euros[fund_name].append(sheet_name)
                    else:
                        fund_sheets_euros[fund_name] = []
                        fund_sheets_euros[fund_name].append(sheet_name)
                        
                    # This means the fund commitment 
                    fund_commitment = None     
                    fund_commitment_euros = df.iloc[3,1]
                else:
                    # Store the names of all the fund sheets associated with each fund  
                    if fund_name in fund_sheets:
                        fund_sheets[fund_name].append(sheet_name)
                    else:
                        fund_sheets[fund_name] = []
                        fund_sheets[fund_name].append(sheet_name)

                    fund_commitment = df.iloc[3,1]        
                    fund_commitment_euros = None
                    

                fund_details = fund_details.append({'Name': fund_name
                                                ,'Vintage': fund_vintage
                                                ,'Commitment': fund_commitment
                                                ,'Commitment_Euros': fund_commitment_euros
                                                ,'Currency': fund_currency}
                                                , ignore_index=True
                                                )
    logging.info(f'\t\t\t----------------------------')
    print(f'\t\t\t----------------------------')
    return fund_details, fund_sheets, fund_sheets_euros, all_fund_sheets      

def calc_fund_op_type(src_row,investee_fund=True): 
    """
        Inputs:
            src_row: data containing the information from the source file that needs to be mapped to the dst_row_num
        Outputs:
            fund_op_type: the fund operation type based on the amounts present on that row 
    """     
    try:
        # IF: Commitment
        if src_row['commitment'] != 0 and investee_fund:
            fund_op_type = 'IF: Commitment'
            fund_op_code = '' 
        elif src_row['commitment'] != 0 and src_row['commitment'] is not None and investee_fund == False:
            fund_op_type = src_row['description']
            fund_op_code = ''
        elif src_row['description'] == 'MF: Commitment':
            # Delete this
            fund_op_type = src_row['description']
            fund_op_code = ''
        elif src_row['description'] == 'MF: Transfer':
            fund_op_type = src_row['description']
            fund_op_code = ''   
        # IF: Call
        elif (src_row['investments'] + src_row['fees'] > 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0):
            if investee_fund:
                fund_op_type = 'IF: Call'
            else:
                fund_op_type = 'MF: Call'
            fund_op_code = 'CC'
        # Call is equal and opposite to fee 
        elif (src_row['investments'] + src_row['fees'] == 0
        and src_row['investments'] != 0
        and src_row['fees'] != 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0):
            if investee_fund:
                if 'impairment' in str.lower(src_row['description']):
                    fund_op_type = 'IF: Impairment'
                else:
                    fund_op_type = 'IF: Call'
            else:
                fund_op_type = 'MF: Call'
            fund_op_code = 'CC' 
        # IF: Return Of Call
        elif (src_row['investments'] + src_row['fees'] < 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0):
            if investee_fund:
                fund_op_type = 'IF: Return Of Call'
            else:
                fund_op_type = 'MF: Return Of Call (Negative Call)'
            fund_op_code = 'CD' 
        # IF: Distribution
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] + src_row['capital gains'] != 0):
            if investee_fund:
                fund_op_type = 'IF: Distribution'
            else:
                fund_op_type = 'MF: Distribution'
            fund_op_code = 'CD' 
        # Return of Capital is equal and opposite to Capital Gain
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] + src_row['capital gains'] == 0
        and src_row['return of capital'] != 0 
        and src_row['capital gains'] != 0):
            if investee_fund:
                if 'impairment' in str.lower(src_row['description']):
                    fund_op_type = 'IF: Impairment'
                else:
                    fund_op_type = 'IF: Distribution'
            else:
                fund_op_type = 'MF: Distribution'
            fund_op_code = 'CD' 
        # IF: Mixed operation
        elif (src_row['investments'] + src_row['fees'] != 0
        and src_row['return of capital'] + src_row['capital gains'] != 0):
            if investee_fund:
                fund_op_type = 'IF: Mixed operation'
            else:
                fund_op_type = 'MF: Mixed operation'
            fund_op_code = 'CCCD' 
        # IF: Accounting Valuation
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0
        and src_row['fair value'] >= 0
        and '(est)' in str.lower(src_row['description'])):
            if investee_fund:
                fund_op_type = 'IF: Accounting Valuation'
            else:
                fund_op_type = 'MF: Net Asset Value'
            fund_op_code = '' 
        # IF: Official NAV
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0 
        and src_row['fair value'] > 0
        and '(est)' not in str.lower(src_row['description'])):
            if investee_fund:
                fund_op_type = 'IF: Official NAV'
            else:
                fund_op_type = 'MF: Net Asset Value'
            fund_op_code = 'NAV' 
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0 
        and src_row['fair value'] < 0
        and not investee_fund):
            fund_op_type = 'MF: Net Asset Value'
            fund_op_code = 'NAV' 
        elif (src_row['investments'] == 0
        and src_row['fees'] == 0
        and src_row['return of capital'] == 0
        and src_row['capital gains'] == 0 
        and src_row['fair value'] == 0):
            fund_op_type = 'REMOVE'
            fund_op_code = 'REMOVE' 
        else:
            if investee_fund:
                fund_op_type = 'IF: Other'
            else:
                fund_op_type = 'MF: Other'
            fund_op_code = ''   
        # IF: Liquidation
    except:
        logger.exception(src_row)

    # Return the fund_op_type
    try:
        return fund_op_type, fund_op_code
    except UnboundLocalError:
        print(src_row, investee_fund)

def append_row(fund_name, fund_ccy, file_name, src_row, src_row_euros, dst_rows,fx_rates=None,investee_fund=True):
    """ 
        Inputs:
            fund: the name of the fund 
            currency: the currency of the fund 
            file_name: the formatted name of the file we'll use in rows F and G
            src_row: data containing the information from the source file that needs to be mapped to the dst_row_num
            src_row_euros: if the fund currency != investor currency, then this is the corresponding row from the sheet with investor currency (euro) amounts 
            dst_rows: a dataframe to collate the destination rows we will insert into the import template 
            investee_fund: boolean that tells us if it's an investee fund. If False, it's a managed fund and we don't need to insert accounting NAV rows.
        Process the inputs and append a row into our record of the output, dst_rows
    """
    # Variables
    log_row_ind = False 
    categorised_ind = True
    negative_redraw_ind = False

    # Work out the operation type
    fund_op_type, fund_op_code = calc_fund_op_type(src_row, investee_fund)

    # Every time we find an Official NAV, insert the latest Accounting Valuation if one exists
    if (fund_op_code == 'NAV' 
    and src_row['latest accounting nav'] is not None 
    and src_row['latest accounting nav'] != 0
    and investee_fund
    ):
        
        accounting_nav_row = {
            'commitment': 0,
            'commitment_euros': None,
            'date': src_row['date'],
            'description': 'NAV (est)',
            'investments': 0,
            'return of capital': 0,
            'capital gains': 0,
            'fees': 0,
            'fair value': src_row['latest accounting nav']
        }
        accounting_nav_row_euros = {
            'commitment': 0,
            'commitment_euros': None,
            'date': src_row['date'],
            'description': 'NAV (est)',
            'investments': 0,
            'return of capital': 0,
            'capital gains': 0,
            'fees': 0,
            'fair value': src_row_euros['latest accounting nav']
        } 

        if file_name == 'X Shares' and fund_name == 'Zodiac' and src_row['date'] >= dt.date(2019,3,31):
            accounting_nav_row = {
                'commitment': 0,
                'commitment_euros': None,
                'date': src_row['date'],
                'description': 'NAV (est)',
                'investments': 0,
                'return of capital': 0,
                'capital gains': 0,
                'fees': 0,
                'fair value': src_row['fair value']
            }
            accounting_nav_row_euros = {
                'commitment': 0,
                'commitment_euros': None,
                'date': src_row['date'],
                'description': 'NAV (est)',
                'investments': 0,
                'return of capital': 0,
                'capital gains': 0,
                'fees': 0,
                'fair value': src_row_euros['fair value']
            } 

        # Insert total fund commitment row                
        dst_rows = append_row(fund_name                                                   
                             ,fund_ccy
                             ,file_name                                     
                             ,accounting_nav_row    
                             ,accounting_nav_row_euros                                            
                             ,dst_rows    
                             ,fx_rates=fx_rates
                             ,investee_fund=investee_fund                                      
                             )                                                     

    if src_row['commitment'] == 0:  
        src_row['commitment'] = None # We no longer need this to be 0 so set this to None so we don't insert zeros 

    # if file_name == 'Y Shares' and investee_fund:
    #     # for Y shares all call operations are just commitments 
    #     investor  = f'{file_name} - LR'
    #     fund = f'{file_name} - LR in {fund_name}' 
    #     fund_op_type = 'IF: Commitment'
    #     fund_op_code = ''
    #     if src_row['investments'] is not None:
    #         inv = src_row['investments']
    #     else:
    #         inv = 0
    #     if src_row['fees'] is not None:
    #         fee = src_row['fees']
    #     else:
    #         fee = 0
    #     src_row['commitment'] = inv + fee
    #     src_row['investments'] = 0 
    #     src_row['fees'] = 0
    if file_name in ['AGp Shares','AHp Shares','AHs Shares','AIp Shares','AIs Shares'] and investee_fund:
        file_name = file_name.replace('p ',' ')
        file_name = file_name.replace('s ',' ')
        investor  = f'{file_name} - LR'
        fund = f'{file_name} - LR in {fund_name}'   
    elif investee_fund:
        investor  = f'{file_name} - LR'
        fund = f'{file_name} - LR in {fund_name}' 
    else:
        investor  = f'{file_name}'
        fund = f'{fund_name} - LR'   

    if fund_op_type == 'IF: Other':
        log_uncategorised_fund_ops(fund_name, file_name, src_row)
        return dst_rows

    try:
        if fund_op_type == 'REMOVE':
            return dst_rows
        else:
            tot_fund_ccy = sum(filter(None,[src_row['commitment'],src_row['investments'],src_row['fees'],src_row['return of capital'],src_row['capital gains'],src_row['fair value']]))
            tot_investor_ccy = sum(filter(None,[src_row['commitment_euros'],src_row_euros['investments'],src_row_euros['fees'],src_row_euros['return of capital'],src_row_euros['capital gains'],src_row_euros['fair value']]))
            # Calculate the fx rate and store it in the fx_rates_record array. Be careful not to name it fx_rates because that's already dataframe we've passed in. 
            fx_rates_record = list()
            if tot_fund_ccy != 0 and tot_investor_ccy != 0 and round(tot_fund_ccy,2) != round(tot_investor_ccy,2):
                fx_rate = tot_investor_ccy/tot_fund_ccy
                fx_rates_record.append(fx_rate)
            elif round(tot_fund_ccy,2) != 0 and (tot_investor_ccy != tot_investor_ccy or round(tot_investor_ccy,2) == 0):
                fx_rate = None
            else:
                fx_rate = 1
                fx_rates_record.append(fx_rate)

            # Make sure the fx rate we calculate matches no matter what value we use. 
            # Where any value is populated, calculate the fx_rate and store it in our list

            if src_row['investments'] is not None and src_row_euros['investments'] is not None and src_row['investments'] != 0 and src_row_euros['investments'] != 0: 
                fx_investments = src_row_euros['investments']/src_row['investments']
                fx_rates_record.append(fx_investments)
            else:
                fx_investments = None
            if src_row['fees'] is not None and src_row_euros['fees'] is not None and src_row['fees'] != 0 and src_row_euros['fees'] != 0: 
                fx_fees = src_row_euros['fees']/src_row['fees']
                fx_rates_record.append(fx_fees)
            else:
                fx_fees = None
            if src_row['return of capital'] is not None and src_row_euros['return of capital'] is not None and src_row['return of capital'] != 0 and src_row_euros['return of capital'] != 0: 
                fx_return_of_capital = src_row_euros['return of capital']/src_row['return of capital']
                fx_rates_record.append(fx_return_of_capital)
            else:
                fx_return_of_capital = None
            if src_row['capital gains'] is not None and src_row_euros['capital gains'] is not None and src_row['capital gains'] != 0 and src_row_euros['capital gains'] != 0:
                fx_capital_gains = src_row_euros['capital gains']/src_row['capital gains']
                fx_rates_record.append(fx_capital_gains)
            else:
                fx_capital_gains = None
            if src_row['fair value'] is not None and src_row_euros['fair value'] is not None and src_row['fair value'] != 0 and src_row_euros['fair value'] != 0: 
                fx_fair_value = src_row_euros['fair value']/src_row['fair value']
                fx_rates_record.append(fx_fair_value)
            else:
                fx_fair_value = None
            # Find out if we calculated multiple different fx rates
            fx_rates_record = pd.Series(fx_rates_record)
            fx_rates_record = fx_rates_record[fx_rates_record != 0]
            fx_rates_rounded = fx_rates_record.round(4)
            
            if isinstance(src_row['date'], int):
                fund_op_date = dt.date(src_row['date'],1,1)
            else:
                fund_op_date = dt.date(src_row['date'].year,src_row['date'].month,src_row['date'].day)

            if len(fx_rates_rounded.unique()) > 1:
                # We have an fx_rate mismatch. This usually happens at the quarter end dates
                month_end_date = dt.date(2022,3,31)
                # This will create a list of all month end dates from 2005-11-30 to 2022-03-31
                month_end_dates = [month_end_date - pd.DateOffset(months=(x*4)) for x in range(50)]
                tomorrow = fund_op_date + pd.DateOffset(days=1)
                if tomorrow in month_end_dates:
                    # It's month end tomorrow so change the fund op date to tomorrow    
                    logging.info(f"\t\tFund op date moved to month end for {fund_op_type} in {fund} on {fund_op_date} ({src_row['description']})")
                    print(f"\t\tFund op date moved to month end for {fund_op_type} in {fund} on {fund_op_date} ({src_row['description']})")
                    fund_op_date = tomorrow
                    fx_rate = fx_rates_rounded[0]
                else:
                    log_fx_mismatch(file_name
                                   ,fund_op_type
                                   ,fund
                                   ,src_row
                                   ,src_row_euros
                                   ,fund_ccy
                                   ,fx_rates
                                   ,fx_investments
                                   ,fx_fees
                                   ,fx_return_of_capital
                                   ,fx_capital_gains
                                   ,fx_fair_value)
                    fx_rate = None
                # fair_val_fx = round(src_row_euros['fair value']/src_row['fair value'],4)
                # fx_rates_rounded = fx_rates_rounded[fx_rates_rounded != fair_val_fx]
                # if len(fx_rates_rounded.unique()) != 1:
                #     print("Unable to find correct fx rate!")
                #     fx_rate = None
                # else:
                #     tot_fund_ccy = sum(filter(None,[src_row['commitment'],src_row['investments'],src_row['fees'],src_row['return of capital'],src_row['capital gains']]))
                #     tot_investor_ccy = sum(filter(None,[src_row['commitment_euros'],src_row_euros['investments'],src_row_euros['fees'],src_row_euros['return of capital'],src_row_euros['capital gains']]))
                #     fx_rate = tot_investor_ccy/tot_fund_ccy
            elif len(fx_rates_rounded.unique()) == 1:
                fx_rate = fx_rates_rounded.unique()[0]
            else:
                fx_rate = None

            dst_row = {'fund_op_type': fund_op_type
                      ,'fund_name': fund_name
                      ,'investor': investor
                      ,'fund': fund
                      ,'date': fund_op_date  
                      ,'fund_op_code': fund_op_code
                      ,'description': src_row['description']
                      ,'settlement_date': fund_op_date 
                      ,'op_currency': fund_ccy
                      ,'commitment_fund_ccy': src_row['commitment']
                      ,'commitment_investor_ccy': src_row['commitment_euros']
                      ,'investments_fund_ccy': src_row['investments']  
                      ,'investments_investor_ccy': src_row_euros['investments'] 
                      ,'fees_fund_ccy': src_row['fees'] 
                      ,'fees_investor_ccy': src_row_euros['fees']
                      ,'fees_fund_ccy_inside_commitment': 0
                      ,'fees_investor_ccy_inside_commitment': 0
                      ,'roc_fund_ccy':  src_row['return of capital'] 
                      ,'roc_investor_ccy': src_row_euros['return of capital']
                      ,'capital_gains_fund_ccy': src_row['capital gains']  
                      ,'capital_gains_investor_ccy': src_row_euros['capital gains']
                      ,'fair_value_fund_ccy': src_row['fair value'] 
                      ,'fair_value_investor_ccy': src_row_euros['fair value']  
                      ,'undrawn_fund_ccy': None
                      ,'undrawn_investor_ccy': None
                      ,'redraw_fund_ccy': None
                      ,'redraw_investor_ccy': None
                      ,'impairment_fund_ccy': None
                      ,'impairment_investor_ccy': None
                      ,'fx_rate': fx_rate
                      ,'orig_fee': src_row['fees']
                      }
    except:
        logger.exception('Append failed!')

    if investee_fund == False:
        dst_row['issue'] = src_row['issue']
        if fund_op_type == 'IF: Commitment': 
            date = dt.date(src_row['date'].year,1,1)
            dst_row['date'] = date
            dst_row['no_shares'] = None
            dst_row['nominal'] = src_row['shares_issued']
        elif fund_op_type == 'MF: Transfer':  
            dst_row['no_shares'] = src_row['shares_issued']
            dst_row['commitment_investor_ccy'] = src_row['shares_issued']
            dst_row['nominal'] = src_row['shares_issued']
        else:
            dst_row['nominal'] = None
        dst_row['shares_issued'] = src_row['shares_issued']
        dst_row['investor_trunc'] = file_name


    if fund_op_type == 'IF: Commitment' and investee_fund is False:
        date = dt.date(src_row['date'].year,1,1)
        dst_row['date'] = date

    if 'undrawn' in src_row.keys() and investee_fund:
        if dst_rows[dst_rows['fund']==fund]['undrawn_fund_ccy'].last_valid_index() and src_row['undrawn'] == src_row['undrawn']:
            # there is already a row in dst_rows where undrawn is populated so calculate the change 
            change_undrawn_fund_ccy = src_row['undrawn'] - dst_rows.loc[dst_rows[dst_rows['fund']==fund]['undrawn_fund_ccy'].last_valid_index(),'undrawn_fund_ccy']
            change_undrawn_investor_ccy = src_row_euros['undrawn'] - dst_rows.loc[dst_rows[dst_rows['fund']==fund]['undrawn_investor_ccy'].last_valid_index(),'undrawn_investor_ccy']
            if (change_undrawn_fund_ccy != change_undrawn_fund_ccy) or (dst_rows.loc[dst_rows[dst_rows['fund']==fund]['undrawn_fund_ccy'].last_valid_index(),'undrawn_fund_ccy'] == 0):
                change_undrawn_fund_ccy = 0
                change_undrawn_investor_ccy = 0
        elif src_row['undrawn'] == src_row['undrawn']:
            # first row use the commitment instead of the last valid index
            change_undrawn_fund_ccy = src_row['undrawn'] - dst_rows.loc[dst_rows[dst_rows['fund']==fund]['commitment_fund_ccy'].first_valid_index(),'commitment_fund_ccy']
            change_undrawn_investor_ccy = src_row_euros['undrawn'] - dst_rows.loc[dst_rows[dst_rows['fund']==fund]['commitment_investor_ccy'].first_valid_index(),'commitment_investor_ccy']
        else:
            change_undrawn_fund_ccy = 0
            change_undrawn_investor_ccy = 0
            
        dst_row['change_undrawn_fund_ccy'] = change_undrawn_fund_ccy
        dst_row['change_undrawn_investor_ccy'] = change_undrawn_investor_ccy
        
        net_change_undrawn_fund_ccy = round((change_undrawn_fund_ccy + dst_row['investments_fund_ccy']),2)
        net_change_undrawn_investor_ccy = round((change_undrawn_investor_ccy + dst_row['investments_investor_ccy']),2)

        # Determine what fraction of the fee amount is inside commitment and what fraction of the distributed amount is redrawable 
        if net_change_undrawn_fund_ccy != 0:
            if (dst_row['fees_fund_ccy'] != 0
            and src_row['return of capital'] == 0 
            and src_row['capital gains'] == 0):
                # We can count that the redrawable amount = 0 so the undrawn delta should tell us everything we need to know 
                dst_row['fees_fund_ccy_inside_commitment'] = -net_change_undrawn_fund_ccy
                dst_row['fees_investor_ccy_inside_commitment'] = -net_change_undrawn_investor_ccy
                dst_row['fees_fund_ccy'] = round((dst_row['fees_fund_ccy'] - dst_row['fees_fund_ccy_inside_commitment']),2)
                dst_row['fees_investor_ccy'] = round((dst_row['fees_investor_ccy'] - dst_row['fees_investor_ccy_inside_commitment']),2) 
                dst_row['redraw_fund_ccy'] = 0
                dst_row['redraw_investor_ccy'] = 0
            elif (dst_row['fees_fund_ccy'] == 0
            and src_row['return of capital'] + src_row['capital gains'] != 0):
                # The fee is 0 so we can work out the fraction of the distributed amount that is redrawable 
                dst_row['redraw_fund_ccy'] = -net_change_undrawn_fund_ccy
                dst_row['redraw_investor_ccy'] = -net_change_undrawn_investor_ccy
            elif (dst_row['fees_fund_ccy'] == 0
            and src_row['return of capital'] == 0 
            and src_row['capital gains'] == 0):
                # Nothing to categorise. Probably a nav operation. 
                pass
            elif (round((net_change_undrawn_fund_ccy + dst_row['fees_fund_ccy']),2) == 0
            and src_row['return of capital'] + src_row['capital gains'] != 0):
                # We assume the fee is fully inside commitment, with redrawable amount = 0
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = -net_change_undrawn_fund_ccy
                dst_row['fees_investor_ccy_inside_commitment'] = -net_change_undrawn_investor_ccy
                dst_row['fees_fund_ccy'] = 0
                dst_row['fees_investor_ccy'] = 0
                dst_row['redraw_fund_ccy'] = 0
                dst_row['redraw_investor_ccy'] = 0
            elif round((net_change_undrawn_fund_ccy - dst_row['roc_fund_ccy']),2) == 0:
                # We assume the fee is fully outside commitment, with redrawable amount = roc
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = 0
                dst_row['fees_investor_ccy_inside_commitment'] = 0
                dst_row['redraw_fund_ccy'] = abs(dst_row['roc_fund_ccy'])
                dst_row['redraw_investor_ccy'] = abs(dst_row['roc_fund_ccy'])
                if dst_row['roc_fund_ccy'] < 0 and dst_row['roc_fund_ccy'] != -dst_row['capital_gains_fund_ccy']:
                    negative_redraw_ind = True
            elif round((net_change_undrawn_fund_ccy - dst_row['capital_gains_fund_ccy']),2) == 0:
                # We assume the fee is fully outside commitment, with redrawable amount = cg
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = 0
                dst_row['fees_investor_ccy_inside_commitment'] = 0
                dst_row['redraw_fund_ccy'] = abs(dst_row['capital_gains_fund_ccy'])
                dst_row['redraw_investor_ccy'] = abs(dst_row['capital_gains_fund_ccy'])
                if dst_row['capital_gains_fund_ccy'] < 0 and dst_row['roc_fund_ccy'] != -dst_row['capital_gains_fund_ccy']:
                    negative_redraw_ind = True     
            elif round((net_change_undrawn_fund_ccy + dst_row['fees_fund_ccy'] - dst_row['roc_fund_ccy']),2) == 0:
                # We assume the fee is fully inside commitment, with redrawable amount = roc
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = dst_row['fees_fund_ccy']
                dst_row['fees_investor_ccy_inside_commitment'] = dst_row['fees_investor_ccy']
                dst_row['fees_fund_ccy'] = 0
                dst_row['fees_investor_ccy'] = 0
                dst_row['redraw_fund_ccy'] = abs(dst_row['roc_fund_ccy'])
                dst_row['redraw_investor_ccy'] = abs(dst_row['roc_investor_ccy'])
                if dst_row['capital_gains_fund_ccy'] < 0 and dst_row['roc_fund_ccy'] != -dst_row['capital_gains_fund_ccy']:
                    negative_redraw_ind = True 
            elif round((net_change_undrawn_fund_ccy + dst_row['fees_fund_ccy'] - dst_row['capital_gains_fund_ccy']),2) == 0:
                # We assume the fee is fully inside commitment, with redrawable amount = cg
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = dst_row['fees_fund_ccy']
                dst_row['fees_investor_ccy_inside_commitment'] = dst_row['fees_investor_ccy']
                dst_row['fees_fund_ccy'] = 0
                dst_row['fees_investor_ccy'] = 0
                dst_row['redraw_fund_ccy'] = abs(dst_row['capital_gains_fund_ccy'])
                dst_row['redraw_investor_ccy'] = abs(dst_row['capital_gains_investor_ccy'])
                if dst_row['capital_gains_fund_ccy'] < 0 and dst_row['roc_fund_ccy'] != -dst_row['capital_gains_fund_ccy']:
                    negative_redraw_ind = True 
            elif round((net_change_undrawn_fund_ccy + dst_row['fees_fund_ccy'] - (dst_row['roc_fund_ccy'] + dst_row['capital_gains_fund_ccy'])),2) == 0:
                # Chances are roc is redrawable and fee is inside commitment - possible source of error
                log_row_ind = True 
                dst_row['fees_fund_ccy_inside_commitment'] = dst_row['fees_fund_ccy']
                dst_row['fees_investor_ccy_inside_commitment'] = dst_row['fees_investor_ccy']
                dst_row['fees_fund_ccy'] = 0
                dst_row['fees_investor_ccy'] = 0
                dst_row['redraw_fund_ccy'] = dst_row['roc_fund_ccy'] + dst_row['capital_gains_fund_ccy']
                dst_row['redraw_investor_ccy'] = dst_row['roc_investor_ccy'] + dst_row['capital_gains_investor_ccy']
            else:
                log_row_ind = True 
                categorised_ind = False   
        else:
            # for some reason undrawn is null and there are now previous rows with undrawn populated
            dst_row['change_undrawn_fund_ccy'] = 0
            dst_row['change_undrawn_investor_ccy'] = 0

        dst_row['undrawn_fund_ccy'] = src_row['undrawn']
        dst_row['undrawn_investor_ccy'] = src_row_euros['undrawn']
    else:
        try:
            dst_row['change_undrawn_fund_ccy'] = 0
            dst_row['change_undrawn_investor_ccy'] = 0
        except UnboundLocalError:
            print(f"\t\tUnboundLocalError: {src_row}")

    # if dst_row['fees_fund_ccy_inside_commitment'] != 0 and dst_row['fees_fund_ccy'] != 0:
    #     print(f"\t\tWarning: row where fee is not fully in or outside commitment: {dst_row['fund']},{dst_row['date']},{dst_row['description']},{src_row['investments']},{src_row['fees']},{src_row['return of capital']},{src_row['capital gains']},{src_row['fair value']},{-round(change_undrawn_fund_ccy,2)}")
    #     logging.warning(f"\t\tWarning: row where fee is not fully in or outside commitment: {dst_row['fund']},{dst_row['date']},{dst_row['description']},{src_row['investments']},{src_row['fees']},{src_row['return of capital']},{src_row['capital gains']},{src_row['fair value']},{-round(change_undrawn_fund_ccy,2)}")

    dst_rows = dst_rows.append(dst_row, ignore_index=True)

    if log_row_ind:
        # Location of the file 
        uncategorised_fees_cap_gains = 'C:/Users/RajContractor/Documents/Python Files/Dev/LR Migration/Migrated/Fees and Return of Capital Categorisation.xlsx'
        # Add relevant columns
        src_row['fees_outside_commitment'] = dst_row['fees_fund_ccy']
        src_row['fees_inside_commitment'] = dst_row['fees_fund_ccy_inside_commitment']
        src_row['redrawable_amount'] = dst_row['redraw_fund_ccy']
        src_row['share'] = investor 
        src_row['fund'] = fund_name
        src_row['undrawn delta'] = -change_undrawn_fund_ccy
        if negative_redraw_ind:
            src_row['redraw negative'] = 'Yes'
        else:
            src_row['redraw negative'] = 'No'

        if categorised_ind:
            src_row['categorised'] = 'Yes'
        else:
            src_row['categorised'] = 'No'

        # Read in what's there and add the new row 
        try:
            fund_ops = pd.read_excel(uncategorised_fees_cap_gains, index_col=None)
            fund_ops = fund_ops.append(src_row, ignore_index=True)
        except FileNotFoundError:
            fund_ops = pd.DataFrame()
            fund_ops = fund_ops.append(src_row, ignore_index=True)

        # Output everything to excel
        fund_ops.to_excel(uncategorised_fees_cap_gains,index=False, columns=['share','fund','date','description','undrawn delta','investments','fees','fees_inside_commitment','fees_outside_commitment','return of capital','capital gains','redrawable_amount','fair value','categorised','redraw negative'])

    return dst_rows

def insert_row(dst_row, dst_row_num, dst_active_sheet, investee_fund=True):
    """ 
        Inputs:
            dst_row: the row we want to insert 
            dst_row_num: the active row in the destination file 
            dst_active_sheet: the active sheet in the destination file 
        Take the dst_row dataframe and transfer its contents to our excel template  
    """
    # Split out 'IF: Other' operations into a separate file    
    if dst_row['fund_op_type'] == 'IF: Other':
        separate_other_fund_ops(dst_row)
        return dst_row_num
    
    # if dst_row['fund_op_type'] == 'IF: Accounting Valuation':
    # if dst_row['fund_op_type'] == 'IF: Official NAV':

    try:
        if investee_fund:
            # Insert the values in the destination row 
            dst_active_sheet[f'C{dst_row_num}'] = dst_row['fund_op_type']                                           # -------------------------------------- Type    
            dst_active_sheet[f'D{dst_row_num}'] = dst_row['fund_name']                                              # -------------------------------------- Fund.Fund 
            # dst_active_sheet[f'E{dst_row_num}'] = only for managed funds                                          # -------------------------------------- 
            dst_active_sheet[f'F{dst_row_num}'] = dst_row['investor']                                               # -------------------------------------- Investors (BP).Subscriber.Investor.Investor
            dst_active_sheet[f'G{dst_row_num}'] = dst_row['fund']                                                   # -------------------------------------- Investors (BP).Subscriber.Main vehicle.Fund.Fund
            dst_active_sheet[f'H{dst_row_num}'] = dst_row['date']                                                   # -------------------------------------- Date
            dst_active_sheet[f'I{dst_row_num}'] = dst_row['index']                                                  # -------------------------------------- Index
            #dst_active_sheet[f'K{dst_row_num}'] = dst_row['description']                                            # -------------------------------------- Operation
            dst_active_sheet[f'L{dst_row_num}'] = 'Shared'                                                          # -------------------------------------- Region
            #dst_active_sheet[f'M{dst_row_num}'] = dst_row['description_trunc']                                     # -------------------------------------- Reason
            dst_active_sheet[f'N{dst_row_num}'] = dst_row['settlement_date']                                        # -------------------------------------- Settlement date
            #dst_active_sheet[f'O{dst_row_num}'] = GL date
            #dst_active_sheet[f'P{dst_row_num}'] = Send date 
            dst_active_sheet[f'Q{dst_row_num}'] = dst_row['op_currency']                                            # -------------------------------------- Operation currency 
            dst_active_sheet[f'R{dst_row_num}'] = 'Commitment'                                                      # -------------------------------------- Share/Series
            dst_active_sheet[f'S{dst_row_num}'] = 'No'                                                              # -------------------------------------- Draft
            # dst_active_sheet[f'T{dst_row_num}'] = Bank Account
            dst_active_sheet[f'U{dst_row_num}'] = 'No'                                                              # -------------------------------------- Cancelled
            dst_active_sheet[f'V{dst_row_num}'] = 'Commitment'                                                      # -------------------------------------- Investors (BP).Fund share.Category
            dst_active_sheet[f'W{dst_row_num}'] = 1 #dst_row['index']                                               # -------------------------------------- Investors (BP).Fund share.Share Index
            # dst_active_sheet[f'X{dst_row_num}'] = Investors (BP).Shares committed
            dst_active_sheet[f'Z{dst_row_num}'] = dst_row['commitment_fund_ccy']                                    # --------------------------------------  Investors (BP).Share Commitment 
            dst_active_sheet[f'AA{dst_row_num}'] = dst_row['commitment_investor_ccy']                               # -------------------------------------- Investors (BP).Share Commitment
            # dst_active_sheet[f'AB{dst_row_num}'] = src_row['commitment']                                          # -------------------------------------- Investors (BP).FX Rate -> Investor
            # dst_active_sheet[f'AC{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AD{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AE{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AF{dst_row_num}'] = src_row['commitment']
            #dst_active_sheet[f'AG{dst_row_num}'] = dst_row['commitment_investor_ccy']                               # -------------------------------------- Investors (BP).Share Commitment
            # dst_active_sheet[f'AI{dst_row_num}'] = src_row['commitment']
            dst_active_sheet[f'AJ{dst_row_num}'] = dst_row['investments_fund_ccy']                                  # -------------------------------------- Investors (BP).Share Call: Investments
            dst_active_sheet[f'AK{dst_row_num}'] = dst_row['investments_investor_ccy']                              # -------------------------------------- Investors (BP).Investor Call: Investments
            #dst_active_sheet[f'AL{dst_row_num}'] = src_row_euros['investments']                                    # -------------------------------------- Share.Allocation : Fund Call: Legal Fees outside commitment
            dst_active_sheet[f'AN{dst_row_num}'] = dst_row['legal_fees_fund_ccy']                                   # -------------------------------------- Investors (BP).Share Call: Legal Fees outside commitment
            dst_active_sheet[f'AO{dst_row_num}'] = dst_row['legal_fees_investor_ccy']                               # -------------------------------------- Investors (BP).Investor Call: Legal Fees outside commitment
            dst_active_sheet[f'AR{dst_row_num}'] = dst_row['fees_fund_ccy_inside_commitment']                       # -------------------------------------- Investors (BP).Share Call: Management Fees in commitment
            dst_active_sheet[f'AS{dst_row_num}'] = dst_row['fees_investor_ccy_inside_commitment']                   # -------------------------------------- Investors (BP).Investor Call: Management Fees in commitment
            dst_active_sheet[f'AV{dst_row_num}'] = dst_row['fees_fund_ccy']                                         # -------------------------------------- Investors (BP).Share Call: Management Fees outside commitment
            dst_active_sheet[f'AY{dst_row_num}'] = dst_row['org_costs_fund_ccy']                                    # -------------------------------------- Investors (BP).Share Call: Organizational Costs outside commitment
            dst_active_sheet[f'AZ{dst_row_num}'] = dst_row['org_costs_investor_ccy']                                # -------------------------------------- Investors (BP).Investor Call: Organizational Costs outside commitment
            dst_active_sheet[f'BG{dst_row_num}'] = dst_row['other_expenses_fund_ccy']                               # -------------------------------------- Investors (BP).Share Call: Other Expenses outside commitment
            dst_active_sheet[f'BH{dst_row_num}'] = dst_row['partnership_expenses_fund_ccy']                         # -------------------------------------- Investors (BP).Share Call: Partnership Expenses outside commitment
            dst_active_sheet[f'BI{dst_row_num}'] = dst_row['redraw_fund_ccy']                                       # -------------------------------------- Investors (BP).Share Redrawable amount
            dst_active_sheet[f'BR{dst_row_num}'] = dst_row['fees_investor_ccy']                                     # -------------------------------------- Investors (BP).Investor Call: Management Fees outside commitment
            dst_active_sheet[f'BS{dst_row_num}'] = dst_row['other_expenses_investor_ccy']                           # -------------------------------------- Investors (BP).Investor Call: Other Expenses outside commitment
            dst_active_sheet[f'BT{dst_row_num}'] = dst_row['partnership_expenses_investor_ccy']                     # -------------------------------------- Investors (BP).Investor Call: Partnership Expenses outside commitment
            dst_active_sheet[f'BU{dst_row_num}'] = dst_row['redraw_investor_ccy']                                   # -------------------------------------- Investors (BP).Investor Redrawable amount
            dst_active_sheet[f'BV{dst_row_num}'] = dst_row['investments_op_ccy']                                    # -------------------------------------- Investors (BP).Op Call: Investments
            dst_active_sheet[f'BW{dst_row_num}'] = dst_row['legal_fees_op_ccy']                                     # -------------------------------------- Investors (BP).Op Call: Legal Fees outside commitment
            dst_active_sheet[f'BX{dst_row_num}'] = dst_row['fees_op_ccy_inside_commitment']                         # -------------------------------------- Investors (BP).Op Call: Management Fees in commitment
            dst_active_sheet[f'BY{dst_row_num}'] = dst_row['fees_op_ccy']                                           # -------------------------------------- Investors (BP).Op Call: Management Fees outside commitment
            dst_active_sheet[f'BZ{dst_row_num}'] = dst_row['org_costs_op_ccy']                                      # -------------------------------------- Investors (BP).Op Call: Organizational Costs outside commitment
            dst_active_sheet[f'CA{dst_row_num}'] = dst_row['other_expenses_op_ccy']                                 # -------------------------------------- Investors (BP).Op Call: Other Expenses outside commitment
            dst_active_sheet[f'CB{dst_row_num}'] = dst_row['partnership_expenses_op_ccy']                           # -------------------------------------- Investors (BP).Op Call: Partnership Expenses outside commitment
            dst_active_sheet[f'CC{dst_row_num}'] = dst_row['redraw_op_ccy']                                         # -------------------------------------- Investors (BP).Op Redrawable amount
            dst_active_sheet[f'CM{dst_row_num}'] = dst_row['roc_fund_ccy']                                          # -------------------------------------- Investors (BP).Share Dist: Return of Capital
            dst_active_sheet[f'CN{dst_row_num}'] = dst_row['capital_gains_fund_ccy']                                # -------------------------------------- Investors (BP).Share Dist: Realized Gain / (Loss)
            dst_active_sheet[f'CO{dst_row_num}'] = dst_row['dividend_fund_ccy']                                     # -------------------------------------- Investors (BP).Share Dist: Dividends
            dst_active_sheet[f'CP{dst_row_num}'] = dst_row['interest_fund_ccy']                                     # -------------------------------------- Investors (BP).Share Dist: Interests
            dst_active_sheet[f'CQ{dst_row_num}'] = dst_row['other_income_fund_ccy']                                 # -------------------------------------- Investors (BP).Share Dist: Other Income
            dst_active_sheet[f'CR{dst_row_num}'] = dst_row['withholding_tax_fund_ccy']                              # -------------------------------------- Investors (BP).Share Dist: Withholding Tax
            dst_active_sheet[f'CS{dst_row_num}'] = dst_row['carry_fund_ccy']                                        # -------------------------------------- Investors (BP).Share Dist: Carry
            dst_active_sheet[f'CT{dst_row_num}'] = dst_row['sub_close_interest_dist_fund_ccy']                      # -------------------------------------- Investors (BP).Share Dist: Subsequent Close Interest          
            dst_active_sheet[f'DC{dst_row_num}'] = dst_row['roc_investor_ccy']                                      # -------------------------------------- Investors (BP).Investor Dist: Return of Capital
            dst_active_sheet[f'DD{dst_row_num}'] = dst_row['capital_gains_investor_ccy']                            # -------------------------------------- Investors (BP).Investor Dist: Realized Gain / (Loss)
            dst_active_sheet[f'DE{dst_row_num}'] = dst_row['dividend_investor_ccy']                                 # -------------------------------------- Investors (BP).Investor Dist: Dividends
            dst_active_sheet[f'DF{dst_row_num}'] = dst_row['interest_investor_ccy']                                 # -------------------------------------- Investors (BP).Investor Dist: Interests
            dst_active_sheet[f'DG{dst_row_num}'] = dst_row['other_income_investor_ccy']                             # -------------------------------------- Investors (BP).Investor Dist: Other Income
            dst_active_sheet[f'DH{dst_row_num}'] = dst_row['withholding_tax_investor_ccy']                          # -------------------------------------- Investors (BP).Investor Dist: Withholding Tax
            dst_active_sheet[f'DI{dst_row_num}'] = dst_row['carry_investor_ccy']                                    # -------------------------------------- Investors (BP).Investor Dist: Carry
            dst_active_sheet[f'DJ{dst_row_num}'] = dst_row['sub_close_interest_dist_investor_ccy']                  # -------------------------------------- Investors (BP).Investor Dist: Subsequent Close Interest 
            dst_active_sheet[f'DK{dst_row_num}'] = dst_row['roc_op_ccy']                                            # -------------------------------------- Investors (BP).Op Dist: Return of Capital
            dst_active_sheet[f'DL{dst_row_num}'] = dst_row['capital_gains_op_ccy']                                  # -------------------------------------- Investors (BP).Op Dist: Realized Gain / (Loss)
            dst_active_sheet[f'DM{dst_row_num}'] = dst_row['dividend_op_ccy']                                       # -------------------------------------- Investors (BP).Op Dist: Dividends
            dst_active_sheet[f'DN{dst_row_num}'] = dst_row['interest_op_ccy']                                       # -------------------------------------- Investors (BP).Op Dist: Interests
            dst_active_sheet[f'DO{dst_row_num}'] = dst_row['other_income_op_ccy']                                   # -------------------------------------- Investors (BP).Op Dist: Other Income
            dst_active_sheet[f'DP{dst_row_num}'] = dst_row['withholding_tax_op_ccy']                                # -------------------------------------- Investors (BP).Op Dist: Withholding Tax
            dst_active_sheet[f'DS{dst_row_num}'] = dst_row['fair_value_fund_ccy']                                   # -------------------------------------- Investors (BP).Share Valuation
            dst_active_sheet[f'DU{dst_row_num}'] = dst_row['fair_value_investor_ccy']                               # -------------------------------------- Investors (BP).Investor Valuation
            dst_active_sheet[f'DV{dst_row_num}'] = dst_row['fair_value_op_ccy']                                     # -------------------------------------- Investors (BP).Op Valuation
            dst_active_sheet[f'ED{dst_row_num}'] = dst_row['sub_close_interest_call_fund_ccy']                      # -------------------------------------- Investors (BP).Share Call: Subsequent Close Interest outside commitment
            dst_active_sheet[f'EF{dst_row_num}'] = dst_row['sub_close_interest_call_investor_ccy']                  # -------------------------------------- Investors (BP).Investor Call: Subsequent Close Interest outside commitment
            dst_active_sheet[f'EG{dst_row_num}'] = dst_row['sub_close_interest_call_op_ccy']                        # -------------------------------------- Investors (BP).Op Call: Subsequent Close Interest outside commitment
            dst_active_sheet[f'EK{dst_row_num}'] = dst_row['sub_close_interest_call_op_ccy_inside_commitment']      # -------------------------------------- Investors (BP).Op Call: Subsequent Close Interest
            dst_active_sheet[f'EN{dst_row_num}'] = dst_row['fair_value_fund_ccy']                                   # -------------------------------------- Investors (BP).Share Estimate
            #dst_active_sheet[f'EO{dst_row_num}'] = dst_row['fair_value_investor_ccy']                               # -------------------------------------- Investors (BP).Fund Estimate 
            dst_active_sheet[f'EP{dst_row_num}'] = dst_row['fair_value_investor_ccy']                               # -------------------------------------- Investors (BP).Investor Estimate
            dst_active_sheet[f'EQ{dst_row_num}'] = dst_row['fair_value_op_ccy']                                     # -------------------------------------- Investors (BP).Op Estimate
            dst_active_sheet[f'ER{dst_row_num}'] = dst_row['description']                                           # -------------------------------------- Comments
            dst_active_sheet[f'ES{dst_row_num}'] = dst_row['legal_fees_fund_ccy_inside_commitment']                 # -------------------------------------- Investors (BP).Share Call: Legal Fees in commitment
            dst_active_sheet[f'ET{dst_row_num}'] = dst_row['legal_fees_investor_ccy_inside_commitment']             # -------------------------------------- Investors (BP).Investor Call: Legal Fees in commitment
            dst_active_sheet[f'EU{dst_row_num}'] = dst_row['org_costs_fund_ccy_inside_commitment']                  # -------------------------------------- Investors (BP).Share Call: Organizational Costs in commitment
            dst_active_sheet[f'EV{dst_row_num}'] = dst_row['org_costs_investor_ccy_inside_commitment']              # -------------------------------------- Investors (BP).Investor Call: Organizational Costs in commitment
            dst_active_sheet[f'EW{dst_row_num}'] = dst_row['partnership_expenses_fund_ccy_inside_commitment']       # -------------------------------------- Investors (BP).Share Call: Partnership Expenses in commitment
            dst_active_sheet[f'EX{dst_row_num}'] = dst_row['partnership_expenses_investor_ccy_inside_commitment']   # -------------------------------------- Investors (BP).Investor Call: Partnership Expenses in commitment
            dst_active_sheet[f'EY{dst_row_num}'] = dst_row['other_expenses_fund_ccy_inside_commitment']             # -------------------------------------- Investors (BP).Share Call: Other Expenses in commitment
            dst_active_sheet[f'EZ{dst_row_num}'] = dst_row['other_expenses_investor_ccy_inside_commitment']         # -------------------------------------- Investors (BP).Investor Call: Other Expenses in commitment
            dst_active_sheet[f'FA{dst_row_num}'] = dst_row['working_capital_fund_ccy_inside_commitment']            # -------------------------------------- Investors (BP).Share Call: Working Capital in commitment
            dst_active_sheet[f'FB{dst_row_num}'] = dst_row['working_capital_investor_ccy_inside_commitment']        # -------------------------------------- Investors (BP).Investor Call: Working Capital in commitment
            dst_active_sheet[f'FC{dst_row_num}'] = dst_row['locked']                                                # -------------------------------------- Locked
            dst_active_sheet[f'FD{dst_row_num}'] = dst_row['fx_rate']                                               # -------------------------------------- Investors (BP).o/i FX Rate
            #dst_active_sheet[f'FE{dst_row_num}'] = dst_row['impairment_fund_ccy']                                   # -------------------------------------- Investors (BP).Fund Impairment
            dst_active_sheet[f'FF{dst_row_num}'] = dst_row['impairment_investor_ccy']                               # -------------------------------------- Investors (BP).Investor Impairment
            dst_active_sheet[f'FG{dst_row_num}'] = dst_row['impairment_op_ccy']                                     # -------------------------------------- Investors (BP).Op Impairment
            dst_active_sheet[f'FH{dst_row_num}'] = dst_row['impairment_fund_ccy']                                   # -------------------------------------- Investors (BP).Share Impairment
        else:
            # Insert the values in the destination row 
            dst_active_sheet[f'C{dst_row_num}'] = dst_row['fund_op_type']                   # -------------------------------------- Type    
            dst_active_sheet[f'D{dst_row_num}'] = f"{dst_row['fund_name']} - LR"            # -------------------------------------- Fund.Fund 
            dst_active_sheet[f'E{dst_row_num}'] = dst_row['investor']                       # -------------------------------------- Investors (BP).Subscriber.Investor.Investor
            #dst_active_sheet[f'F{dst_row_num}'] = dst_row['investor']                      # -------------------------------------- Investors (BP).Subscriber.Investor.Investor
            #dst_active_sheet[f'G{dst_row_num}'] = dst_row['fund']                          # -------------------------------------- Investors (BP).Subscriber.Main vehicle.Fund.Fund
            dst_active_sheet[f'H{dst_row_num}'] = dst_row['date']                           # -------------------------------------- Date
            dst_active_sheet[f'I{dst_row_num}'] = dst_row['index']                       # -------------------------------------- Index
            dst_active_sheet[f'K{dst_row_num}'] = dst_row['fund_op_code']                   # -------------------------------------- Operation
            if isinstance(dst_row['desc'],float):
                pass
            elif len(dst_row['desc']) < 254:
                dst_active_sheet[f'L{dst_row_num}'] = dst_row['desc']                # -------------------------------------- Reason                                  
                dst_active_sheet[f'M{dst_row_num}'] = dst_row['desc']                # -------------------------------------- Comments
            else:
                logging.warning(f"{dst_row['fund']} description not inserted because it was too long:\n\t{dst_row['desc']}")
            dst_active_sheet[f'N{dst_row_num}'] = 'Shared'                                  # -------------------------------------- Region
            dst_active_sheet[f'O{dst_row_num}'] = dst_row['settlement_date']                # -------------------------------------- Settlement date
            #dst_active_sheet[f'P{dst_row_num}'] = Send date 
            #dst_active_sheet[f'Q{dst_row_num}'] =  
            dst_active_sheet[f'R{dst_row_num}'] = dst_row['op_currency']                    # -------------------------------------- Operation currency 
            dst_active_sheet[f'S{dst_row_num}'] = 'Commitment'                              # -------------------------------------- Share/Series
            dst_active_sheet[f'T{dst_row_num}'] = 'No'                                      # -------------------------------------- Draft
            #dst_active_sheet[f'U{dst_row_num}'] = ''                                       # -------------------------------------- Bank account
            dst_active_sheet[f'V{dst_row_num}'] = 'No'                                      # -------------------------------------- Cancelled
            dst_active_sheet[f'W{dst_row_num}'] = 'Commitment'                              # -------------------------------------- Investors (BP).Fund share.Category
            # dst_active_sheet[f'X{dst_row_num}'] = Investors (BP).Shares committed
            dst_active_sheet[f'X{dst_row_num}'] = 1 #dst_row['index']                       # -------------------------------------- Investors (BP).Fund share.Share Index
            dst_active_sheet[f'Y{dst_row_num}'] = dst_row['shares_issued']                  # -------------------------------------- Investors (BP).Shares committed              
            dst_active_sheet[f'Z{dst_row_num}'] = dst_row['no_shares']                      # -------------------------------------- Investors (BP).Nb Shares
            dst_active_sheet[f'AA{dst_row_num}'] = dst_row['commitment_investor_ccy']       # -------------------------------------- Investors (BP).Share Commitment
            # dst_active_sheet[f'AB{dst_row_num}'] = src_row['commitment']                  # -------------------------------------- Investors (BP).FX Rate -> Investor
            # dst_active_sheet[f'AC{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AD{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AE{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AF{dst_row_num}'] = src_row['commitment']
            # dst_active_sheet[f'AI{dst_row_num}'] = src_row['commitment']
            #dst_active_sheet[f'AJ{dst_row_num}'] = dst_row['investments_fund_ccy']          # -------------------------------------- Investors (BP).Share Call: Investments
            #dst_active_sheet[f'AK{dst_row_num}'] = dst_row['investments_investor_ccy']      # -------------------------------------- Investors (BP).Investor Call: Investments
            #dst_active_sheet[f'AL{dst_row_num}'] = src_row_euros['investments']            # -------------------------------------- Share.Allocation : Fund Call: Legal Fees outside commitment
            #dst_active_sheet[f'AR{dst_row_num}'] = dst_row['fees_fund_ccy']                # -------------------------------------- Investors (BP).Share Call: Management Fees in commitment
            #dst_active_sheet[f'AS{dst_row_num}'] = dst_row['fees_investor_ccy']            # -------------------------------------- Investors (BP).Investor Call: Management Fees in commitment
            #dst_active_sheet[f'CL{dst_row_num}'] = dst_row['roc_investor_ccy']              # -------------------------------------- Investors (BP).Share Dist: Return of Capital
            #dst_active_sheet[f'CM{dst_row_num}'] = dst_row['roc_investor_ccy']             # -------------------------------------- Investors (BP).Investor Dist: Return of Capital
            #dst_active_sheet[f'CN{dst_row_num}'] = dst_row['capital_gains_fund_ccy']        # -------------------------------------- Investors (BP).Share Dist: Realized Gain / (Loss)            
            #dst_active_sheet[f'DC{dst_row_num}'] = dst_row['roc_investor_ccy']             # -------------------------------------- Investors (BP).Investor Dist: Return of Capital
            #dst_active_sheet[f'DD{dst_row_num}'] = dst_row['capital_gains_investor_ccy']   # -------------------------------------- Investors (BP).Investor Dist: Realized Gain / (Loss)
            dst_active_sheet[f'FG{dst_row_num}'] = dst_row['nominal']                     # -------------------------------------- Investors (BP).Share Call: Nominal
            dst_active_sheet[f'FI{dst_row_num}'] = dst_row['premium']                       # -------------------------------------- Investors (BP).Share Call: Premium
            dst_active_sheet[f'FK{dst_row_num}'] = 0                                        # -------------------------------------- Investors (BP).Share Dist: Return of Nominal
            if dst_row['return_of_premium'] is not None:
                dst_active_sheet[f'FM{dst_row_num}'] = round(dst_row['return_of_premium'],2)    # -------------------------------------- Investors (BP).Share Dist: Return of Premium
            if dst_row['fair_value_investor_ccy'] is not None:
                dst_active_sheet[f'FO{dst_row_num}'] = round(dst_row['fair_value_investor_ccy'],2) # -------------------------------------- Investors (BP).Share Valuation pre-carried
    except KeyError:
        logger.exception(f'Key Error:')
        logger.exception(dst_row)
    return dst_row_num + 1

def compile_data(src_file, file_name,fx_rates=None,env='UAT'):
    """
        Inputs:
            src_file: the path to the source LR report containing the fund ops we need to summarise  
            file_name: contains the share class information
            fx_rates: dataframe containing fx rates
            env: determines which versions of import files we use
    """

    # Define the dataframe that will store our destination rows:
    dst_rows = pd.DataFrame({'fund_op_type': pd.Series([], dtype='str')
                            ,'fund_name': pd.Series([], dtype='str')
                            ,'investor': pd.Series([], dtype='str')
                            ,'fund': pd.Series([], dtype='str')
                            ,'date': pd.Series([], dtype='str')
                            ,'fund_op_code': pd.Series([], dtype='str')
                            ,'description': pd.Series([], dtype='str')
                            ,'settlement_date': pd.Series([], dtype='str')
                            ,'op_currency': pd.Series([], dtype='str')
                            ,'commitment_fund_ccy': pd.Series([], dtype='float')
                            ,'commitment_investor_ccy': pd.Series([], dtype='float')
                            ,'investments_fund_ccy': pd.Series([], dtype='float')
                            ,'investments_investor_ccy': pd.Series([], dtype='float')
                            ,'fees_fund_ccy': pd.Series([], dtype='float')
                            ,'fees_investor_ccy': pd.Series([], dtype='float')
                            ,'roc_fund_ccy':  pd.Series([], dtype='float')
                            ,'roc_investor_ccy': pd.Series([], dtype='float')
                            ,'capital_gains_fund_ccy': pd.Series([], dtype='float')  
                            ,'capital_gains_investor_ccy': pd.Series([], dtype='float')
                            ,'fair_value_fund_ccy': pd.Series([], dtype='float') 
                            ,'fair_value_investor_ccy': pd.Series([], dtype='float') 
                            ,'undrawn_fund_ccy': pd.Series([], dtype='float') 
                            ,'undrawn_investor_ccy': pd.Series([], dtype='float') 
                            })

    # Find the sheets we care about and store them in a dictionary    
    logger.info('...Extracting Data')
    print('...Extracting Data')

    # Read in the investors in each share and assign the best matching name to each fund in dst_rows 
    print(file_name)
    if file_name in ['AGp Shares','AHp Shares','AHs Shares','AIp Shares','AIs Shares']:
        file_name = file_name.replace('p ',' ')
        file_name = file_name.replace('s ',' ')

    share = str.split(file_name)[0]
    sheet_count = 1 # Count the sheets

    fund_details, fund_sheets, fund_sheets_euros, all_fund_sheets = extract_funds(src_file,share,env=env)



    # loop through our funds
    logger.info('...Working On:')
    print('...Working On:')

    for fund_name, sheet_names in all_fund_sheets.items():

        # Find/decide the values we're going to insert for the commitment row
        fund_commitment = fund_details[fund_details['Name'] == fund_name]['Commitment'].max()
        fund_commitment_euros = fund_details[fund_details['Name'] == fund_name]['Commitment_Euros'].max()
        fund_vintage = fund_details[fund_details['Name'] == fund_name]['Vintage'].iloc[0]
        if type(fund_details[(fund_details['Currency'] != 'EUR') & (fund_details['Name'] == fund_name)]['Currency'].max()) == float:           
            fund_ccy = fund_details[(fund_details['Name'] == fund_name)]['Currency'].max()
        else:
            fund_ccy = fund_details[(fund_details['Currency'] != 'EUR') & (fund_details['Name'] == fund_name)]['Currency'].max()    


        if np.isnan(fund_commitment):
            fund_commitment = fund_commitment_euros

        src_row = {
            'commitment': fund_commitment,
            'commitment_euros': fund_commitment_euros,
            'date': fund_vintage,
            'description': 'Commitment',
            'investments': None,
            'return of capital': None,
            'capital gains': None,
            'fair value': None,
            'fees': None
        }

        # Insert total fund commitment row  
        dst_rows = append_row(fund_name                                                   
                             ,fund_ccy
                             ,file_name                                     
                             ,src_row    
                             ,src_row       
                             ,dst_rows
                             ,fx_rates=fx_rates
                             ,investee_fund=True)

        #logger.info(dst_rows.iloc[-1])                                                                                                                                        

        # Each fund might have data across multiple sheets, so loop through the sheets
        for sheet_name in sheet_names:

            if fund_name in fund_sheets and '€' in sheet_name:
                # fund_sheets only contains funds that are not in the investor currency, so there should be 2 sheets for this fund
                # We want to skip the investor currency sheet because we'll insert this data when we run into the fund currency counterpart
                pass 
            else:

                # Print the fund to monitor the progress
                logger.info(f'\t{sheet_count}. Fund: {fund_name}, Sheet: {sheet_name}')
                print(f'\t{sheet_count:02d}. Fund: {fund_name}, Sheet: {sheet_name}')
                sheet_count = sheet_count + 1

                # Read in the rest of the data as a dataframe
                xl = pd.ExcelFile(src_file)     
                src_rows = xl.parse(sheet_name,skiprows=9,usecols='A:J')
                src_rows.dropna(subset = ['Date'], inplace=True)
                src_rows['Commitment'] = 0
                src_rows['commitment_euros'] = None

                # Format dataframe 
                src_rows.columns = src_rows.columns.str.lower()
                src_rows['description'] = src_rows['description'].fillna(' ')
                src_rows_undrawn = src_rows['undrawn']
                src_rows = src_rows.fillna(0)
                src_rows['undrawn'] = src_rows_undrawn

                # If there fund currency isn't the investor currency, extract the amounts from the investor currency sheet 
                src_rows_euros = pd.DataFrame() 
                if fund_ccy != 'EUR' and '€' not in sheet_name:
                    sheet_name_euros_1 = sheet_name + ' €'
                    sheet_name_euros_2 = sheet_name + ' (€)'
                    sheet_name_euros_3 = sheet_name + ' € '
                    sheet_name_euros_4 = sheet_name + ' (€) '
                    try:
                        if sheet_name_euros_1 in fund_sheets_euros[fund_name]:
                            src_rows_euros = xl.parse(sheet_name_euros_1,skiprows=9,usecols='A:J')
                            src_rows_euros.dropna(subset = ['Date'], inplace=True)
                            src_rows_euros.columns = src_rows_euros.columns.str.lower()
                            sheet_name_euros = sheet_name_euros_1
                        elif sheet_name_euros_2 in fund_sheets_euros[fund_name]:
                            src_rows_euros = xl.parse(sheet_name_euros_2,skiprows=9,usecols='A:J')
                            src_rows_euros.dropna(subset = ['Date'], inplace=True)
                            src_rows_euros.columns = src_rows_euros.columns.str.lower()
                            sheet_name_euros = sheet_name_euros_2
                        elif sheet_name_euros_3 in fund_sheets_euros[fund_name]:
                            src_rows_euros = xl.parse(sheet_name_euros_3,skiprows=9,usecols='A:J')
                            src_rows_euros.dropna(subset = ['Date'], inplace=True)
                            src_rows_euros.columns = src_rows_euros.columns.str.lower()
                            sheet_name_euros = sheet_name_euros_3
                        elif sheet_name_euros_4 in fund_sheets_euros[fund_name]:
                            src_rows_euros = xl.parse(sheet_name_euros_4,skiprows=9,usecols='A:J')
                            src_rows_euros.dropna(subset = ['Date'], inplace=True)
                            src_rows_euros.columns = src_rows_euros.columns.str.lower()
                            sheet_name_euros = sheet_name_euros_4
                    except KeyError:
                        logger.warning(f'\t\t{sheet_name} is in {fund_ccy} but there is no equivalent sheet in €!')

                # Check if two sheets exist for the same fund that they both contain the same number of fund operations. 
                if not src_rows_euros.empty:
                    if len(src_rows_euros) != len(src_rows):
                        print(f'\t\tThe number of fund operations in {sheet_name} does not match {sheet_name_euros}!')
                        logger.warning(f'\t\tThe number of fund operations in {sheet_name} does not match {sheet_name_euros}!')
                else:
                    # If src_rows_euros was empty it means our fund is in euros, so make src_rows_euros a copy of src_rows 
                    src_rows_euros = src_rows.copy()

                # Set fair value to 0 where it's negative as requested in 'Other' Fund Operations file sent by LR. The two exceptions have been fixed in the source files. 
                src_rows.loc[src_rows['fair value'] < 0,'fair value'] = 0
                src_rows_euros.loc[src_rows_euros['fair value'] < 0,'fair value'] = 0

                # Fill out the latest accounting nav column
                src_rows['latest accounting nav'] = src_rows['fair value'].shift(1)
                src_rows_euros['latest accounting nav'] = src_rows_euros['fair value'].shift(1)

                # Map the source columns to the destination columns 
                for index, src_row in src_rows.iterrows():
                    # put this in a try-except block in case fund ccy != investor ccy and the euros sheet has fewer rows  
                    try:
                        src_row_euros = src_rows_euros.loc[index]
                    except IndexError:
                        logger.warning(f'\tRow {index} ignored...')
                    except KeyError:
                        logger.warning(f'\tRow {index} ignored...')
                    dst_rows = append_row(fund_name                                                   
                                         ,fund_ccy
                                         ,file_name                                     
                                         ,src_row    
                                         ,src_row_euros       
                                         ,dst_rows
                                         ,fx_rates=fx_rates
                                         ,investee_fund=True)

    # Take care of missing dates 
    null_dates = dst_rows[dst_rows['fund_op_type']=='IF: Commitment']['date'].isna()
    null_date_funds = dst_rows[dst_rows['fund_op_type']=='IF: Commitment'][null_dates]['fund_name']
    for fund in null_date_funds:
        ndf_rows = dst_rows.loc[dst_rows['fund_name']==fund]
        alt_date = ndf_rows[ndf_rows['fund_op_type'].isin(['IF: Call','IF: Mixed operation'])]['date'].min()
        dst_rows.loc[(dst_rows['fund_op_type']=='IF: Commitment')&(dst_rows['fund_name']==fund),'date'] = alt_date
        dst_rows.loc[(dst_rows['fund_op_type']=='IF: Commitment')&(dst_rows['fund_name']==fund),'settlement_date'] = alt_date

    int_date_funds = dst_rows[dst_rows['date'].astype(str).str.isdigit()]['fund_name']
    for fund in int_date_funds:
        idf_rows = dst_rows.loc[dst_rows['fund_name']==fund]
        alt_date = idf_rows[idf_rows['fund_op_type'].isin(['IF: Call','IF: Mixed operation'])]['date'].min()
        dst_rows.loc[(dst_rows['fund_op_type']=='IF: Commitment')&(dst_rows['fund_name']==fund),'date'] = alt_date
        dst_rows.loc[(dst_rows['fund_op_type']=='IF: Commitment')&(dst_rows['fund_name']==fund),'settlement_date'] = alt_date

    # Replace None with NaN so we can aggregate
    dst_rows.fillna(value=np.nan,inplace=True) 

    ################################################################################################################################################################################################################
    # Categorise each fee and capital gain based on the ones that Jenny + I were able to categorise 
    # descriptions_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Data Migration/Investee Fund Operation Compilation Files/Compilation - All Fees, Return of Capital and Capital Gains.xlsx'
    # desc_xl = pd.ExcelFile(descriptions_file)
    # for sheet in desc_xl.book.worksheets:
    #     if sheet.title == 'Fees':
    #         desc_fees = desc_xl.parse(sheet.title)
    #     # All capital gains within the return of capital sheet are classified as Realised Gain/Loss so we don't need to read that sheet in.
    #     elif sheet.title == 'Capital Gains':
    #         desc_cap_gains = desc_xl.parse(sheet.title)
    # dst_rows = categorise_fees_cap_gains(dst_rows,desc_fees,desc_cap_gains)
    ################################################################################################################################################################################################################

    # Find the rows where we weren't able to figure out which amounts were in/outside commitment or recallable and categorise them
    dst_rows['date'] = dst_rows['date'].apply(lambda x: dt.datetime(x.year,x.month,x.day))
    
    dst_rows = integrate_misc_categorisation_files(dst_rows)

    # Get the descriptions for categorising fees and capital gains as returned by LR, and add rows to dst_rows where we're splitting up an amount
    dst_rows = integrate_LR_compilation_file(dst_rows)
    

    # Add index column (number the fund ops of the same type happening on the same day)
    ind = dst_rows.groupby(['fund'
                            #,'investor'
                            #,'fund_name'
                            #,'op_currency'
                            #,'fund_op_type' # eFront for some reason updates instead of inserting if there is a call and return of call on the same date with index = 1 
                            #,'fund_op_code'
                            ,'date']).cumcount()
    ind = ind.apply(lambda x: x + 1)
    dst_rows['index'] = ind

    # Add Locked column
    dst_rows['locked'] = 'FALSE'
    #dst_rows.loc[dst_rows['date'] >= dt.datetime(2021, 1, 1), 'locked'] = 'FALSE'

    # Add op_ccy cols
    dst_rows = add_op_cols(dst_rows)

    # Not sure why but some empty rows are there at the end sometimes - e.g. AB Shares
    dst_rows = dst_rows[dst_rows.fund_op_type.notnull()]

    # Round everything to 2 d.p. Some of the values will already be rounded to 2 d.p. in the 'append' function
    fx_rate = dst_rows['fx_rate'] # We don't want to round fx rates, so store that 
    dst_rows.fillna(0,inplace=True) # If a column contains null values then the round function doesn't work, so lets fill them with 0s first 
    dst_rows = dst_rows.round(2) # Round the rows to 2 d.p.
    dst_rows.replace(0,np.nan,inplace=True) # Replace 0s with nulls again. Important for valuations, so we don't accidentally enter a valuation of 0
    dst_rows['fx_rate'] = fx_rate # replace the fx_rates column to the pre-rounded figures 
  

    # Only for testing: Check if latest fund op is Accounting NAV. if it is, insert Official NAV 
    for fund in list(dst_rows['fund_name'].unique()):
        dst_rows_fund = dst_rows[dst_rows['fund_name'] == fund]
        if dst_rows_fund['fund_op_type'].iloc[-1] == 'IF: Accounting Valuation': 
            dst_row = dst_rows_fund.iloc[-1].copy()
            dst_row['fund_op_type'] = 'IF: Official NAV'
            dst_row['index'] = dst_row['index'] + 10 # Accounting NAVs will be o
            dst_rows = dst_rows.append(dst_row,ignore_index=True)

    # Sort the rows
    dst_rows.sort_values(by=['fund_name','date','fund_op_type','index'],inplace=True,ignore_index=True) 

    return dst_rows

def categorise_fees_cap_gains(dst_rows, desc_fees, desc_cap_gains):
    """ 
        Inputs:
            dst_rows: the processed data containing the fees we want to categorise
            desc_fees: all the labelled descriptions for all fees
            desc_cap_gains: all the labelled descriptions for all capital gains  
        Classify the fees  
    """
    # Create new fee columns 
    dst_rows['legal_fees_fund_ccy'] = None
    dst_rows['legal_fees_investor_ccy'] = None
    dst_rows['org_costs_fund_ccy'] = None
    dst_rows['org_costs_investor_ccy'] = None
    dst_rows['partnership_expenses_fund_ccy'] = None
    dst_rows['partnership_expenses_investor_ccy'] = None
    dst_rows['other_expenses_fund_ccy'] = None
    dst_rows['other_expenses_investor_ccy'] = None
    dst_rows['working_capital_fund_ccy'] = None
    dst_rows['working_capital_investor_ccy'] = None
    dst_rows['sub_close_interest_call_fund_ccy'] = None
    dst_rows['sub_close_interest_call_investor_ccy'] = None
    dst_rows['legal_fees_fund_ccy_inside_commitment'] = None
    dst_rows['legal_fees_investor_ccy_inside_commitment'] = None
    dst_rows['org_costs_fund_ccy_inside_commitment'] = None
    dst_rows['org_costs_investor_ccy_inside_commitment'] = None
    dst_rows['partnership_expenses_fund_ccy_inside_commitment'] = None
    dst_rows['partnership_expenses_investor_ccy_inside_commitment'] = None
    dst_rows['other_expenses_fund_ccy_inside_commitment'] = None
    dst_rows['other_expenses_investor_ccy_inside_commitment'] = None
    dst_rows['working_capital_fund_ccy_inside_commitment'] = None
    dst_rows['working_capital_investor_ccy_inside_commitment'] = None
    dst_rows['sub_close_interest_call_fund_ccy_inside_commitment'] = None
    dst_rows['sub_close_interest_call_investor_ccy_inside_commitment'] = None
    # Create new distribution columns 
    dst_rows['dividend_fund_ccy'] = None
    dst_rows['dividend_investor_ccy'] = None
    dst_rows['interest_fund_ccy'] = None
    dst_rows['interest_investor_ccy'] = None
    dst_rows['other_income_fund_ccy'] = None
    dst_rows['other_income_investor_ccy'] = None
    dst_rows['realised_gain_fund_ccy'] = None
    dst_rows['realised_gain_investor_ccy'] = None
    dst_rows['withholding_tax_fund_ccy'] = None
    dst_rows['withholding_tax_investor_ccy'] = None
    dst_rows['carry_fund_ccy'] = None
    dst_rows['carry_investor_ccy'] = None
    dst_rows['sub_close_interest_dist_fund_ccy'] = None
    dst_rows['sub_close_interest_dist_investor_ccy'] = None
    # Extract the fees and capital gains rows from dst_rows
    dst_rows_fees = dst_rows[dst_rows['fees_fund_ccy'] > 0]
    dst_rows_cap_gains = dst_rows[dst_rows['capital_gains_fund_ccy'] > 0]
    # Sort through all the fees and make sure it's categorised appropriately 
    for dst_row in dst_rows_fees.itertuples():
        if dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Legal Fees')]['Description'].to_list():
            # Legal Fees outside commitment
            dst_rows.loc[dst_row.Index,'legal_fees_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'legal_fees_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Legal Fees inside commitment
            dst_rows.loc[dst_row.Index,'legal_fees_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'legal_fees_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
        elif dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Working Capital')]['Description'].to_list():
            # Working Capital outside commitment
            dst_rows.loc[dst_row.Index,'working_capital_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'working_capital_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Working Capital inside commitment
            dst_rows.loc[dst_row.Index,'working_capital_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'working_capital_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
        elif dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Subsequent Close Interest')]['Description'].to_list():
            # Subsequent Close Interest outside commitment
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Subsequent Close Interest inside commitment
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
        elif dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Partnership Expenses')]['Description'].to_list():
            # Partnership Expenses outside commitment
            dst_rows.loc[dst_row.Index,'partnership_expenses_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'partnership_expenses_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Partnership Expenses inside commitment
            dst_rows.loc[dst_row.Index,'partnership_expenses_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'partnership_expenses_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
        elif dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Organizational Costs')]['Description'].to_list():
            # Organizational Costs outside commitment
            dst_rows.loc[dst_row.Index,'org_costs_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'org_costs_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Organizational Costs inside commitment
            dst_rows.loc[dst_row.Index,'org_costs_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'org_costs_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
        elif dst_row.description in desc_fees[(desc_fees['Unnamed: 22'] == dst_row.investor)&(desc_fees['Mapping'] == 'Other Expenses')]['Description'].to_list():
            # Other Expenses outside commitment
            dst_rows.loc[dst_row.Index,'other_expenses_fund_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'other_expenses_investor_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy'] = 0
            # Other Expenses inside commitment
            dst_rows.loc[dst_row.Index,'other_expenses_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'other_expenses_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
            dst_rows.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
    # Sort through all the capital gains and make sure it's categorised appropriately 
    for dst_row in dst_rows_cap_gains.itertuples():
        if dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Interests')]['Description'].to_list():
            # Interests
            dst_rows.loc[dst_row.Index,'interest_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'interest_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        # elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Realised Gain/Loss')]['Description'].to_list():
        #     # Realised Gain/Loss
        #     dst_rows.loc[dst_row.Index,'realised_gain_fund_ccy'] = dst_row.capital_gains_fund_ccy
        #     dst_rows.loc[dst_row.Index,'realised_gain_investor_ccy'] = dst_row.capital_gains_investor_ccy
        #     dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
        #     dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Carry')]['Description'].to_list():
            # Realised Gain/Loss
            dst_rows.loc[dst_row.Index,'carry_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'carry_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Dividends')]['Description'].to_list():
            # Dividends
            dst_rows.loc[dst_row.Index,'dividend_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'dividend_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Other Income')]['Description'].to_list():
            # Other Income
            dst_rows.loc[dst_row.Index,'other_income_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'other_income_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Subsequent Close Interest')]['Description'].to_list():
            # Subsequent Close Interest
            dst_rows.loc[dst_row.Index,'sub_close_interest_dist_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'sub_close_interest_dist_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
        elif dst_row.description in desc_cap_gains[(desc_cap_gains['Unnamed: 12'] == dst_row.investor)&(desc_cap_gains['Mapping'] == 'Withholding Tax')]['Description'].to_list():
            # Withholding Tax
            dst_rows.loc[dst_row.Index,'withholding_tax_fund_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'withholding_tax_investor_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
            dst_rows.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0

    return dst_rows

def log_uncategorised_fund_ops(fund_name, file_name, src_row):
    """
        Inputs:
            src_row: the source row that we can't categorise
        Insert the src row in a separate file for reference
    """
    # Location of the file 
    other_fund_operations_src = 'C:/Users/RajContractor/Documents/Python Files/Dev/LR Migration/Migrated/Other Fund Operations Src.xlsx'

    # Read in what's there and add the new row 
    src_row['fund'] = fund_name 
    src_row['share'] = file_name 
    try:
        other_fund_ops_src = pd.read_excel(other_fund_operations_src, index_col=None)
        other_fund_ops_src = other_fund_ops_src.append(src_row, ignore_index=True)
    except FileNotFoundError:
        other_fund_ops_src = pd.DataFrame()
        other_fund_ops_src = other_fund_ops_src.append(src_row, ignore_index=True)

    # Output everything to excel
    other_fund_ops_src.to_excel(other_fund_operations_src,index=False, columns=['share','fund','date','description','investments', 'fees','return of capital','capital gains','fair value'])

def separate_other_fund_ops(dst_row):
    """
        Inputs:
            dst_row: the processed row that would have been included in the import file 
        Insert the dst row in a separate file so we can categorise and import it later
    """
    # Location of the file 
    other_fund_operations_dst = 'C:/Users/RajContractor/Documents/Python Files/Dev/LR Migration/Migrated/Other Fund Operations.xlsx'
    
    # Read in what's there and add the new row 
    try:
        other_fund_ops_dst = pd.read_excel(other_fund_operations_dst, index_col=None)
        other_fund_ops_dst = other_fund_ops_dst.append(dst_row, ignore_index=True)
    except FileNotFoundError:
        other_fund_ops_dst = pd.DataFrame()
        other_fund_ops_dst = other_fund_ops_dst.append(dst_row, ignore_index=True)

    # Output everything to excel
    other_fund_ops_dst.to_excel(other_fund_operations_dst,index=False)

def log_fx_mismatch(file_name,fund_op_type, fund,src_row,src_row_euros,fund_ccy,fx_rates,fx_investments,fx_fees,fx_return_of_capital,fx_capital_gains,fx_fair_value):
    """
        Inputs:
            file_name: same as the share class 
            fund_op_type: the type of fund operation containing the fx rate mismatch
            fund: the fund containing the above fund operation
            src_row: the source data row that contains the fx rate mismatch
            src_row_euros: as above but contains the euros amounts
            fund_ccy: fund currency
            fx_rates: dataframe containing fx rates 
        Insert the source row in a separate file so we can send to LR
    """
    # Location of the file 
    fx_rate_mismatch_doc = 'C:/Users/RajContractor/Documents/Python Files/Dev/LR Migration/Migrated/FX_Rate_Mismatch.xlsx'

    # Convert to dataframe
    fx_rates = fx_rates[(fx_rates['src_ccy']==fund_ccy)&(fx_rates['dst_ccy']=='EUR')&(fx_rates['date']==src_row['date'])]['rate'].values
    if len(fx_rates) != 1:
        if src_row['date'] <= dt.datetime(2021,3,31):
            print(f"\t\tNo FX Rate imported for {src_row['date']}")
    elif src_row['date'] <= dt.datetime(2021,3,31):
        fx_rate = fx_rates[0]
        fx_rate = round(fx_rate,4)
        if fx_investments is not None:
            fx_investments = round(fx_investments,4)
        if fx_fees is not None:
            fx_fees = round(fx_fees,4)
        if fx_return_of_capital is not None:
            fx_return_of_capital = round(fx_return_of_capital,4)
        if fx_capital_gains is not None:
            fx_capital_gains = round(fx_capital_gains,4)
        if fx_fair_value is not None:
            fx_fair_value = round(fx_fair_value,4)

        row = pd.DataFrame({'File':[file_name]
                        ,'Fund':[fund]
                        ,'Date':[src_row['date']]
                        ,'Description':[src_row['description']]
                        ,'Expected Fx Rate':[fx_rate] 
                        ,'Investment Fund CCY':[src_row['investments']]
                        ,'Investment Investor CCY':[src_row_euros['investments']]
                        ,'Investment Fx Rate':[fx_investments]
                        ,'Investment Fx Matches': ['Yes' if fx_investments == fx_rate else 'No']
                        ,'Fee Fund CCY':[src_row['fees']]
                        ,'Fee Investor CCY':[src_row_euros['fees']]
                        ,'Fee Fx Rate':[fx_fees]
                        ,'Fee Fx Matches': ['Yes' if fx_fees == fx_rate else 'No']
                        ,'RoC Fund CCY':[src_row['return of capital']]
                        ,'RoC Investor CCY':[src_row_euros['return of capital']]
                        ,'RoC Fx Rate':[fx_return_of_capital]
                        ,'RoC Fx Matches': ['Yes' if fx_return_of_capital == fx_rate else 'No']
                        ,'CG Fund CCY':[src_row['capital gains']]
                        ,'CG Investor CCY':[src_row_euros['capital gains']]
                        ,'CG Fx Rate':[fx_capital_gains]
                        ,'CG Fx Matches': ['Yes' if fx_capital_gains == fx_rate else 'No']
                        ,'Fair Value Fund CCY':[src_row['fair value']]
                        ,'Fair Value Investor CCY':[src_row_euros['fair value']]
                        ,'Fair Value Fx Rate':[fx_fair_value]
                        ,'Fair Value Fx Matches': ['Yes' if fx_fair_value == fx_rate else 'No']
                        })
        try:
            fx_rate_mismatch = pd.read_excel(fx_rate_mismatch_doc, index_col=None)
            fx_rate_mismatch = fx_rate_mismatch.append(row, ignore_index=True)
        except FileNotFoundError:
            fx_rate_mismatch = pd.DataFrame()
            fx_rate_mismatch = fx_rate_mismatch.append(row, ignore_index=True)

        # Output everything to excel
        fx_rate_mismatch.to_excel(fx_rate_mismatch_doc,index=False)

def replace_fund_name(share,fund_name_report):
    """
        Hardcode what the fund name should be for certain funds 
        Inputs:
            share - name of the share, e.g. 'AH' 
            fund_name_report - this is what's in the 'Name' field when a fund sheet is found in the LR report 
        Output:
            fund_name_compare - this is the name we should use in our comparison with the official Fund import file 
    """
    # Based on LR's emails, AachenMünchener Lebensversicherung AG was renamed to Generali Deutschland Lebensversicherung AG
    fund_name_report = fund_name_report.replace('AachenMünchener Lebensversicherung AG', 'Generali Deutschland Lebensversicherung AG')
    fund_name_report = fund_name_report.replace('AachenMunchener Lebensversicherung AG', 'Generali Deutschland Lebensversicherung AG')

    if share == 'AH':
        fund_name_report = fund_name_report.replace('NVP I Co-Invest SCSp', 'Novalpina I Co-Invest')
       # fund_name_report = fund_name_report.replace('NB Renaissance Partners III', 'Neuberger Berman Renaissance Partners III - Fund')
        fund_name_report = fund_name_report.replace('Stirling Square Capital Partners Fund IV - Coinv', 'Stirling Square Capital Partners Fund IV - Coinvestment')
        fund_name_report = fund_name_report.replace('PASF IV', 'Portfolio Advisor Seconday Fund IV')
    elif share == 'AI':
        fund_name_report = fund_name_report.replace('Novacap TMT VI Coinvestment', 'Novacap TMT V Co-Investment (Logibec)')
        fund_name_report = fund_name_report.replace('Antin Infrastructure Partners Co-Investment', 'Gauss Co-Invest')
    elif share == 'N':
        fund_name_report = fund_name_report.replace('Black River Food Fund 2 LP', 'ITV - Blackriver Food Fund 2')
        fund_name_report = fund_name_report.replace('Permira I 4b', 'Permira Europe I L.P. 4B')
        fund_name_report = fund_name_report.replace('Permira I 3', 'Permira Europe I L.P. 3 / L.P. 4')
        fund_name_report = fund_name_report.replace('Advent B', 'Advent Euro-Italian Direct Investment Program L.P. - Class B')
        fund_name_report = fund_name_report.replace('Permira II', 'Permira Europe II L.P. 2')
        fund_name_report = fund_name_report.replace('MEIF', 'Macquarie European Infrastructure Fund L.P.')
        fund_name_report = fund_name_report.replace('GSIP I', 'GS Infrastructure Partners I L.P.')
        fund_name_report = fund_name_report.replace('GSMP V', 'GS Mezzanine Partners V Offshore LP')
        fund_name_report = fund_name_report.replace('REI', 'Renewable Energy Investments')
    elif share == 'O':
        fund_name_report = fund_name_report.replace('Rhone III', 'Rhône Offshore Partners III L.P.')
        fund_name_report = fund_name_report.replace('Rhone II', 'Rhône Offshore Partners II L.P.')
        fund_name_report = fund_name_report.replace('21 CP III', '21 Centrale Partners III')
        fund_name_report = fund_name_report.replace('Equinox II', 'Equinox Two S.C.A')
        fund_name_report = fund_name_report.replace('CHF II', 'China Harvest Fund II L.P.')
        fund_name_report = fund_name_report.replace('RIH', 'Renewable Investments Holding')
        fund_name_report = fund_name_report.replace('21 CP IV', '21 Central Partners IV')
        fund_name_report = fund_name_report.replace('Lehman Brothers II', 'Lehman Brothers Offshore Investment Partners II L.P.')
    elif share in ['AA','AB','AC','AD','AE']:
        fund_name_report = fund_name_report.replace('Coinv', 'Co-inv')
        fund_name_report = fund_name_report.replace('coinv', 'co-inv')
    elif share == 'AF':
        fund_name_report = fund_name_report.replace('ACON Equity Partners IV - Coinvestment', 'ACON - Coinvestment')
    elif share == 'W':
        fund_name_report = fund_name_report.replace('JAB Consumer Fund II (follow on)', 'JAB Consumer Fund - Global Consumer Brands')





    return fund_name_report

######################################################################################
#                                 Investee Fund Ops                                  #
######################################################################################
def migrate_investee_data(src_file, dst_file, file_name,fx_rates, env):
    """
        Inputs:
            src_file: the path to the source LR report containing the fund ops we need to migrate 
            dst_file: the path to the output file. This is the eFront import file which will be used to migrate the fund ops from the src_file.  
            file_name: what our source file is called. We need this to determine the fund we're working with.  
            fx_rates: dataframe containing fx rates
            env: 'DEV' or 'UAT'. Determines which version of import files we use. 
        Migrate investee fund ops.
    """
    # Define the template 
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/N2. Fund Operations - Test Template.xlsx'
    wb = oxl.load_workbook(template_file)
    dst_active_sheet = wb.active
    dst_row_num = 5 # Ignore headers in the template 
    
    # Read in the 'other' fund operations already classified by LR. We're doing it here rather than in the function where this is needed so it's only read in once.
    other_fund_op_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/Other Fund Operations.xlsx'
    other_fund_ops = pd.read_excel(other_fund_op_file, index_col=None)
    other_fund_ops['fair_value'] = other_fund_ops['Fair Value']

    # Create our dst_rows containing all the fund ops 
    dst_rows = compile_data(src_file,file_name,fx_rates,env=env)

    # Insert
    for i, dst_row in dst_rows.iterrows():
        dst_row_num = insert_row(dst_row                                                        # pass in the row we want to insert                                                
                                ,dst_row_num                                                    # this is just the current row count 
                                ,dst_active_sheet                                               # the active sheet we're editing
                                )

    wb.save(dst_file)

def add_op_cols(dst_rows):
    """add the Op ccy columns to dst_rows"""
    dst_rows['commitment_op_ccy'] = None
    dst_rows['investments_op_ccy'] = None
    dst_rows['legal_fees_op_ccy'] = None
    dst_rows['fees_op_ccy_inside_commitment'] = None
    dst_rows['fees_op_ccy'] = None
    dst_rows['org_costs_op_ccy'] = None
    dst_rows['other_expenses_op_ccy'] = None
    dst_rows['partnership_expenses_op_ccy'] = None
    dst_rows['redraw_op_ccy'] = None

    dst_rows['roc_op_ccy'] = None
    dst_rows['capital_gains_op_ccy'] = None
    dst_rows['dividend_op_ccy'] = None
    dst_rows['interest_op_ccy'] = None
    dst_rows['other_income_op_ccy'] = None
    dst_rows['withholding_tax_op_ccy'] = None
    dst_rows['fair_value_op_ccy'] = None
    dst_rows['sub_close_interest_call_op_ccy'] = None
    dst_rows['sub_close_interest_call_op_ccy_inside_commitment'] = None

    dst_rows['impairment_op_ccy'] = None

    # For stirling add the op ccy amounts
    for dst_row in dst_rows.itertuples():
        if (dst_row.investor == 'AH Shares - LR') and (dst_row.fund_name == 'Stirling Square Capital Partners Fund IV - Coinvestment'):
            dst_rows.loc[dst_row.Index,'commitment_op_ccy'] = dst_row.commitment_fund_ccy
            dst_rows.loc[dst_row.Index,'investments_op_ccy'] = dst_row.investments_fund_ccy
            dst_rows.loc[dst_row.Index,'legal_fees_op_ccy'] = dst_row.legal_fees_fund_ccy
            dst_rows.loc[dst_row.Index,'fees_op_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_op_ccy'] = dst_row.fees_fund_ccy
            dst_rows.loc[dst_row.Index,'org_costs_op_ccy'] = dst_row.org_costs_fund_ccy
            dst_rows.loc[dst_row.Index,'other_expenses_op_ccy'] = dst_row.other_expenses_fund_ccy
            dst_rows.loc[dst_row.Index,'partnership_expenses_op_ccy'] = dst_row.partnership_expenses_fund_ccy
            dst_rows.loc[dst_row.Index,'redraw_op_ccy'] = dst_row.redraw_fund_ccy 

            dst_rows.loc[dst_row.Index,'roc_op_ccy'] = dst_row.org_costs_fund_ccy
            dst_rows.loc[dst_row.Index,'capital_gains_op_ccy'] = dst_row.capital_gains_fund_ccy
            dst_rows.loc[dst_row.Index,'dividend_op_ccy'] = dst_row.dividend_fund_ccy
            dst_rows.loc[dst_row.Index,'interest_op_ccy'] = dst_row.interest_fund_ccy
            dst_rows.loc[dst_row.Index,'other_income_op_ccy'] = dst_row.other_income_fund_ccy
            dst_rows.loc[dst_row.Index,'withholding_tax_op_ccy'] = dst_row.withholding_tax_fund_ccy
            dst_rows.loc[dst_row.Index,'fair_value_op_ccy'] = dst_row.fair_value_fund_ccy
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_op_ccy'] = dst_row.sub_close_interest_call_fund_ccy
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_op_ccy_inside_commitment'] = dst_row.sub_close_interest_call_fund_ccy_inside_commitment
            
            dst_rows.loc[dst_row.Index,'impairment_op_ccy'] = dst_row.impairment_fund_ccy

            dst_rows.loc[dst_row.Index,'commitment_fund_ccy'] = dst_row.commitment_investor_ccy
            dst_rows.loc[dst_row.Index,'investments_fund_ccy'] = dst_row.investments_investor_ccy
            dst_rows.loc[dst_row.Index,'legal_fees_fund_ccy'] = dst_row.legal_fees_investor_ccy
            dst_rows.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
            dst_rows.loc[dst_row.Index,'fees_fund_ccy'] = dst_row.fees_investor_ccy
            dst_rows.loc[dst_row.Index,'org_costs_fund_ccy'] = dst_row.org_costs_investor_ccy
            dst_rows.loc[dst_row.Index,'other_expenses_fund_ccy'] = dst_row.other_expenses_investor_ccy
            dst_rows.loc[dst_row.Index,'partnership_expenses_fund_ccy'] = dst_row.partnership_expenses_investor_ccy
            dst_rows.loc[dst_row.Index,'redraw_fund_ccy'] = dst_row.redraw_investor_ccy

            dst_rows.loc[dst_row.Index,'roc_fund_ccy'] = dst_row.roc_investor_ccy 
            dst_rows.loc[dst_row.Index,'capital_gains_fund_ccy'] = dst_row.capital_gains_investor_ccy
            dst_rows.loc[dst_row.Index,'dividend_fund_ccy'] = dst_row.dividend_investor_ccy
            dst_rows.loc[dst_row.Index,'interest_fund_ccy'] = dst_row.interest_investor_ccy
            dst_rows.loc[dst_row.Index,'other_income_fund_ccy'] = dst_row.other_income_investor_ccy
            dst_rows.loc[dst_row.Index,'withholding_tax_fund_ccy'] = dst_row.withholding_tax_investor_ccy
            dst_rows.loc[dst_row.Index,'fair_value_fund_ccy'] = dst_row.fair_value_investor_ccy 
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_fund_ccy'] = dst_row.sub_close_interest_call_investor_ccy
            dst_rows.loc[dst_row.Index,'sub_close_interest_call_fund_ccy_inside_commitment'] = dst_row.sub_close_interest_call_investor_ccy_inside_commitment

            dst_rows.loc[dst_row.Index,'impairment_fund_ccy'] = dst_row.impairment_investor_ccy

    return dst_rows

def integrate_misc_categorisation_files(dst_rows):
    # Create a copy of dst_rows - this is the one we will edit 
    dst_rows_copy = dst_rows.copy()

    # Read in the data we will use to categorise all our fees/cap gains
    comp_file_lr = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Data Migration/Investee Fund Operation Compilation Files/Returned by LR/Copy of Fees and Return of Capital Categorisation.xlsx'
    comp_fund_ops = pd.read_excel(comp_file_lr, index_col=None)

    # Find the share we're interested in (i.e. the share of our source data) and replace s/p
    share_src = dst_rows['investor'].unique()[0]

    # This next bit shouldn't be necessary but I'm afraid to remove it 
    if share_src in ['AGp Shares - LR','AHp Shares - LR','AHs Shares - LR','AIp Shares - LR','AIs Shares - LR']:
        share_src = share_src.replace('s S',' S')
        share_src = share_src.replace('p S',' S')

    # Only keep the rows we're interested in
    comp_fund_ops = comp_fund_ops[(comp_fund_ops['Share'] == share_src)]

    # Formate date as datetime
    comp_fund_ops['Date'] = comp_fund_ops['Date'].apply(lambda x: dt.datetime(x.year,x.month,x.day))
    comp_fund_ops['Investments'] = comp_fund_ops['Investments'].astype('float64')
    comp_fund_ops['Fees'] = comp_fund_ops['Fees'].astype('float64')
    comp_fund_ops['Return of Capital'] = comp_fund_ops['Return of Capital'].astype('float64')
    comp_fund_ops['Capital Gains'] = comp_fund_ops['Capital Gains'].astype('float64')

    if share_src == 'O Shares - LR':
        # One bespoke change they seem to want
        dst_rows.loc[(dst_rows['fund_name']=='Equinox Two S.C.A')&(dst_rows['date']==dt.datetime(2015,1,8)),'roc_fund_ccy'] = 202622.58
        dst_rows.loc[(dst_rows['fund_name']=='Equinox Two S.C.A')&(dst_rows['date']==dt.datetime(2015,1,8)),'roc_investor_ccy'] = 202622.58
        dst_rows.loc[(dst_rows['fund_name']=='Equinox Two S.C.A')&(dst_rows['date']==dt.datetime(2015,1,8)),'capital_gains_fund_ccy'] = 0.00
        dst_rows.loc[(dst_rows['fund_name']=='Equinox Two S.C.A')&(dst_rows['date']==dt.datetime(2015,1,8)),'capital_gains_investor_ccy'] = 0.00
        dst_rows.loc[(dst_rows['fund_name']=='Equinox Two S.C.A')&(dst_rows['date']==dt.datetime(2015,1,8)),'capital_gains_fund_ccy']
        dst_rows_copy.loc[(dst_rows_copy['fund_name']=='Equinox Two S.C.A')&(dst_rows_copy['date']==dt.datetime(2015,1,8)),'roc_fund_ccy'] = 202622.58
        dst_rows_copy.loc[(dst_rows_copy['fund_name']=='Equinox Two S.C.A')&(dst_rows_copy['date']==dt.datetime(2015,1,8)),'roc_investor_ccy'] = 202622.58
        dst_rows_copy.loc[(dst_rows_copy['fund_name']=='Equinox Two S.C.A')&(dst_rows_copy['date']==dt.datetime(2015,1,8)),'capital_gains_fund_ccy'] = 0.00
        dst_rows_copy.loc[(dst_rows_copy['fund_name']=='Equinox Two S.C.A')&(dst_rows_copy['date']==dt.datetime(2015,1,8)),'capital_gains_investor_ccy'] = 0.00

    for i, fund_op in comp_fund_ops.iterrows():
        fund = fund_op['Fund']

        # Find the relevant row    
        relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fund_op['Date'])&(round(dst_rows['investments_fund_ccy'],2) == round(fund_op['Investments'],2))&(round(dst_rows['roc_fund_ccy'],2) == round(fund_op['Return of Capital'],2))&(round(dst_rows['capital_gains_fund_ccy'],2) == round(fund_op['Capital Gains'],2))&(round(dst_rows['fair_value_fund_ccy'],2)==round(fund_op['Fair Value'],2))&(round(dst_rows['orig_fee'],2) == round(fund_op['Fees'],2))].copy()
        if len(relevant_row) != 1:    
            # We didn't manage to find the relevant row so throw an error 
            print(f"\t\tWarning: row not found for fund op - {share_src} - {fund} - {fund_op['Date']} - {fund_op['Description']} - {fund_op['Investments']} - {fund_op['Fees']} - {fund_op['Return of Capital']} - {fund_op['Capital Gains']}")
            logging.warning(f"\t\tWarning: row not found for fund op - {share_src} - {fund} - {fund_op['Date']} - {fund_op['Description']} - {fund_op['Investments']} - {fund_op['Fees']} - {fund_op['Return of Capital']} - {fund_op['Capital Gains']}")
        else:
            # Find the index of the row we need to update 
            ind = relevant_row.index.to_list()[0]
            # Fee
            if fund_op['Fees'] > 0 and fund_op['AddedInd'] in [0,1]:
                dst_rows_copy.loc[ind,'fees_fund_ccy_inside_commitment'] = fund_op['Fee inside commitment']
                dst_rows_copy.loc[ind,'fees_investor_ccy_inside_commitment'] = round((relevant_row.loc[ind,'fx_rate']*fund_op['Fee inside commitment']),2)
                dst_rows_copy.loc[ind,'fees_fund_ccy'] = fund_op['Fee outside commitment']
                dst_rows_copy.loc[ind,'fees_investor_ccy'] = round((relevant_row.loc[ind,'fx_rate']*fund_op['Fee outside commitment']),2)
            elif fund_op['AddedInd'] > 1:
                # Update the description to what they've provided
                relevant_row['description'] = fund_op['Description']

                # The fee has been split - this is not the first row, so create a copy of the relevant row and set all non-fee values to 0
                relevant_row['investments_fund_ccy'] = 0
                relevant_row['investments_investor_ccy'] = 0
                relevant_row['roc_fund_ccy'] = 0
                relevant_row['roc_investor_ccy'] = 0
                relevant_row['capital_gains_fund_ccy'] = 0
                relevant_row['capital_gains_investor_ccy'] = 0
                if fund_op['Fee outside commitment'] + fund_op['Fee inside commitment'] > 0:
                    relevant_row['fund_op_type'] = 'IF: Call'
                    relevant_row['fund_op_code'] = 'CC'
                else:
                    relevant_row['fund_op_type'] = 'IF: Return Of Call'
                    relevant_row['fund_op_code'] = 'CD'

                relevant_row['fees_fund_ccy'] = fund_op['Fee outside commitment']
                relevant_row['fees_investor_ccy'] = round((relevant_row.loc[ind,'fx_rate']*fund_op['Fee outside commitment']),2)
                relevant_row['fees_fund_ccy_inside_commitment'] = fund_op['Fee inside commitment']
                relevant_row['fees_investor_ccy_inside_commitment'] = round((relevant_row.loc[ind,'fx_rate']*fund_op['Fee inside commitment']),2)

                if fund_op['AddedInd'] > 100:
                    # The original fee has been split out into it's own fund operation, so we need to null any fees on the original row
                    dst_rows_copy.loc[ind,'fees_fund_ccy_inside_commitment'] = 0
                    dst_rows_copy.loc[ind,'fees_investor_ccy_inside_commitment'] = 0
                    dst_rows_copy.loc[ind,'fees_fund_ccy'] = 0
                    dst_rows_copy.loc[ind,'fees_investor_ccy'] = 0
                dst_rows_copy = pd.concat([dst_rows_copy, relevant_row], ignore_index=True)
            # Redraw
            redraw_fund_ccy = fund_op['return of capital redrawable amount'] + fund_op['capital gain redrawable amount']
            if redraw_fund_ccy > 0:
                dst_rows_copy.loc[ind,'redraw_fund_ccy'] = redraw_fund_ccy
                dst_rows_copy.loc[ind,'redraw_investor_ccy'] = round((relevant_row.loc[ind,'fx_rate']*redraw_fund_ccy),2)

    return dst_rows_copy

def integrate_LR_compilation_file(dst_rows):
    """ 
        Inputs:
            dst_rows: the processed data containing the fees we want to categorise
        Get the categorised fees/capital gains data, add rows where one amount has been split in two and classify the fees/capital gains in our data  
    """
    # Read in Compilation data
    comp_file_lr = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Data Migration/Investee Fund Operation Compilation Files/Returned by LR/Compilation_LR_Clean.xlsx'
    comp_fees = pd.read_excel(comp_file_lr, index_col=None, sheet_name='Fees')
    # All rows with non-zero capital gains from return of capital/fees sheet should have been copied to the Capital Gains sheet, so we don't need to read in the return of capital sheet
    comp_cap_gains = pd.read_excel(comp_file_lr, index_col=None, sheet_name='Capital Gains')

    # Make a copy of dst_rows, because we'll need a place to store our updated amounts
    dst_rows_copy = dst_rows.copy()
    dst_rows_copy['mapping'] = None

    # Find the share we're interested in (i.e. the share of our source data) and replace s/p
    share_src = dst_rows['investor'].unique()[0]
    share_src = share_src.replace(' - LR','')
    if share_src in ['AGp Shares','AHp Shares','AHs Shares','AIp Shares','AIs Shares']:
        share_src = share_src.replace('s ',' ')
        share_src = share_src.replace('p ',' ')

    # Only keep the rows we're interested in
    comp_fees_share = comp_fees[comp_fees['Share'] == share_src]
    comp_fees_share = comp_fees_share[(comp_fees_share.Mapping != '?')&(comp_fees_share.Mapping == comp_fees_share.Mapping)]
    comp_cap_gains_share = comp_cap_gains[comp_cap_gains['Share'] == share_src]
    comp_cap_gains_share = comp_cap_gains_share[(comp_cap_gains_share.Mapping != '?')&(comp_cap_gains_share.Mapping == comp_cap_gains_share.Mapping)]

    # Convert our date column to datetime in case it isn't already
    comp_fees_share['Date'] = comp_fees_share['Date'].apply(lambda x: dt.datetime(x.year,x.month,x.day))
    comp_cap_gains_share['Date'] = comp_cap_gains_share['Date'].apply(lambda x: dt.datetime(x.year,x.month,x.day))

    # Transfer all our capital gains mappings to dst_rows_copy
    for row in comp_cap_gains_share.itertuples():
        # Add "- LR" next to the fund name if in N and O shares 
        fund = row.investee_fund
        alt_fund = fund + ' - LR'

        # Find the relevant row
        if fund in dst_rows['fund_name'].to_list():
            # Some shares are split into multiple files, and this row might correspond to a fund in the other file
            relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == row.Date)&(dst_rows['description'] == row.Description)&(round(dst_rows['capital_gains_fund_ccy'],2) == round(row.cg_amount,2))].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # If N/O shares we might need to add - LR at the end of the fund
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == row.Date)&(dst_rows['description'] == row.Description)&(round(dst_rows['capital_gains_fund_ccy'],2) == round(row.cg_amount,2))].copy()
            if len(relevant_row) != 1:
                # maybe the capital gains amount has changed 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == row.Date)&(dst_rows['description'] == row.Description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == row.Date)&(dst_rows['description'] == row.Description)].copy()
            if len(relevant_row) != 1:
                # maybe the description has changed 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == row.Date)&(round(dst_rows['capital_gains_fund_ccy'],2) == round(row.cg_amount,2))].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == row.Date)&(round(dst_rows['capital_gains_fund_ccy'],2) == round(row.cg_amount,2))].copy()
            if len(relevant_row) != 1:
                # maybe the description and amount has changed 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == row.Date)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == row.Date)].copy()
            if len(relevant_row) != 1:
                print(f"\t\tWarning: row not found for distributed amount - {row.Share} - {fund} - {row.Date} - {row.Description} - {row.roc_amount} - {row.cg_amount}")
                logging.warning(f"\t\tWarning: row not found for distributed amount - {row.Share} - {fund} - {row.Date} - {row.Description} - {row.roc_amount} - {row.cg_amount}")
            else:
                if row.Mapping == 'Subsequent Close Interest': 
                    mapping = 'SCI - Dist'
                else:
                    mapping = row.Mapping
                if row.SplitInd == 1:
                    # First row - update description and amount if applicable
                    dst_rows_copy.loc[relevant_row.index,'capital_gains_fund_ccy'] = row.cg_amount
                    dst_rows_copy.loc[relevant_row.index,'capital_gains_investor_ccy'] = round((row.cg_amount * dst_rows_copy.loc[relevant_row.index,'fx_rate']),2)
                    dst_rows_copy.loc[relevant_row.index,'description'] = row.Description
                    dst_rows_copy.loc[relevant_row.index,'mapping'] = mapping
                elif row.SplitInd > 1:
                    # New fund op added 
                    relevant_row['description'] = row.Description
                    # The distribution has been split - this is not the first row, so create a copy of the relevant row and set all non-distribution values to 0
                    relevant_row['investments_fund_ccy'] = 0
                    relevant_row['investments_investor_ccy'] = 0
                    relevant_row['fees_fund_ccy'] = 0
                    relevant_row['fees_investor_ccy'] = 0
                    relevant_row['fees_fund_ccy_inside_commitment'] = 0
                    relevant_row['fees_investor_ccy_inside_commitment'] = 0
                    relevant_row['roc_fund_ccy'] = 0
                    relevant_row['roc_investor_ccy'] = 0
                    relevant_row['capital_gains_fund_ccy'] = row.cg_amount
                    relevant_row['capital_gains_investor_ccy'] = round((row.cg_amount * dst_rows_copy.loc[row.Index,'fx_rate']),2)
                    
                    # Update the fund op type as it may no longer be applicable 
                    relevant_row['fund_op_type'] = 'IF: Distribution'
                    relevant_row['fund_op_code'] = 'CD'
                    relevant_row['mapping'] = mapping
                    dst_rows_copy = pd.concat([dst_rows_copy, relevant_row], ignore_index=True)
                else:
                    dst_rows_copy.loc[relevant_row.index,'description'] = row.Description
                    dst_rows_copy.loc[relevant_row.index,'mapping'] = mapping
                
    # Before we try to match the fees we need to strip the descriptions of any leading/trailing spaces and replace spaces with _
    # https://stackoverflow.com/questions/69513863/identical-strings-dont-match-in-python-an-issue-with-spaces-leading-trailing          
    dst_rows['old_description'] = dst_rows['description'].str.strip().apply(lambda x: '_'.join(x.split()))
    comp_fees_share['old_description'] = comp_fees_share['old_description'].str.strip().apply(lambda x: '_'.join(x.split()))

    # Transfer all our fee mappings to dst_rows_copy and split amounts/rows as appropriate 
    for fee in comp_fees_share.itertuples():
        fund = fee.investee_fund
        alt_fund = fund + ' - LR'

        if fund in dst_rows['fund_name'].to_list():
            # Find the relevant row in our data
            relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) == round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) == round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount is inside commitment
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) == round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) == round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount has changed
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) != round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) != round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount is inside commitment and has changed 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) != round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) != round(fee.fee_amount,2))&(dst_rows['old_description'] == fee.old_description)].copy()
            if len(relevant_row) != 1:
                # maybe the description has changed
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) == round(fee.fee_amount,2))].copy() 
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) == round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount is inside commitment and the description has changed
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) == round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) == round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount has changed and the description has changed
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) != round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy'],2) != round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1:
                # maybe the fee amount is inside commitment and has changed and the description has changed 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) != round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares']:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(round(dst_rows['fees_fund_ccy_inside_commitment'],2) != round(fee.fee_amount,2))].copy()
            if len(relevant_row) != 1 and fee.SplitInd != 0:
                # the amount is not going to match because this row has been split 
                relevant_row = dst_rows[(dst_rows['fund_name'] == fund)&(dst_rows['date'] == fee.Date)&(dst_rows['old_description'].values[0] == fee.old_description)].copy()
            if len(relevant_row) != 1 and share_src in ['N Shares','O Shares'] and fee.SplitInd != 0:
                # As above but look for a fund with - LR at the end 
                relevant_row = dst_rows[(dst_rows['fund_name'] == alt_fund)&(dst_rows['date'] == fee.Date)&(dst_rows['old_description'].values[0] == fee.old_description)].copy()
            if len(relevant_row) != 1:
                print(f"\t\tWarning: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
                logging.warning(f"\t\tWarning: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
            else:
                # Get the index of the row in our source data
                i = relevant_row.index.to_list()[0]

                # Update the amounts 
                if fee.SplitInd in [1,0,-1]:
                    # The fee might have been split, in which case this is the first row
                    # Update the description in case it has been changed
                    dst_rows_copy.loc[i,'description'] = fee.Description
                    # Update the mapping
                    dst_rows_copy.loc[i,'mapping'] = fee.Mapping
                    # Update the amount
                    if relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] != 0 and relevant_row.loc[i,'fees_fund_ccy'] == 0:
                        # Find the ratio of the old amount to the new amount
                        ratio = fee.fee_amount/relevant_row.loc[i,'fees_fund_ccy_inside_commitment']
                        # Find the corresponding amount in the investor currency and apply the ratio to that
                        investor_ccy_amount = ratio * relevant_row.loc[i,'fees_investor_ccy_inside_commitment']
                        # Update the old amount
                        dst_rows_copy.loc[i,'fees_fund_ccy_inside_commitment'] = round(fee.fee_amount,2)
                        dst_rows_copy.loc[i,'fees_investor_ccy_inside_commitment'] = round(investor_ccy_amount,2)
                    elif relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] == 0 and relevant_row.loc[i,'fees_fund_ccy'] != 0:
                        # Find the ratio of the old amount to the new amount
                        ratio = fee.fee_amount/relevant_row.loc[i,'fees_fund_ccy']
                        # Find the corresponding amount in the investor currency and apply the ratio to that
                        investor_ccy_amount = ratio * relevant_row.loc[i,'fees_investor_ccy']
                        # Update the old amount
                        dst_rows_copy.loc[i,'fees_fund_ccy'] = round(fee.fee_amount,2)
                        dst_rows_copy.loc[i,'fees_investor_ccy'] = round(investor_ccy_amount,2)
                    elif round(fee.fee_amount,2) == round((relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] + relevant_row.loc[i,'fees_fund_ccy']),2):
                        # They just want to separate the fee inside commitment and fee outside commitment. We have already done that so ignore this row 
                        pass 
                    elif relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] == 0 and relevant_row.loc[i,'fees_fund_ccy'] == 0:
                        # No fee. This is not necessarily an issue. This should only happen if we split the fee in the integrate_misc_categorisation_files function
                        print(f"\t\tFee amounts 0: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")
                        logging.info(f"\t\tFee amounts 0: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")
                    else:
                        # Any rows in our compilation file corresponding to mixed fees should have SplitInd >100, so throw an error if we end up here
                        print(relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment'])
                        raise NameError('Unexpected Data Found')
                    
                    if fee.SplitInd == -1:
                        # Zero the return of capital and capital gain amount because it has been split into another fund op
                        dst_rows_copy.loc[i,'roc_fund_ccy'] = 0
                        dst_rows_copy.loc[i,'roc_investor_ccy'] = 0 
                        dst_rows_copy.loc[i,'capital_gains_fund_ccy'] = 0 
                        dst_rows_copy.loc[i,'capital_gains_investor_ccy'] = 0 

                        # # update the fund op
                        # if fee.fee_amount > 0:
                        #     dst_rows_copy.loc[i,'fund_op_type'] = 'IF: Call'
                        #     dst_rows_copy.loc[i,'fund_op_code'] = 'CC'
                        # elif fee.fee_amount < 0:
                        #     dst_rows_copy.loc[i,'fund_op_type'] = 'IF: Return Of Call'
                        #     dst_rows_copy.loc[i,'fund_op_code'] = 'CD'

                    # update the fund op
                    if relevant_row.loc[i,'fund_op_type'] != 'IF: Mixed operation':
                        if round(fee.fee_amount,2) + round(relevant_row.loc[i,'investments_fund_ccy'],2) > 0 or fee.Mapping == 'Subsequent Close Interest':
                            dst_rows_copy.loc[i,'fund_op_type'] = 'IF: Call'
                            dst_rows_copy.loc[i,'fund_op_code'] = 'CC'
                        elif fee.fee_amount < 0:
                            dst_rows_copy.loc[i,'fund_op_type'] = 'IF: Return Of Call'
                            dst_rows_copy.loc[i,'fund_op_code'] = 'CD'
                    
                elif fee.SplitInd > 1 and fee.SplitInd <= 100:
                    # Update the description to what they've provided
                    relevant_row['description'] = fee.Description

                    # Add the mapping
                    relevant_row['mapping'] = fee.Mapping

                    # The fee has been split - this is not the first row, so create a copy of the relevant row and set all non-fee values to 0
                    relevant_row['investments_fund_ccy'] = 0
                    relevant_row['investments_investor_ccy'] = 0
                    relevant_row['roc_fund_ccy'] = 0
                    relevant_row['roc_investor_ccy'] = 0
                    relevant_row['capital_gains_fund_ccy'] = 0
                    relevant_row['capital_gains_investor_ccy'] = 0
                    
                    # Update the fund op type as it may no longer be applicable 
                    if fee.fee_amount > 0 or fee.Mapping == 'Subsequent Close Interest':
                        relevant_row['fund_op_type'] = 'IF: Call'
                        relevant_row['fund_op_code'] = 'CC'
                    elif fee.fee_amount < 0 and fee.Mapping != 'Subsequent Close Interest':
                        relevant_row['fund_op_type'] = 'IF: Return Of Call'
                        relevant_row['fund_op_code'] = 'CD'
                    
                    # The fee has been split - this is not the first row, so we need to update the amounts on our copy of the relevant row so we can
                    if relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] != 0 and relevant_row.loc[i,'fees_fund_ccy'] == 0:
                        # Find the ratio of the old amount to the new amount
                        ratio = fee.fee_amount/relevant_row.loc[i,'fees_fund_ccy_inside_commitment']
                        # Find the corresponding amount in the investor currency and apply the ratio to that
                        investor_ccy_amount = ratio * relevant_row.loc[i,'fees_investor_ccy_inside_commitment']
                        # Update the old amount
                        relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] = round(fee.fee_amount,2)
                        relevant_row.loc[i,'fees_investor_ccy_inside_commitment'] = round(investor_ccy_amount,2)
                    elif relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] == 0 and relevant_row.loc[i,'fees_fund_ccy'] != 0:
                        # Find the ratio of the old amount to the new amount
                        ratio = fee.fee_amount/relevant_row.loc[i,'fees_fund_ccy']
                        # Find the corresponding amount in the investor currency and apply the ratio to that
                        investor_ccy_amount = ratio * relevant_row.loc[i,'fees_investor_ccy']
                        # Update the old amount
                        relevant_row.loc[i,'fees_fund_ccy'] = round(fee.fee_amount,2)
                        relevant_row.loc[i,'fees_investor_ccy'] = round(investor_ccy_amount,2)
                    else:
                        # We dont expect any rows in our compilation file corresponding to mixed fees, so throw an error if we end up here
                        print(f"\t\tWarning >1: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
                        print(f"\t\tWarning >1: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")
                        logging.warning(f"\t\tWarning >1: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
                        logging.warning(f"\t\tWarning >1: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")
                    
                    # Add the new row to dst_rows_copy
                    dst_rows_copy = pd.concat([dst_rows_copy, relevant_row], ignore_index=True)            
                elif fee.SplitInd < -1:
                    # Update the description to what they've provided
                    relevant_row['description'] = fee.Description
                    # The fee has been split - this is not the first row, so create a copy of the relevant row and set all non-fee values to 0
                    relevant_row['investments_fund_ccy'] = 0
                    relevant_row['investments_investor_ccy'] = 0
                    relevant_row['fees_fund_ccy'] = 0
                    relevant_row['fees_investor_ccy'] = 0
                    relevant_row['fees_fund_ccy_inside_commitment'] = 0
                    relevant_row['fees_investor_ccy_inside_commitment'] = 0
                    
                    # Update the fund op type as it may no longer be applicable 
                    relevant_row['fund_op_type'] = 'IF: Distribution'
                    relevant_row['fund_op_code'] = 'CD'
                    relevant_row['mapping'] = fee.Mapping
                    dst_rows_copy = pd.concat([dst_rows_copy, relevant_row], ignore_index=True) 
                elif fee.SplitInd > 100:
                    # This is a mixed fee that has been split - part of it is inside commitment and part of it is outside commitment.   
                    # Update the description to what they've provided
                    relevant_row['description'] = fee.Description

                    # Update the mapping
                    relevant_row['mapping'] = fee.Mapping                  

                    # We need to zero the fee on the relevant row in our data because we've split it into it's own separate fund op, away from any investment
                    dst_rows_copy.loc[i,'fees_fund_ccy'] = 0
                    dst_rows_copy.loc[i,'fees_fund_ccy_inside_commitment'] = 0
                    dst_rows_copy.loc[i,'fees_investor_ccy'] = 0
                    dst_rows_copy.loc[i,'fees_investor_ccy_inside_commitment'] = 0

                    # The fee has been split - this is not the first row, so create a copy of the relevant row and set all non-fee values to 0
                    relevant_row['investments_fund_ccy'] = 0
                    relevant_row['investments_investor_ccy'] = 0
                    relevant_row['roc_fund_ccy'] = 0
                    relevant_row['roc_investor_ccy'] = 0
                    relevant_row['capital_gains_fund_ccy'] = 0
                    relevant_row['capital_gains_investor_ccy'] = 0
                    
                    # Update the fund op type as it may no longer be applicable 
                    if fee.fee_amount > 0 or fee.Mapping == 'Subsequent Close Interest':
                        relevant_row['fund_op_type'] = 'IF: Call'
                        relevant_row['fund_op_code'] = 'CC'
                    elif fee.fee_amount < 0 and fee.Mapping != 'Subsequent Close Interest':
                        relevant_row['fund_op_type'] = 'IF: Return Of Call'
                        relevant_row['fund_op_code'] = 'CD'
                    
                    # Set the fee amount based on the LR Compilation file 
                    if round(relevant_row.loc[i,'fees_fund_ccy_inside_commitment'],2) == round(fee.fee_amount,2):
                        # The actual amounts should not change, so we don't need to calculate anything. Just set the other amount to 0 because it'll get added by SplitInd = 102
                        relevant_row.loc[i,'fees_fund_ccy'] = 0
                        relevant_row.loc[i,'fees_investor_ccy'] = 0
                        print(round(relevant_row.loc[i,'fees_fund_ccy_inside_commitment'],2))
                        print(round(relevant_row.loc[i,'fees_investor_ccy_inside_commitment'],2))
                    elif round(relevant_row.loc[i,'fees_fund_ccy'],2) == round(fee.fee_amount,2):
                        # The actual amounts should not change, so we don't need to calculate anything. Just set the other amount to 0 because it'll get added by SplitInd = 102 
                        relevant_row.loc[i,'fees_fund_ccy_inside_commitment'] = 0
                        relevant_row.loc[i,'fees_investor_ccy_inside_commitment'] = 0
                        print(round(relevant_row.loc[i,'fees_fund_ccy'],2))
                        print(round(relevant_row.loc[i,'fees_investor_ccy'],2))
                    else:
                        # We dont expect any rows in our compilation file corresponding to mixed fees, so throw an error if we end up here
                        print(f"\t\tWarning 101: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
                        print(f"\t\tWarning 101: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")
                        logging.warning(f"\t\tWarning 101: row not found for fee: {fee.Share} - {fund} - {fee.Date} - {fee.old_description} - {fee.fee_amount}")
                        logging.warning(f"\t\tWarning 101: {relevant_row.loc[i,'date'], relevant_row.loc[i,'description'], relevant_row.loc[i,'fees_fund_ccy'],relevant_row.loc[i,'fees_fund_ccy_inside_commitment']}")

                    # Add the new row to dst_rows_copy
                    dst_rows_copy = pd.concat([dst_rows_copy, relevant_row], ignore_index=True)

    # Create new fee columns 
    dst_rows_copy['legal_fees_fund_ccy'] = None
    dst_rows_copy['legal_fees_investor_ccy'] = None
    dst_rows_copy['org_costs_fund_ccy'] = None
    dst_rows_copy['org_costs_investor_ccy'] = None
    dst_rows_copy['partnership_expenses_fund_ccy'] = None
    dst_rows_copy['partnership_expenses_investor_ccy'] = None
    dst_rows_copy['other_expenses_fund_ccy'] = None
    dst_rows_copy['other_expenses_investor_ccy'] = None
    dst_rows_copy['working_capital_fund_ccy'] = None
    dst_rows_copy['working_capital_investor_ccy'] = None
    dst_rows_copy['sub_close_interest_call_fund_ccy'] = None
    dst_rows_copy['sub_close_interest_call_investor_ccy'] = None
    dst_rows_copy['legal_fees_fund_ccy_inside_commitment'] = None
    dst_rows_copy['legal_fees_investor_ccy_inside_commitment'] = None
    dst_rows_copy['org_costs_fund_ccy_inside_commitment'] = None
    dst_rows_copy['org_costs_investor_ccy_inside_commitment'] = None
    dst_rows_copy['partnership_expenses_fund_ccy_inside_commitment'] = None
    dst_rows_copy['partnership_expenses_investor_ccy_inside_commitment'] = None
    dst_rows_copy['other_expenses_fund_ccy_inside_commitment'] = None
    dst_rows_copy['other_expenses_investor_ccy_inside_commitment'] = None
    dst_rows_copy['working_capital_fund_ccy_inside_commitment'] = None
    dst_rows_copy['working_capital_investor_ccy_inside_commitment'] = None
    dst_rows_copy['sub_close_interest_call_fund_ccy_inside_commitment'] = None
    dst_rows_copy['sub_close_interest_call_investor_ccy_inside_commitment'] = None
    
    # Create new distribution columns 
    dst_rows_copy['dividend_fund_ccy'] = None
    dst_rows_copy['dividend_investor_ccy'] = None
    dst_rows_copy['interest_fund_ccy'] = None
    dst_rows_copy['interest_investor_ccy'] = None
    dst_rows_copy['other_income_fund_ccy'] = None
    dst_rows_copy['other_income_investor_ccy'] = None
    dst_rows_copy['realised_gain_fund_ccy'] = None
    dst_rows_copy['realised_gain_investor_ccy'] = None
    dst_rows_copy['withholding_tax_fund_ccy'] = None
    dst_rows_copy['withholding_tax_investor_ccy'] = None
    dst_rows_copy['carry_fund_ccy'] = None
    dst_rows_copy['carry_investor_ccy'] = None
    dst_rows_copy['sub_close_interest_dist_fund_ccy'] = None
    dst_rows_copy['sub_close_interest_dist_investor_ccy'] = None

    # Loop through dst_rows_copy and move the amounts based on the mapping
    for dst_row in dst_rows_copy.itertuples():
        if dst_row.mapping == dst_row.mapping:
            # our mapping is populated so use it 
            if dst_row.fund_op_type == 'IF: Impairment' or dst_row.mapping == 'Impairment':
                # Don't worry about operation currency here as we'll add it in the add_op_cols function later. We want to add the cap gain amount here (should be -ve)
                dst_rows_copy.loc[dst_row.Index,'impairment_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'impairment_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'roc_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'roc_investor_ccy'] = 0
            elif dst_row.mapping == 'Legal Fees':
                dst_rows_copy.loc[dst_row.Index,'legal_fees_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'legal_fees_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Legal Fees inside commitment
                dst_rows_copy.loc[dst_row.Index,'legal_fees_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'legal_fees_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            elif dst_row.mapping == 'Working Capital':
                # Working Capital outside commitment
                dst_rows_copy.loc[dst_row.Index,'working_capital_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'working_capital_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Working Capital inside commitment
                dst_rows_copy.loc[dst_row.Index,'working_capital_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'working_capital_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            elif dst_row.mapping == 'Subsequent Close Interest':
                # Subsequent Close Interest outside commitment
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_call_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_call_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Subsequent Close Interest inside commitment
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_call_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_call_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            elif dst_row.mapping == 'Partnership Expenses':
                # Partnership Expenses outside commitment
                dst_rows_copy.loc[dst_row.Index,'partnership_expenses_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'partnership_expenses_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Partnership Expenses inside commitment
                dst_rows_copy.loc[dst_row.Index,'partnership_expenses_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'partnership_expenses_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            elif dst_row.mapping == 'Organizational Costs':
                # Organizational Costs outside commitment
                dst_rows_copy.loc[dst_row.Index,'org_costs_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'org_costs_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Organizational Costs inside commitment
                dst_rows_copy.loc[dst_row.Index,'org_costs_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'org_costs_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            elif dst_row.mapping == 'Other Expenses':
                # Other Expenses outside commitment
                dst_rows_copy.loc[dst_row.Index,'other_expenses_fund_ccy'] = dst_row.fees_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'other_expenses_investor_ccy'] = dst_row.fees_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy'] = 0
                # Other Expenses inside commitment
                dst_rows_copy.loc[dst_row.Index,'other_expenses_fund_ccy_inside_commitment'] = dst_row.fees_fund_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'other_expenses_investor_ccy_inside_commitment'] = dst_row.fees_investor_ccy_inside_commitment
                dst_rows_copy.loc[dst_row.Index,'fees_fund_ccy_inside_commitment'] = 0
                dst_rows_copy.loc[dst_row.Index,'fees_investor_ccy_inside_commitment'] = 0
            # Distributions
            elif dst_row.mapping == 'Interests':
                # Interests
                dst_rows_copy.loc[dst_row.Index,'interest_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'interest_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
            elif dst_row.mapping == 'Carry':
                # Realised Gain/Loss
                dst_rows_copy.loc[dst_row.Index,'carry_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'carry_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
            elif dst_row.mapping == 'Dividends':
                # Dividends
                dst_rows_copy.loc[dst_row.Index,'dividend_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'dividend_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
            elif dst_row.mapping == 'Other Income':
                # Other Income
                dst_rows_copy.loc[dst_row.Index,'other_income_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'other_income_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
            elif dst_row.mapping == 'SCI - Dist':
                # Subsequent Close Interest
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_dist_fund_ccy'] = dst_row.capital_gains_fund_ccy
                dst_rows_copy.loc[dst_row.Index,'sub_close_interest_dist_investor_ccy'] = dst_row.capital_gains_investor_ccy
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0
            elif dst_row.mapping == 'Withholding Tax':
                # Withholding Tax
                dst_rows_copy.loc[dst_row.Index,'withholding_tax_fund_ccy'] = -dst_row.capital_gains_fund_ccy # The amount will appear with the opposite sign in eFront to what's in the source file, but the IRR trace will be correct. 
                dst_rows_copy.loc[dst_row.Index,'withholding_tax_investor_ccy'] = -dst_row.capital_gains_investor_ccy # The amount will appear with the opposite sign in eFront to what's in the source file, but the IRR trace will be correct. 
                dst_rows_copy.loc[dst_row.Index,'capital_gains_fund_ccy'] = 0
                dst_rows_copy.loc[dst_row.Index,'capital_gains_investor_ccy'] = 0

    return dst_rows_copy 

######################################################################################
#                                  Managed Fund Ops                                  #
######################################################################################
def migrate_managed_data(input_file, dst_file, env='UAT', debug=False):
    """
        Inputs:
            input_file: the path to the MASTERFILE containing all the managed fund ops for different shares 
            dst_file: the path to the eFront import file which will contain the fund ops from the input_file  
        Migrate managed fund ops.
    """

    # ---------------------1---------------------
    # Define the template and import key data
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/N2. Managed Fund Operations - Test Template.xlsx'

    investor_details_file = 'C:/Users/RajContractor/Documents/Lion River/LR Reports/Managed Funds/Investor_Details.xlsx'
    investor_details = pd.read_excel(investor_details_file)
    
    # ---------------------2--------------------- 
    # Define our dst_rows dataframe 
    dst_rows = pd.DataFrame()
    dst_rows_unsplit = dst_rows

    # ---------------------3---------------------
    # Open our masterfile, loop through each share in investor_details and find the sheet containing fund ops for that share  
    xl = pd.ExcelFile(input_file)
    for sheet in xl.book.worksheets:
        if sheet.sheet_state == 'hidden':
            xl.book.remove(sheet)
    desc_summary = pd.DataFrame(columns=['share','date','fund_op_type','desc'])

    # ---------------------4---------------------
    # Get the official list of investors. Make sure every investor in our Com_Nom file matches the investors import (it should do already)
    if env == 'UAT':
        # Use the UAT import files
        descriptions_map_file = 'C:/Users/RajContractor/Documents/Lion River/LR Reports/Managed Funds/DescriptionsMap_UAT.xlsx'
        investors_import_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/06.1 Investors.xlsx'
    else:
        # Use the DEV import files
        descriptions_map_file = 'C:/Users/RajContractor/Documents/Lion River/LR Reports/Managed Funds/DescriptionsMap_DEV.xlsx'
        investors_import_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/06.1 DEV Investors.xlsx'
    investors_import = pd.read_excel(investors_import_file, index_col=None, header=2, usecols='C:H',sheet_name='Investors Import')[1:]
    descriptions_map = pd.read_excel(descriptions_map_file,index_col=0) 
    # Replace any old names with the new ones
    old_names = investors_import['Description'].dropna()
    old_names = old_names.apply(lambda x: x.split('\"')[1])
    old_names = old_names.apply(lambda x: x.strip())
    for index, name in old_names.iteritems():
        for row in investor_details.itertuples():
            if row.investor.strip() == name:
                investor_details.loc[row.Index,'investor'] = investors_import.loc[index,'Investor']
    # Replace all the names with the best match from the list of investors we've imported into the system
    official_investors = investors_import['Investor'].to_list()
    for row in investor_details.itertuples():
        investor_details.loc[row.Index,'investor'] = process.extractOne(row.investor,official_investors)[0]

    # Define new dictionaries so that we make a log of all the descriptions that we attribute to an investor
    new_descriptions = {}
    new_descriptions_shares = {}
    # Define a new dictionary so that we make a log of all the descriptions that we don't attribute to any given investor 
    split_descriptions_shares = {}
    x_investor = None 

    for share in list(investor_details['share'].unique()):
        # Read in the sheet for the current share from the masterfile, store the info in a dataframe and format it
        if share in xl.sheet_names: 
            input_rows = xl.parse(f'{share}',skiprows=9,usecols='A:J')

            input_rows.dropna(subset = ['Date'], inplace=True) # Drop all rows where date isn't populated
            input_rows['commitment'] = 0 # Make commitment = 0 for these rows 
            input_rows['commitment_euros'] = None
            input_rows.columns = input_rows.columns.str.lower() 
            input_rows['description'].fillna(' ', inplace=True) # Take care of null values
            input_rows.fillna(0, inplace=True)  
            input_rows['split_by_investor'] = True # Assume we want to split each row across all valid investors. Later we will set this to False if appropriate. 

            # Fetch the investor details for this share. This will be a list of all investors for this share, along with their commitment and nominal
            # Loop through each investor for the share
            inv_det = investor_details[investor_details['share']==share]
            issue_tot = inv_det.issue.sum()
            previous_description = ''
            previous_date = None

            # Go through our rows and insert the ones that have been split out by investors
            print(f"Share: {share}")
            logging.info(f"Share: {share}")
            print(f"\tSubstituting description with best matched investor...")
            logging.info(f"\tSubstituting description with best matched investor...")
            
            # Fill out the latest accounting nav column. This column isn't required for managed funds but we need to add it to prevent errors. 
            input_rows['latest accounting nav'] = input_rows['fair value'].shift(1)

            for index, input_row in input_rows.iterrows():
                if input_row['description'] != ' ' and debug==False:
                    # We've found a row where the description is populated
                    # Find all the valid investors for the date of this fund operation 
                    investors = inv_det[inv_det['transfer_ind']==0]
                    for i, investor in inv_det[inv_det['transfer_ind']==1].iterrows():
                        if input_row['date'] >= investor['transfer_date'] and investor['transfer'] > 0 and investor['multi_row'] == 0:
                            # New investor, date after transfer. Add new investor to list.  
                            investors = investors.append(investor)
                        elif input_row['date'] <= investor['transfer_date'] and investor['transfer'] < 0 and investor['multi_row'] == 0:
                            # Old investor, date before transfer. Add old investor to list. 
                            investors = investors.append(investor)
                        elif investor['transfer'] < 0 and -investor['transfer'] < investor['issue'] and investor['multi_row'] == 0:
                            # An investor did not transfer all of their shares 
                            investors = investors.append(investor)
                        elif input_row['date'] < investor['transfer_date'] and investor['transfer'] > 0 and investor['issue'] > 0:
                            # Old investor that had even more shares added to them - fix for W shares Generali Deutschland AG
                            investors = investors.append(investor)
                        elif investor['multi_row'] == 1:
                            second_row = inv_det[(inv_det['investor']==investor['investor'])&(inv_det['multi_row'] == 2)]
                            # If the date of the fund op is within the period the investor was present then add them to the available investors
                            if second_row['transfer_date'].values[0] > investor['transfer_date']:
                                entry_date = investor['transfer_date']
                                exit_date = second_row['transfer_date'].values[0]
                            else:
                                entry_date = second_row['transfer_date']
                                exit_date = investor['transfer_date'].values[0]
                            if input_row['date'] >= entry_date and input_row['date'] <= exit_date:
                                investors = investors.append(investor)
                            # If the investor never fully exited then add them to the available investors
                            if (second_row['transfer'].values[0] + second_row['issue'].values[0]) != -(investor['transfer'] + investor['issue']):
                                investors = investors.append(investor)

                    # For the list of valid investors, find the investor it's most likely to be
                    description_raw = input_row['description']
                    description = description_raw.lower()
                    description = description.replace('france','generali vie s.a.')
                    description = description.replace('germany','deutschland')
                    description = description.replace('italy','assicurazioni')
                    ratio = investors['investor'].apply(lambda x: fuzz.token_sort_ratio(description, x.replace('ITV - ','').lower()))
                    ratio_max = ratio.max()
                    row = investors.loc[ratio.idxmax()]

                    print(f'\n\t\tRow {index}, {ratio_max}% match found')
                    logger.info(f'\t\tRow {index}, {ratio_max}% match found')

                    if ratio_max != 100 and description_raw not in list(descriptions_map.index):
                        # Exact match not found. We haven't looked at this row before. Get the user to check and decide what the right investor should be 
                        row, user_input = investor_user_check(description_raw,input_row,investors,inv_det,row)
                        correct_investor = row['investor']
                    elif description_raw in list(descriptions_map.index):
                        # We've looked at this investor before, use the decision we made last time 
                        correct_investor = descriptions_map.loc[description_raw,'official_name']
                        try:
                            if correct_investor in list(investors['investor']):
                                print(f"\t\tDescription:   '{description_raw}'\n\t\tCorrect match: '{correct_investor}'")
                                user_input = 'y'
                            else:
                                # We haven't found a perfect match and we need the user to check what the substitution should be
                                row, user_input = investor_user_check(description_raw,input_row,investors,inv_det,row)
                                correct_investor = row['investor']
                        except ValueError:
                            print('VALUE ERROR', correct_investor, '\n', investors['investor'], '\n', list(investors['investor']))
                    else:
                        # We found a 100% match, no need for user input
                        correct_investor = row['investor']
                        print(f"\t\tDescription:   '{description_raw}'\n\t\tCorrect match: '{correct_investor}'")
                        user_input = 'y'

                    if user_input.lower() == 'y':
                        # Log it                     
                        logger.info(f"\t\tDescription:   '{description_raw}' Correct match: '{correct_investor}'")
                        if description_raw not in list(new_descriptions.keys()):
                            new_descriptions[description_raw] = correct_investor
                            new_descriptions_shares[description_raw] = [share]
                        else:
                            if share not in new_descriptions_shares[description_raw]:
                                new_descriptions_shares[description_raw].append(share)

                        # Remember not to split this by investors
                        input_rows.loc[index, 'split_by_investor'] = False

                        # Figure out what the number of shares issues is     
                        if row['transfer_ind'] == 0 or (row['transfer_ind'] == 1 and row['transfer'] < 0):
                            # No transfer. We want to only populate shares_issued on commitments and transfers. 
                            input_row['issue'] = row['issue']
                            input_row['shares_issued'] = None
                        elif row['transfer_ind'] == 1 and row['transfer'] > 0:
                            # new investor. We want to only populate shares_issued on commitments and transfers. 
                            input_row['issue'] = row['transfer']
                            input_row['shares_issued'] = None
                            
                        # Process this row and add it to our dest_rows record 
                        dst_rows_unsplit = append_row(share                                                   
                                            ,'EUR'
                                            ,correct_investor                                    
                                            ,input_row    
                                            ,input_row       
                                            ,dst_rows_unsplit
                                            ,fx_rates = None
                                            ,investee_fund=False)

                        if len(dst_rows_unsplit) > 1:
                            penultimate_fund_op = dst_rows_unsplit.iloc[-2]['fund_op_type']
                            latest_fund_op = dst_rows_unsplit.iloc[-1]['fund_op_type'] 
                            penultimate_investor = dst_rows_unsplit.iloc[-2]['investor']  
                            latest_investor = dst_rows_unsplit.iloc[-1]['investor']    
                            # Save the previous description on this day
                            if previous_date is None or input_row['date'] > previous_date:
                                # New date so reset the description
                                previous_date = input_row['date']
                                previous_description = input_row['description']
                            elif input_row['date'] == previous_date and penultimate_fund_op != latest_fund_op:
                                # New fund operation so reset the description
                                previous_description = input_row['description']
                            elif input_row['date'] == previous_date and penultimate_fund_op == latest_fund_op and penultimate_investor == latest_investor:
                                # Same fund operation on the same investor as before so reset the description
                                previous_description = input_row['description']
                            elif input_row['date'] == previous_date and penultimate_fund_op == latest_fund_op and penultimate_investor != latest_investor:
                                # Same fund operation on a different investor - append the description to the previous one
                                previous_description = previous_description + '; ' + input_row['description']

                        elif len(dst_rows_unsplit) == 1:
                            penultimate_fund_op = None
                            latest_fund_op = dst_rows_unsplit.iloc[-1]['fund_op_type']
                            previous_date = input_row['date']
                            previous_description = input_row['description']

                        desc_summary = desc_summary.append({'share': share, 'date': previous_date, 'fund_op_type': latest_fund_op,'desc': previous_description},ignore_index=True)
                        
                    else:
                        if description_raw not in list(split_descriptions_shares.keys()):
                            split_descriptions_shares[description_raw] = [share]
                        else:
                            if share not in split_descriptions_shares[description_raw]:
                                split_descriptions_shares[description_raw].append(share)
                        print(f"\t\tRow split by investors")
                        logger.info(f"\t\tRow split by investors")                                                 

        # Output to an excelfile 
        descriptions_map_df = pd.DataFrame({'official_name': new_descriptions.values(), 'shares': new_descriptions_shares.values()}, index=new_descriptions.keys())
        split_descriptions_map_df = pd.DataFrame({'shares': split_descriptions_shares.values()}, index=split_descriptions_shares.keys())
        if env == 'UAT':
            # Use the UAT import files
            with pd.ExcelWriter('C:/Users/RajContractor/Documents/Lion River/LR Reports/Managed Funds/DescriptionsMap_UAT.xlsx') as writer:
                descriptions_map_df.to_excel(writer,sheet_name='Substituted Descriptions') 
                split_descriptions_map_df.to_excel(writer,sheet_name='Split Descriptions')
        else:
            # Use the UAT import files
            with pd.ExcelWriter('C:/Users/RajContractor/Documents/Lion River/LR Reports/Managed Funds/DescriptionsMap_DEV.xlsx') as writer:
                descriptions_map_df.to_excel(writer,sheet_name='Substituted Descriptions') 
                split_descriptions_map_df.to_excel(writer,sheet_name='Split Descriptions')
        
        for index, row in inv_det.iterrows():
            if row['transfer_ind'] == 0:
                # We have found an original investor
                # Start by inserting the commitment for each original investor for the given share class.
                src_row = {
                    'commitment': row['commitment'],
                    'commitment_euros': row['commitment'],
                    'date': dt.datetime(row['issue_date'].year,1,1),
                    'description': 'MF: Commitment',
                    'investments': None,
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['issue'],
                    'shares_issued': row['issue']
                }

                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,src_row    
                                    ,src_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)

            elif (row['transfer_ind'] ==1 and row['transfer'] < 0 and row['multi_row'] == 0):
                # We have found an original investor that was transferred 
                # Insert a commitment row
                src_row = {
                    'commitment': row['commitment'],
                    'commitment_euros': row['commitment'],
                    'date': dt.datetime(row['issue_date'].year,1,1),
                    'description': 'MF: Commitment',
                    'investments': None,
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['issue'],
                    'shares_issued': row['issue']
                }
                # Insert a transfer operation - negative sign gets taken care by the fact that the transfer amount will be -ve
                transfer_row = {
                    'commitment': row['commitment']*row['transfer']/row['issue'],
                    'commitment_euros': row['commitment']*row['transfer']/row['issue'],
                    'date': row['transfer_date'],
                    'description': 'MF: Transfer',
                    'investments': None,
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['transfer'],
                    'shares_issued': row['transfer']
                }

                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,transfer_row    
                                    ,transfer_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)

                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,src_row    
                                    ,src_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)
            elif (row['transfer_ind'] ==1 and row['transfer'] > 0 and row['issue'] > 0):
                # We have found an original investor that had more shares transferred to them at a alter date
                # Insert a commitment row
                src_row = {
                    'commitment': row['commitment'],
                    'commitment_euros': row['commitment'],
                    'date': dt.datetime(row['issue_date'].year,1,1),
                    'description': 'MF: Commitment',
                    'investments': None,
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['issue'],
                    'shares_issued': row['issue']
                }
                # Insert the transfer row
                transfer_row = {
                    'commitment': row['commitment']*(row['transfer']+row['issue'])/row['issue'],
                    'commitment_euros': row['commitment']*(row['transfer']+row['issue'])/row['issue'],
                    'date': row['transfer_date'],
                    'description': 'MF: Transfer',
                    'investments': None,
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['transfer'],
                    'shares_issued': row['transfer']
                }
                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,src_row    
                                    ,src_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)

                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,transfer_row    
                                    ,transfer_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)
            else:   
                # We have found a row for a transferred investor 
                # Insert a transfer operation 
                transfer_row = {
                    'commitment': row['commitment'],
                    'commitment_euros': row['commitment'],
                    'date': row['transfer_date'],
                    'description': 'MF: Transfer',
                    'investments': None, # Everything is null for now - we'll populate it later
                    'return of capital': None,
                    'capital gains': None,
                    'fair value': None,
                    'fees': None,
                    'issue': row['transfer'],
                    'shares_issued': row['transfer']
                }
                dst_rows = append_row(share                                                   
                                    ,'EUR'
                                    ,row['investor']                                    
                                    ,transfer_row    
                                    ,transfer_row       
                                    ,dst_rows
                                    ,fx_rates=None
                                    ,investee_fund=False)
                  
            # Transfer each fund op from the masterfile to the dst_rows dataframe 
            for index, input_row in input_rows.iterrows():

                if input_row['split_by_investor'] == True:
                    
                    add_row_ind, investor, issue = calc_investor(input_row, row)
                    #print(f"Investor: {investor} Date: {input_row['date']} Add: {add_row_ind}")

                    if add_row_ind:
                        if share == 'T Shares' and (x_investor is None or x_investor != investor):
                            x_investor = investor
                            print(f"Investor: {investor}\tIssue Total: {issue_tot}\tIssue: {issue}\tRatio: {issue/issue_tot}")
                        src_row = {
                            'commitment': 0,
                            'commitment_euros': 0,
                            'date': input_row['date'],
                            'description': input_row['description'],
                            'investments': input_row['investments']*issue/issue_tot,
                            'return of capital': input_row['return of capital']*issue/issue_tot,
                            'capital gains': input_row['capital gains']*issue/issue_tot,
                            'fair value': input_row['fair value']*issue/issue_tot,
                            'fees': input_row['fees']*issue/issue_tot
                        }

                        # Not needed for managed funds, but we need the latest accounting nav for the append row function
                        if index > 0:
                            src_row['latest accounting nav'] = input_rows.iloc[index-1]['fair value']
                        else:
                            src_row['latest accounting nav'] = None      

                        # Add issue to src_row. We only want to populate shares_issued for commitment and transfer fund ops      
                        src_row['issue'] = issue
                        src_row['shares_issued'] = None                  

                        # Process this row and add it to our dest_rows record 
                        dst_rows = append_row(share                                                   
                                            ,'EUR'
                                            ,investor                                     
                                            ,src_row    
                                            ,src_row       
                                            ,dst_rows
                                            ,fx_rates=None
                                            ,investee_fund=False)
                    

    # ---------------------7---------------------
    # Where the rows were split by investors in the source data, take the max value of desc_summary, because this will contain all the descriptions for the same fund op on the same date 
    desc = desc_summary.groupby(['share','date','fund_op_type'])['desc'].max().reset_index()

    # dst_rows_unsplit contains all the fund ops where the rows were split by investors in the source data
    if not dst_rows_unsplit.empty:
        dst_rows_unsplit['share'] = dst_rows_unsplit['fund_name']
        dst_rows_unsplit['date'] = dst_rows_unsplit.date.astype('datetime64[ns]')
        desc['date'] = desc.date.astype('datetime64[ns]')
        dst_rows_unsplit = pd.merge(dst_rows_unsplit,desc,on=['date','share','fund_op_type'])
        dst_rows['share'] = dst_rows['fund_name'] 
        dst_rows = dst_rows.append(dst_rows_unsplit,ignore_index=True)

    # ---------------------8---------------------
    # Calculate the premium for the first call op and null out the issue for subsequent fund ops    
    for share in list(investor_details['share'].unique()):
        if share in xl.sheet_names: 
            dst_rows_share = dst_rows[dst_rows['fund_name']==share]
            inv_det = investor_details[investor_details['share']==share].copy()

            # Make any bespoke changes. Currently just removing call operations from A Shares
            dst_rows = bespoke_changes(dst_rows,share)

            # ---------------------9---------------------
            # We need to make sure that any transfer operations have the correct amounts 
            # The approach we take is we find all the rows with a negative transfer and then sum up values pre-transfer date and add them to the transfer operation based on the number of shares transferred
            # We usually have a one to one or one to many relationship. One investor transfers out and one or more investors inherit their shares. The one exception is K Shares, where multiple investors transfer into a single new investor 
            if 1 in list(inv_det['transfer_ind']): 
                # We only care about the investors with negative transfers as we care abouts amounts transferring out 
                inv_det_exited = inv_det[(inv_det['transfer_ind']==1)&(inv_det['transfer']<0)]

                for exited_investor in list(inv_det_exited['investor']):
                    
                    # Sum all the rows where shares are issued (this shouldn't be needed as there should just be 1)
                    exited_investor_issue = inv_det[inv_det['investor'] == exited_investor]['issue'].sum()

                    # Total shares transfered FROM the investor                   
                    if len(inv_det[(inv_det['investor'] == exited_investor)&(inv_det['transfer']<0)]) == 1:
                        tot_transfer_out = -inv_det[(inv_det['investor'] == exited_investor)&(inv_det['transfer']<0)]['transfer'].values[0]
                        transfer_date = pd.to_datetime(inv_det[(inv_det['investor'] == exited_investor)&(inv_det['transfer']<0)]['transfer_date'].values[0])       
                    else:
                        # The below was just added if there were any cases where the investor transferred out via multiple transfer operations. If there were such cases we'd need to cater for them, but there aren't.
                        print(f'Warning: Multiple transfers out of {exited_investor} for {share}')
                        logging.warning(f'Multiple transfers out of {exited_investor} for {share}\n')

                    # Total shares transferred TO the investor 
                    exited_investor_positive_transfer = inv_det[(inv_det['investor'] == exited_investor)&(inv_det['transfer'] > 0)]['transfer'].sum()
                    exited_investor_shares = exited_investor_issue + exited_investor_positive_transfer

                    # Throw an error if somehow more shares are going out than coming in 
                    if exited_investor_shares - tot_transfer_out < 0:
                        print(f'Warning: Tranfer out of {exited_investor} is greater than the issue/tranfer in for {share}')
                        logging.warning(f'Tranfer out of {exited_investor} is greater than the issue/tranfer in for {share}\n')

                    # Find all the rows corresponding to the investor who is transferring out and sum up the values prior to the transfer date into dst_rows_share_investor_tot 
                    dst_rows_share_pre_transfer = dst_rows_share[dst_rows_share['date']<transfer_date]
                    dst_rows_share_investor = dst_rows_share_pre_transfer[dst_rows_share_pre_transfer['investor_trunc']==exited_investor]
                    dst_rows_share_investor_tot = dst_rows_share_investor.sum()
                    
                    # Find the latest NAV for the investor who is transferring out 
                    latest_nav_df = dst_rows_share_investor[dst_rows_share_investor['fair_value_fund_ccy'] > 0]
                    latest_nav_df = latest_nav_df.sort_values(by=['date'],ascending=False, ignore_index=True)                    

                    if tot_transfer_out < exited_investor_shares:
                        # The investor isn't completely transferring out
                        ratio = tot_transfer_out/exited_investor_shares
                        # The tot amounts are the amounts getting transferred to the new investors
                        tot_investments = dst_rows_share_investor_tot['investments_fund_ccy']*ratio
                        tot_commitment = dst_rows_share_investor_tot['commitment_investor_ccy']*ratio
                        tot_roc = dst_rows_share_investor_tot['roc_investor_ccy']*ratio
                        tot_cap_gains = dst_rows_share_investor_tot['capital_gains_investor_ccy']*ratio
                        tot_premium = dst_rows_share_investor_tot['premium']*ratio                        
                        if len(latest_nav_df) > 0:
                            tot_fair_value = latest_nav_df.loc[0,'fair_value_investor_ccy'] - latest_nav_df.loc[0,'fair_value_investor_ccy']*ratio
                        else:
                            print(f'Warning: No latest NAV for {exited_investor} in {share}')
                            logging.warning(f'No latest NAV for {exited_investor} in {share}\n')
                            tot_fair_value = 0
                    else:
                        # The investor is completely transferring out
                        # The tot amounts are the amounts getting transferred to the new investors
                        tot_investments = dst_rows_share_investor_tot['investments_fund_ccy']
                        tot_commitment = dst_rows_share_investor_tot['commitment_investor_ccy']
                        tot_roc = dst_rows_share_investor_tot['roc_investor_ccy']
                        tot_cap_gains = dst_rows_share_investor_tot['capital_gains_investor_ccy']
                        tot_premium = dst_rows_share_investor_tot['premium']
                        if len(latest_nav_df) > 0:
                            tot_fair_value = latest_nav_df.loc[0,'fair_value_investor_ccy']
                        else:
                            print(f'Warning: No latest NAV for {exited_investor} in {share}')
                            logging.warning(f'No latest NAV for {exited_investor} in {share}\n')
                            tot_fair_value = 0
                        # The new amounts are the amounts on the transfer operation for the investor transferring out 

                    # Set the amounts on the transfer operation for the investor transferring out 
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = -tot_investments
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = -tot_investments
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = -tot_investments
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'commitment_investor_ccy'] = -tot_commitment
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'roc_fund_ccy'] = -tot_roc
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'roc_investor_ccy'] = -tot_roc
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'capital_gains_fund_ccy'] = -tot_cap_gains
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'capital_gains_investor_ccy'] = -tot_cap_gains
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'fair_value_fund_ccy'] = -tot_fair_value
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'fair_value_investor_ccy'] = -tot_fair_value
                    dst_rows.loc[(dst_rows['investor_trunc'] == exited_investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'premium'] = -tot_premium

                    # Find the investor(s) that will replace the exiting investor
                    # The below will work. Whenever you have multiple investors exiting out, the exit on a unique date, so the incoming investors for that date should match
                    incoming_inv_det = inv_det[(inv_det['transfer_ind'] == 1)&(inv_det['transfer_date'] == transfer_date)&(inv_det['transfer']>0)].copy()
                    # K shares is the only case where two investors exit and one investor enters
                    if share == 'K Shares':
                        incoming_inv_det['ratio'] = incoming_inv_det['transfer']/(tot_transfer_out*2)
                    else:
                        incoming_inv_det['ratio'] = incoming_inv_det['transfer']/tot_transfer_out
                        
                    if incoming_inv_det['transfer'].sum() != tot_transfer_out and share != 'K Shares':
                        print(f'Warning: New investor not found when {exited_investor} exited out of {share}')
                        logging.warning(f'New investor not found when {exited_investor} exited out of {share}\n')

                    # Loop through our incoming investors and set the amounts on the transfer operation based on the number of shares transferred to the investor         
                    for row in incoming_inv_det.itertuples():
                        # What if the incoming investor was already there and they are having more shares added to them?
                        # What if investor transferring out is not fully transferring out?

                        # keep a track of all the values we're transferring
                        sum_investment = 0
                        sum_commitment = 0
                        sum_roc = 0
                        sum_cap_gains = 0
                        sum_fair_value = 0
                        sum_premium = 0
                        if row.Index != incoming_inv_det.index.max():
                            # There are multiple investors transferring in and there are more to come
                            # Add it to the sums
                            set_investment = tot_investments*row.ratio
                            set_commitment = tot_commitment*row.ratio
                            set_roc = tot_roc*row.ratio
                            set_cap_gains = tot_cap_gains*row.ratio
                            set_fair_value = tot_fair_value*row.ratio
                            set_premium = tot_premium*row.ratio
                            sum_investment += set_investment
                            sum_commitment += set_commitment
                            sum_roc += set_roc
                            sum_cap_gains += set_cap_gains
                            sum_fair_value += set_fair_value
                            sum_premium += set_premium
                        else:
                            # We're on the final investor transferring in and set the remaining amount left to be allocated to them
                            set_investment = tot_investments - sum_investment
                            set_commitment = tot_commitment - sum_commitment
                            set_roc = tot_roc - sum_roc
                            set_cap_gains = tot_cap_gains - sum_cap_gains
                            set_fair_value = tot_fair_value - sum_fair_value
                            set_premium = tot_premium - sum_premium

                        # Set this amount on the corresponding transfer operation for the investor transferring in   
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = set_investment
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = set_investment
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'investments_fund_ccy'] = set_investment
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'commitment_investor_ccy'] = set_commitment
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'roc_fund_ccy'] = set_roc
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'roc_investor_ccy'] = set_roc
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'capital_gains_fund_ccy'] = set_cap_gains
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'capital_gains_investor_ccy'] = set_cap_gains
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'fair_value_fund_ccy'] = set_fair_value
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'fair_value_investor_ccy'] = set_fair_value
                        dst_rows.loc[(dst_rows['investor_trunc'] == row.investor)&(dst_rows['fund_op_type']=='MF: Transfer')&(dst_rows['fund_name']==share), 'premium'] = set_premium

            # Populate the shares on the first call operation of every investor so we can populate the premium 
            nominal_investors = inv_det[(inv_det['issue'] > 0)&((inv_det['transfer_ind'] == 0)|((inv_det['full_transfer_ind'] == 1)&(inv_det['multi_row'] == 1)))]
            
            for investor in nominal_investors.itertuples():
                dst_rows_share_investor = dst_rows_share[dst_rows_share['investor'] == investor.investor]
                dst_rows_share_investor_calls = dst_rows_share_investor[dst_rows_share_investor['fund_op_type']=='MF: Call']
                dst_rows_share_investor_calls = dst_rows_share_investor_calls.sort_values(by=['date'])
                if investor.transfer > 0:
                    nominal_amount = investor.transfer
                else:
                    nominal_amount = investor.issue
                if len(dst_rows_share_investor_calls) > 0 and dst_rows_share_investor_calls.first_valid_index() in dst_rows.index:
                    if dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] >= float(nominal_amount):
                        # 'shares_issued' is None for all rows unless set here, 'issue' is always populated 
                        # Nominal == number of shares == issue 
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'no_shares'] = nominal_amount
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'nominal'] = nominal_amount
                        # Premium = call amount - nominal
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'premium'] = dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] - nominal_amount
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] = 0
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_fund_ccy'] = 0
                    else:
                        print(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(nominal_amount)}")
                        logging.warning(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(nominal_amount)}\n")
                else:
                    print('NOMINAL INVESTORS')
                    print(f'Warning: no call operation found for {investor.investor} in {share}')
                    logging.warning(f'No call operation found for {investor.investor} in {share}\n')

            partial_transfer_investors = inv_det[(inv_det['full_transfer_ind'] == 0)&(inv_det['transfer_ind'] == 1)&((inv_det['multi_row'] == 0)|(inv_det['multi_row'] == 1))]            
            for investor in partial_transfer_investors.itertuples():
                dst_rows_share_investor = dst_rows_share[dst_rows_share['investor'] == investor.investor]
                dst_rows_share_investor_calls = dst_rows_share_investor[dst_rows_share_investor['fund_op_type']=='MF: Call']
                dst_rows_share_investor_calls = dst_rows_share_investor_calls.sort_values(by=['date'])
                if len(dst_rows_share_investor_calls) > 0 and dst_rows_share_investor_calls.first_valid_index() in dst_rows.index:
                    if dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'date'] >= investor.transfer_date:
                        # Call operation happens after date of transfer so adjust the nominal by the transferred amount
                        nominal_amount = investor.issue + investor.transfer
                    else:
                        nominal_amount = investor.issue 
                    if dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] >= float(nominal_amount):
                        # 'shares_issued' is None for all rows unless set here, 'issue' is always populated 
                        # Nominal == number of shares == issue 
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'no_shares'] = nominal_amount
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'nominal'] = nominal_amount
                        # Premium = call amount - nominal
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'premium'] = dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] - nominal_amount
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] = 0
                        dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_fund_ccy'] = 0
                    else:
                        print(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(nominal_amount)}")
                        logging.warning(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(nominal_amount)}\n")
                else:
                    print('PARTIAL INVESTORS')
                    print(f'Warning: no call operation found for {investor.investor} in {share}')
                    logging.warning(f'No call operation found for {investor.investor} in {share}\n')

            transferred_investors = inv_det[(inv_det['transfer'] > 0)&(inv_det['issue'] == 0)]            
            for investor in transferred_investors.itertuples():
                if not partial_transfer_investors[partial_transfer_investors['investor']==investor.investor]['investor'].any():
                    dst_rows_share_investor = dst_rows_share[dst_rows_share['investor'] == investor.investor]
                    dst_rows_share_investor_calls = dst_rows_share_investor[dst_rows_share_investor['fund_op_type']=='MF: Call']
                    dst_rows_share_investor_calls = dst_rows_share_investor_calls.sort_values(by=['date'])
                    nominal_amount = investor.transfer
                    if len(dst_rows_share_investor_calls) > 0 and dst_rows_share_investor_calls.first_valid_index() in dst_rows.index:
                        if dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] >= float(nominal_amount):
                            # 'shares_issued' is None for all rows unless set here, 'issue' is always populated 
                            # Nominal == number of shares == transfer 
                            dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'no_shares'] = nominal_amount
                            dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'nominal'] = nominal_amount
                            # Premium = call amount - nominal
                            dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'premium'] = dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] - investor.transfer
                            dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy'] = 0
                            dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_fund_ccy'] = 0
                        else:
                            print(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(investor.transfer)}")
                            logging.warning(f"Warning: no call operation with sufficient investment found for {investor.investor} in {share}\t Investment: {dst_rows.loc[dst_rows_share_investor_calls.first_valid_index(),'investments_investor_ccy']} Issue: {float(investor.transfer)}\n")
                    else:
                        print('TRANSFERRED INVESTORS')
                        print(f'Warning: no call operation found for {investor.investor} in {share}')
                        logging.warning(f'No call operation found for {investor.investor} in {share}\n')

    # ---------------------9--------------------- 
    # We have set the 'premium' amount on transfer/first call operations. Now for the remaining cases we set it to match the investment amount
    null_premiums = dst_rows[dst_rows['premium']!=dst_rows['premium']]
    for row in null_premiums.itertuples():
        dst_rows.loc[row.Index,'premium'] = row.investments_fund_ccy

    # ---------------------10---------------------
    # Sort and format everything
    dst_rows.sort_values(by=['fund_name','date','fund_op_type'],inplace=True,ignore_index=True,ascending=[True,True,False])
    dst_rows = dst_rows.round(2)

    # ---------------------11---------------------
    # Add the return of premium column 
    dst_rows['return_of_premium'] = dst_rows['capital_gains_fund_ccy'] + dst_rows['roc_investor_ccy']

    # ---------------------12---------------------
    # The below doesn't do anything for the main data migration, but is a left as a failsafe. 
    # for any transfer operations, see if there is a call operation or distribution for that investor where shares_issued == premium (call operations) or capital_gains_fund_ccy + roc_investor_ccy (distributions)
    # The only stuff that got removed was for A Shares (5 calls, 1 dist). We explicitly remove those earlier now. Left this in case we run it on historic/future data.
    dst_rows_copy = dst_rows.copy()
    for share in list(dst_rows['share'].unique()):
        dst_rows_share = dst_rows[dst_rows['share'] == share]
        transfers = dst_rows_share[dst_rows_share['fund_op_type'] == 'MF: Transfer']
        for transfer in transfers.itertuples():
            if transfer.shares_issued > 0:
                i = list(dst_rows_share[(dst_rows_share['fund_op_type'] == 'MF: Call')&(dst_rows_share['premium'] == transfer.shares_issued)&(dst_rows_share['investor'] == transfer.investor)&(dst_rows_share['date'] == transfer.date)].index)
            elif transfer.shares_issued < 0:
                i = list(dst_rows_share[(dst_rows_share['fund_op_type'] == 'MF: Distribution')&(-dst_rows_share['return_of_premium'] == transfer.shares_issued)&(dst_rows_share['investor'] == transfer.investor)&(dst_rows_share['date'] == transfer.date)].index)
            if len(i) == 1:
                print(f"{share} - {dst_rows.loc[i[0],'fund_op_type']} removed")
                dst_rows_copy = dst_rows_copy.drop([i[0]])
    dst_rows = dst_rows_copy
    dst_rows_loc = 'C:/Users/RajContractor/Documents/Python Files/Dev/LR Migration/Migrated/dst_rows.xlsx'
    dst_rows.to_excel(dst_rows_loc,index=False)

    # ---------------------13---------------------
    # Add index column 
    dst_rows['index'] = 1
    one = dst_rows.groupby(['fund_name','date','fund_op_type','investor']).cumcount()
    dst_rows['index'] += one    
    prev_fund_op = None 
    prev_fund = None
    prev_investor = None
    prev_date = None
    count = 0
    for dst_row in dst_rows.itertuples():
        if prev_date is None:
            # First row. Ignore the index. Set our 'prev' variables
            prev_fund_op = dst_row.fund_op_type
            prev_fund = dst_row.fund_name 
            prev_investor = dst_row.investor 
            prev_date = dst_row.date
        else:
            # Not the first row. 
            if dst_row.fund_name == prev_fund and dst_row.date == prev_date and dst_row.fund_op_type != prev_fund_op:
                count += 1
            elif dst_row.fund_name != prev_fund or dst_row.date != prev_date:
                # Reset the count when the fund changes or the date changes
                count = 0
            # Update the index
            dst_rows.loc[dst_row.Index,'index'] += count
            # Jot down the values for the next iteration
            prev_fund_op = dst_row.fund_op_type
            prev_fund = dst_row.fund_name 
            prev_investor = dst_row.investor 
            prev_date = dst_row.date

    # ---------------------14--------------------   
    # Insert into our template file and save as a new migrated file      
    # Open the input template, find active sheet and ignore header rows 
    wb = oxl.load_workbook(template_file)
    dst_active_sheet = wb.active
    dst_row_num = 5 # Ignore headers in the template 
    for i, dst_row in dst_rows.iterrows():
        dst_row_num = insert_row(dst_row                                                        # pass in the row we want to insert                                                
                                ,dst_row_num                                                    # this is just the current row count 
                                ,dst_active_sheet                                               # the active sheet we're editing
                                ,False
                                )

    wb.save(dst_file) 

    # Close our source file
    xl.close()

def calc_investor(input_row, row):
    """
        Inputs:
            input_row: a series containing information about a single fund operation 
            row: the series containing details about an individual investor for the share that we're processing 
        Outputs:
            insert_row_ind: bool. If True, a row needs to be inserted in the import template. 
            investor: the investor that should be put on the fund operation insert 
            issue: the nominal that should go along with the fund op insert 
            For managed funds only. 
    """
    # Figure out if the investor has been transferred or not 
    if (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] > input_row['date']
    and row['transfer'] > 0
    and row['issue'] == 0):
        # This is a new investor and the fund op is prior to when they came in 
        add_row_ind = False
        issue = None
        investor = None
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] > input_row['date']
    and row['transfer'] > 0
    and row['issue'] > 0
    and row['multi_row'] == 0):
        # This is an original investor and the fund op is prior to when they had more shares transferred
        add_row_ind = True
        issue = row['issue']
        investor = row['investor']
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] > input_row['date']
    and row['transfer'] < 0
    and row['multi_row'] == 0):
        # This is an old investor and the fund op is prior to when they cashed out  
        add_row_ind = True
        issue = row['issue']
        investor = row['investor']
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] > 0
    and row['multi_row'] == 0):
        # This is a new investor and the fund op is after they came in
        add_row_ind = True
        issue = row['issue'] + row['transfer']
        investor = row['investor']
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] < 0
    and row['full_transfer_ind'] == 1):
        # This is an old investor and the fund op is after they cashed out 
        add_row_ind = False
        issue = None
        investor = None
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] < 0
    and row['full_transfer_ind'] == 0
    and row['multi_row'] == 0):
        # This is an old investor and the fund op is after they cashed out 
        add_row_ind = True
        issue = row['issue'] + row['transfer']
        investor = row['investor']
    elif row['transfer_ind'] == 0: 
        add_row_ind = True
        issue = row['issue']
        investor = row['investor']
    # So far, when deciding to add investors, we have neglected investors that have multiple transfers. That's what we do now
    # In each case, the issue will be 0. The first transfer is positive and the second negative. All except ITV - Generali Assurances Générales S.A. in O Shares fully transfer out (full_transfer_ind == 1).
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] > input_row['date']
    and row['transfer'] > 0
    and row['multi_row'] == 1):
        # The fund op is prior to when they had any shares transferred
        add_row_ind = False
        issue = None
        investor = None
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] > 0
    and row['multi_row'] == 1
    and not pd.isnull(row['second_transfer_date'])
    and row['second_transfer_date'] > input_row['date']):
        # This fund op is between when this investor transferred in and when they transferred out
        add_row_ind = True
        issue = row['transfer']
        investor = row['investor']
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] > 0
    and row['multi_row'] == 1
    and not pd.isnull(row['second_transfer_date'])
    and row['second_transfer_date'] <= input_row['date']
    and row['full_transfer_ind'] == 1):
        # This fund op is after the investor transferred out
        add_row_ind = False
        issue = None
        investor = None
    elif (row['transfer_ind'] == 1 
    and not pd.isnull(row['transfer_date'])
    and row['transfer_date'] <= input_row['date']
    and row['transfer'] > 0
    and row['multi_row'] == 1
    and not pd.isnull(row['second_transfer_date'])
    and row['second_transfer_date'] <= input_row['date']
    and row['full_transfer_ind'] == 0):
        # This is ITV - Generali Assurances Générales S.A. in O Shares
        add_row_ind = True
        issue = row['issue'] + row['transfer']
        investor = row['investor']
    elif row['multi_row'] == 2:
        # We've already taken care of this scenario
        add_row_ind = False
        issue = None
        investor = None
    else:
        print('-----------------ERROR------------------')
        logging.warning('-----------------ERROR------------------')
        print(row)
        logging.warning(row)
        add_row_ind = False
        issue = None
        investor = None
    return add_row_ind, investor, issue

def investor_user_check(description_raw,input_row,investors,inv_det,row):
    # We haven't found a perfect match and we need the user to check what the substitution should be
    print(f"\t\tDescription:   '{description_raw}'\n\t\tClosest match: '{row['investor']}'")
    user_input = input("\t\tY to accept, N to see other options and S to split by all investors for this date: ")

    while user_input.lower() != 'y' and user_input.lower() != 's':
        print('\t\t\tInvestors in this share for this date:')
        print('\t\t\t----------------------------------------')
        for i, investor in investors.iterrows(): 
            print(f"\t\t\t{i}. {investor['investor']}") 
        print('\t\t\t----------------------------------------')
        row_id = int(input("\t\t\tIndex: "))
        row = inv_det.loc[row_id]
        print(f"\t\t\tDescription:   '{input_row['description']}'\n\t\t\tCorrect match: '{row['investor']}'")
        user_input = input("\t\t\tY to accept, N to see other options and S to split by all investors for this date: ")

    return row, user_input

def bespoke_changes(dst_rows,share):
    """
        Make bespoke changes to dst_rows to reflect how LR want the data to be displayed in eFront
    """
    if share == 'A Shares':
        # for A shares only, get rid of the call operation prior to 2017
        dst_rows_share = dst_rows[dst_rows['share'] == 'A Shares']
        dst_rows.drop(dst_rows_share[(dst_rows_share['fund_op_type'] == 'MF: Call')&(dst_rows_share['date'] < dt.date(2017,1,1))].index,inplace=True)
        dst_rows.drop(dst_rows_share[(dst_rows_share['fund_op_type'] == 'MF: Distribution')&(dst_rows_share['date'] < dt.date(2017,1,1))].index,inplace=True) 
    # B Shares 
	# delete two return of call operations and add -28950708.66 to the call operation on 12/05/2009
    # first call op should be one with Nominal = 4998 and Premium = 43426063 that are already there (no change needed)
	# add new call operation with Nominal = 4998, Premium = 43426063 on 12/05/2009 and give this and remaining 2 call ops an index of 2 on this date
    # dst_rows.drop(dst_rows_share[(dst_rows_share['fund_op_type'] == 'MF: Call')&(dst_rows_share['date'] < dt.date(2017,1,1))].index,inplace=True)
    # b_drop = dst_rows[(dst_rows['fund_name'] == 'B Shares')&(dst_rows['investor_trunc'] == 'Assicurazioni Generali S.p.A.')].copy()
    # drop_index = b_drop[b_drop['fund_op_type'] == 'MF: Return Of Call (Negative Call)'].index.to_list()
    # dst_rows = dst_rows.drop((dst_rows['fund_name'] == 'B Shares')&(dst_rows['investor_trunc'] == 'Assicurazioni Generali S.p.A.')))    
    # dst_rows_b = dst_rows[(dst_rows['fund_name'] == 'B Shares')&(dst_rows['fund_op_type'] == 'MF: Call')].copy()

    return dst_rows

######################################################################################
#                                   Bank Operations                                  #
######################################################################################
def migrate_investee_fund_op_bank_ops(fund_ops, dest_file, env='UAT'):
    """ Inputs:
            fund_ops: a dataframe of all investee fund ops
            dst_file: the absolute path to the destination file the transformed data will be saved under 
        Extract the relevant data from our investee fund ops and rearrange it to fit the bank ops import template 
    """
    # Read in the template file - this already has the starting balances
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/BankOp_Template.xlsx'
    wb = oxl.load_workbook(template_file)
    # dst_active_sheet = wb.create_sheet('Investee Fund Ops', 0)
    dst_active_sheet = wb.active
    dst_active_sheet.title = 'Investee Fund Ops'
    dst_row_num = 10 # Ignore headers and starting balances in the template     

    # Read in the official bank ops linked to each fund 
    if env == 'UAT':
        # Use the UAT import files
        bank_account_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/10 Link Bank Accounts to Fund, Company, Investor - TODO!.xlsx'
    else:
        # Use the DEV import files
        bank_account_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/10 DEV Link Bank Accounts to Fund, Company, Investor.xlsx'
    bank_accounts = pd.read_excel(bank_account_file,index_col=None,skiprows=[0,2,3]) 

    # Read the data we need to transform
    fund_ops = fund_ops.fillna(0)
    missing_bank_accounts = []
    
    for fund_op in fund_ops.itertuples():
        if fund_op.OTYPE1 in ['IF: Accounting Valuation','IF: Official NAV','IF: Commitment']:
            pass
        else:
            # Get the bank accounts 
            investee_fund = fund_op.FUND2
            if investee_fund in list(bank_accounts['FUND7']):
                investee_fund_bank_account = bank_accounts[bank_accounts['FUND7']==investee_fund].iloc[0]['ACCOUNTCODE1']
            else:
                if investee_fund not in missing_bank_accounts:
                    missing_bank_accounts.append(investee_fund)
                investee_fund_bank_account = None
            managed_fund = fund_op.FUND18_MV
            if managed_fund in list(bank_accounts['FUND7']):
                managed_fund_bank_account = bank_accounts[bank_accounts['FUND7']==managed_fund].iloc[0]['ACCOUNTCODE1']
            else:
                if managed_fund not in missing_bank_accounts:
                    missing_bank_accounts.append(managed_fund)
                managed_fund_bank_account = None

            # Get the Amounts
                # CALL
                    # Fund CCY
                        # Investments
            call_investment_fund_ccy = fund_op.AMOUNT02111
                        # Fees
                            # Outside Commitment
            call_fees_outside_comm_fund_ccy = fund_op.AMOUNT03111 + fund_op.AMOUNT40111 + fund_op.AMOUNT05111 + fund_op.AMOUNT06111 + fund_op.AMOUNT07111 + fund_op.AMOUNT09111
                            # Inside Commitment
            call_fees_inside_comm_fund_ccy = fund_op.AMOUNT04111 + fund_op.AMOUNT88111 + fund_op.AMOUNT89111 + fund_op.AMOUNT90111 + fund_op.AMOUNT91111 + fund_op.AMOUNT10111
                    # Investor CCY
                        # Investments
            call_investment_investor_ccy = fund_op.AMOUNT02211
                        # Fees
                            # Outside Commitment
            call_fees_outside_comm_investor_ccy = fund_op.AMOUNT03211 + fund_op.AMOUNT05211 + fund_op.AMOUNT40211 + fund_op.AMOUNT06211 + fund_op.AMOUNT07211 + fund_op.AMOUNT09211
                            # Inside Commitment
            call_fees_inside_comm_investor_ccy = fund_op.AMOUNT88211 + fund_op.AMOUNT04211 + fund_op.AMOUNT89211 + fund_op.AMOUNT90211 + fund_op.AMOUNT91211 + fund_op.AMOUNT10211
                # DIST
                    # Func CCY
            dist_fund_ccy = fund_op.AMOUNT24111 + fund_op.AMOUNT17111 + fund_op.AMOUNT13111 + fund_op.AMOUNT14111 + fund_op.AMOUNT15111 + fund_op.AMOUNT23111 + fund_op.AMOUNT29111 + fund_op.AMOUNT28111
                # Investor ccy
            dist_investor_ccy = fund_op.AMOUNT24211 + fund_op.AMOUNT17211 + fund_op.AMOUNT13211 + fund_op.AMOUNT14211 + fund_op.AMOUNT15211 + fund_op.AMOUNT23211 + fund_op.AMOUNT29211 + fund_op.AMOUNT28211

            # if fund_op.OTYPE1 in ['IF: Call']:
            #     # Call
            #     amount_fund_ccy = call_investment_fund_ccy + call_fees_outside_comm_fund_ccy + call_fees_inside_comm_fund_ccy
            #     amount_investor_ccy = call_investment_investor_ccy + call_fees_outside_comm_investor_ccy + call_fees_inside_comm_investor_ccy
            #     if amount_fund_ccy >= 0:
            #         bank_op_type = 'Payment'
            #     else:
            #         bank_op_type = 'Receipt'
            # elif fund_op.OTYPE1 in ['IF: Distribution']:
            #     # Distributions
            #     amount_fund_ccy = dist_fund_ccy
            #     amount_investor_ccy = dist_investor_ccy
            #     if amount_fund_ccy >= 0:
            #         bank_op_type = 'Receipt'
            #     else:
            #         bank_op_type = 'Payment'

            amount_fund_ccy = call_investment_fund_ccy + call_fees_outside_comm_fund_ccy + call_fees_inside_comm_fund_ccy - dist_fund_ccy
            amount_investor_ccy = call_investment_investor_ccy + call_fees_outside_comm_investor_ccy + call_fees_inside_comm_investor_ccy - dist_investor_ccy
            if amount_fund_ccy >= 0:
                bank_op_type = 'Payment'
            else:
                bank_op_type = 'Receipt'
                amount_fund_ccy = -amount_fund_ccy
                amount_investor_ccy = -amount_investor_ccy

            if amount_investor_ccy != amount_investor_ccy or amount_investor_ccy == 0:
                # This is because often there are more fund ops in the fund ccy than the investor ccy - there should only be ~25 cases
                pass 
                # print('\tNo amount found:', managed_fund, investee_fund, fund_op.OTYPE1, fund_op.SETTLEMENTDATE1)
            else:
                amount_investor_ccy = round(amount_investor_ccy,2)
                amount_fund_ccy = round(amount_fund_ccy,2)
                dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
                dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
                dst_active_sheet[f'C{dst_row_num}'] = fund_op.SETTLEMENTDATE1               # -------------------------------------- Bank Op Date CLOSEDATE1
                dst_active_sheet[f'D{dst_row_num}'] = fund_op.CURRENCY31                    # -------------------------------------- Payment currency CURRENCY11
                dst_active_sheet[f'E{dst_row_num}'] = bank_op_type                      # -------------------------------------- Bank Op Type OPTYPE1
                dst_active_sheet[f'F{dst_row_num}'] = managed_fund                      # -------------------------------------- Managed Fund Name LINKEDENTITY
                dst_active_sheet[f'G{dst_row_num}'] = 'Fund'                            # -------------------------------------- Entity Type XX_ENTITYCLASS
                dst_active_sheet[f'H{dst_row_num}'] = managed_fund                      # -------------------------------------- Entity Name in Bank Op FUND5
                dst_active_sheet[f'I{dst_row_num}'] = 'Fund'                            # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
                dst_active_sheet[f'J{dst_row_num}'] = investee_fund                     # -------------------------------------- Counterparty Fund.Fund FUND_CF
                dst_active_sheet[f'K{dst_row_num}'] = investee_fund_bank_account        # -------------------------------------- Bank Account BANKACCOUNTC1
                dst_active_sheet[f'L{dst_row_num}'] = managed_fund_bank_account         # -------------------------------------- Bank Account ACCOUNTCODE2
                dst_active_sheet[f'M{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
                dst_active_sheet[f'N{dst_row_num}'] = amount_fund_ccy                   # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
                dst_active_sheet[f'O{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
                dst_active_sheet[f'P{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
                #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
                dst_active_sheet[f'Q{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Draft DRAFT1
                dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
                dst_active_sheet[f'S{dst_row_num}'] = fund_op.SETTLEMENTDATE1           # -------------------------------------- CLOSEDATE22
                dst_active_sheet[f'T{dst_row_num}'] = fund_op.INDEXOP1                  # -------------------------------------- INDEXOP22
                dst_active_sheet[f'U{dst_row_num}'] = fund_op.OTYPE1                    # -------------------------------------- Fund Op Type OTYPE22
                dst_active_sheet[f'V{dst_row_num}'] = ''                                # -------------------------------------- Comment
                dst_row_num += 1

    wb.save(dest_file)

    # Print missing bank accounts
    if len(missing_bank_accounts) > 0:
        print(f'\n\tNo bank account found for:')
        logging.warning(f'\n\tNo bank account found for:')
        for investor in missing_bank_accounts:
            print(f'\t\t{investor}')
            logging.warning(f'\t\t{investor}')

def migrate_managed_fund_op_bank_ops(input_file, dest_file, env='UAT'):
    """ Inputs:
            input_file: the absolute path to the source file
            dst_file: the absolute path to the destination file the transformed data will be saved under 
        Extract the relevant data from our investee fund ops and rearrange it to fit the bank ops import template 
    """
    # Read in the template file - this already has the starting balances
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/BankOp_Template.xlsx'
    wb = oxl.load_workbook(template_file)
    # dst_active_sheet = wb.create_sheet('Managed Fund Ops', 0)
    dst_active_sheet = wb.active
    dst_active_sheet.title = 'Managed Fund Ops'
    dst_row_num = 5   

    # Read in the official bank ops linked to each fund 
    if env == 'UAT':
        bank_account_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/10 Link Bank Accounts to Fund, Company, Investor - TODO!.xlsx'
    else:
        bank_account_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/10 DEV Link Bank Accounts to Fund, Company, Investor.xlsx'
    share_bank_accounts = pd.read_excel(bank_account_file,index_col=None,skiprows=[0,2,3])
    investor_bank_accounts = pd.read_excel(bank_account_file,index_col=None,sheet_name='Link to Investor',skiprows=[0,2,3])

    # Read the data we need to transform
    all_fund_ops = pd.read_excel(input_file,index_col=None,skiprows=[0,2,3])
    fund_ops = all_fund_ops[pd.to_datetime(all_fund_ops['SETTLEMENTDATE1']) >= dt.datetime(2021,1,1)] 
    missing_bank_accounts = []
    fund_ops = fund_ops.fillna(0)

    for fund_op in fund_ops.itertuples():

        if fund_op.OTYPE1 in ['MF: Net Asset Value','MF: Other','MF: Commitment']:
            pass
        else:
            # Insert the values in the destination row 
            #fund_op.OTYPE1              # Fund Op Type
            #fund_op.FUND2               # Share 
            #fund_op.INVESTOR_NAME_FINV  # Investor
            #fund_op.CLOSEDATE1          # Date
            #fund_op.INDEXOP1            # Index  
            #fund_op.DESCRIPTION1        # Description   
            #fund_op.CURRENCY31          # Currency
            #fund_op.COMMITTEDAMOUNT111  # Committed Amount
            #fund_op.AMOUNT98111         # Nominal 	 
            #fund_op.AMOUNT99111         # Premium
            #fund_op.AMOUNT86111         # Fair Value
            #fund_op.AMOUNT97111         # Return of Premium  


            # Get the bank accounts 
            investor = fund_op.INVESTOR_NAME_FINV            
            if investor in list(investor_bank_accounts['INVESTOR_NAME9']):
                investor_bank_account = investor_bank_accounts[investor_bank_accounts['INVESTOR_NAME9']==investor].iloc[0]['ACCOUNTCODE1']
            else:
                if investor not in missing_bank_accounts:
                    missing_bank_accounts.append(investor)
                investor_bank_account = None
            share = fund_op.FUND2
            share_bank_account = share_bank_accounts[share_bank_accounts['FUND7']==share].iloc[0]['ACCOUNTCODE1']

            if fund_op.OTYPE1 in ['MF: Call']:
                # Call
                amount_investor_ccy = fund_op.AMOUNT98111 + fund_op.AMOUNT99111
                bank_op_type = 'Payment'
            elif fund_op.OTYPE1 in ['MF: Mixed operation','MF: Transfer']:
                # Mixed operation
                amount_investor_ccy = fund_op.AMOUNT98111 + fund_op.AMOUNT99111 - fund_op.AMOUNT97111
                if amount_investor_ccy >= 0:
                    bank_op_type = 'Payment'
                else:
                    bank_op_type = 'Receipt'
            elif fund_op.OTYPE1 == 'MF: Distribution':
                # Distributions
                amount_investor_ccy = fund_op.AMOUNT97111
                bank_op_type = 'Receipt'
            elif fund_op.OTYPE1 == 'MF: Return Of Call (Negative Call)':
                # Return of Call
                amount_investor_ccy = -fund_op.AMOUNT99111
                bank_op_type = 'Receipt'

            if amount_investor_ccy != amount_investor_ccy or amount_investor_ccy == 0:
                print('\tNo amount found:')
                print('\t', fund_op.OTYPE1)
            else:
                amount_investor_ccy = round(amount_investor_ccy,2)
                dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
                dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
                dst_active_sheet[f'C{dst_row_num}'] = fund_op.CLOSEDATE1                # -------------------------------------- Bank Op Date CLOSEDATE1
                dst_active_sheet[f'D{dst_row_num}'] = fund_op.CURRENCY31                # -------------------------------------- Payment currency CURRENCY11
                dst_active_sheet[f'E{dst_row_num}'] = bank_op_type                      # -------------------------------------- Bank Op Type OPTYPE1
                dst_active_sheet[f'F{dst_row_num}'] = investor                          # -------------------------------------- Managed Fund Name LINKEDENTITY
                dst_active_sheet[f'G{dst_row_num}'] = 'Investor Account'                # -------------------------------------- Entity Type XX_ENTITYCLASS
                dst_active_sheet[f'H{dst_row_num}'] = investor                          # -------------------------------------- Entity Name in Bank Op FUND5
                dst_active_sheet[f'I{dst_row_num}'] = 'Fund'                            # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
                dst_active_sheet[f'J{dst_row_num}'] = share                             # -------------------------------------- Counterparty Fund.Fund FUND_CF
                dst_active_sheet[f'K{dst_row_num}'] = share_bank_account                # -------------------------------------- Bank Account BANKACCOUNTC1
                dst_active_sheet[f'L{dst_row_num}'] = investor_bank_account             # -------------------------------------- Bank Account ACCOUNTCODE2
                dst_active_sheet[f'M{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
                dst_active_sheet[f'N{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
                dst_active_sheet[f'O{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
                dst_active_sheet[f'P{dst_row_num}'] = amount_investor_ccy               # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
                #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
                dst_active_sheet[f'Q{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Draft DRAFT1
                dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
                dst_active_sheet[f'S{dst_row_num}'] = fund_op.CLOSEDATE1                # -------------------------------------- CLOSEDATE22
                dst_active_sheet[f'T{dst_row_num}'] = fund_op.INDEXOP1                  # -------------------------------------- INDEXOP22
                dst_active_sheet[f'U{dst_row_num}'] = fund_op.OTYPE1                    # -------------------------------------- Fund Op Type OTYPE22
                dst_active_sheet[f'V{dst_row_num}'] = ''                                # -------------------------------------- Comment
                dst_row_num += 1

    wb.save(dest_file)
    
    # Print missing bank accounts
    if len(missing_bank_accounts) > 0:
        print(f'\n\tNo bank account found for:')
        logging.warning(f'\n\tNo bank account found for:')
        for investor in missing_bank_accounts:
            print(f'\t\t{investor}')
            investor = investor.encode('utf-8')
            logging.warning(f'\t\t{investor}')

def migrate_cash_transfer_bank_ops(dest_file):
    """ Inputs:
            dst_file: the absolute path to the destination file the transformed data will be saved under 
        Extract the relevant data from the cash transfers file and rearrange it to fit the bank ops import template 
    """
    # Read in the template file - this already has the starting balances
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/BankOp_Test.xlsx'
    wb = oxl.load_workbook(template_file)
    # dst_active_sheet = wb.create_sheet('Cash Transfers', 0)
    dst_active_sheet = wb.active
    dst_active_sheet.title = 'Cash Transfers'
    dst_row_num = 5 # Ignore headers and overwrite the starting balances

    cash_transfers_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/LR Data Files/Cash Transfers/Cash Pool BNP.xlsx'
    cash_transfers = pd.read_excel(cash_transfers_file,index_col=None)

    # Drop totals row
    cash_transfers = cash_transfers[0:-1]

    # Format date of cash operation 
    # cash_transfers['Jr/mnd'] = cash_transfers['Jr/mnd'] + '/01'
    # cash_transfers['Jr/mnd'] = pd.to_datetime(cash_transfers['Jr/mnd'],infer_datetime_format=True)
    cash_transfers['Boek.datum'] = cash_transfers['Boek.datum'].apply(lambda x: dt.date(x.year,x.month,x.day))

    # Drop all cash operations prior to 1st Jan 2021
    cash_transfers = cash_transfers[cash_transfers['Boek.datum'] >= dt.date(2021,1,1)]

    #Format columns and drop totals row
    cash_transfers.columns = [c.replace(' ', '_') for c in cash_transfers.columns]
    cash_transfers.columns = [c.replace('/', '_') for c in cash_transfers.columns]
    cash_transfers.columns = [c.replace('.', '_') for c in cash_transfers.columns]
    
    for cash_transfer in cash_transfers.itertuples():
     
        if cash_transfer.Bedrag_EUR >= 0:
            cp_bank_op_type = 'Cash Transfer Out'
            rabo_bank_op_type = 'Cash Transfer In'
            amount = cash_transfer.Bedrag_EUR
        else:
            cp_bank_op_type = 'Cash Transfer In'
            rabo_bank_op_type = 'Cash Transfer Out'
            amount = -cash_transfer.Bedrag_EUR

        # Cash Pool Entry
        dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
        dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
        dst_active_sheet[f'C{dst_row_num}'] = cash_transfer.Boek_datum          # -------------------------------------- Bank Op Date CLOSEDATE1
        dst_active_sheet[f'D{dst_row_num}'] = cash_transfer.Eigen_valuta        # -------------------------------------- Payment currency CURRENCY11
        dst_active_sheet[f'E{dst_row_num}'] = cp_bank_op_type                   # -------------------------------------- Bank Op Type OPTYPE1
        dst_active_sheet[f'F{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Managed Fund Name LINKEDENTITY
        dst_active_sheet[f'G{dst_row_num}'] = 'Fund'                            # -------------------------------------- Entity Type XX_ENTITYCLASS
        dst_active_sheet[f'H{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Entity Name in Bank Op FUND5
        dst_active_sheet[f'I{dst_row_num}'] = 'Fund'                            # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
        dst_active_sheet[f'J{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Counterparty Fund.Fund FUND_CF
        dst_active_sheet[f'K{dst_row_num}'] = 'Cash Pool BNP'                   # -------------------------------------- Bank Account BANKACCOUNTC1
        dst_active_sheet[f'L{dst_row_num}'] = 'Rabobank EUR'                    # -------------------------------------- Bank Account ACCOUNTCODE2
        dst_active_sheet[f'M{dst_row_num}'] = amount                            # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
        dst_active_sheet[f'N{dst_row_num}'] = amount                            # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
        dst_active_sheet[f'O{dst_row_num}'] = amount                            # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
        dst_active_sheet[f'P{dst_row_num}'] = amount                            # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
        #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
        dst_active_sheet[f'Q{dst_row_num}'] = 'FALSE'                            # -------------------------------------- Draft DRAFT1
        dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
        dst_active_sheet[f'S{dst_row_num}'] = cash_transfer.Boek_datum          # -------------------------------------- CLOSEDATE22
        dst_active_sheet[f'T{dst_row_num}'] = 1                                 # -------------------------------------- INDEXOP22
        dst_active_sheet[f'U{dst_row_num}'] = ''                                # -------------------------------------- Fund Op Type OTYPE22
        dst_active_sheet[f'V{dst_row_num}'] = cash_transfer.Tekst               # -------------------------------------- Comment
        dst_row_num += 1

        # Rabobank Entry 
        dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
        dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
        dst_active_sheet[f'C{dst_row_num}'] = cash_transfer.Boek_datum          # -------------------------------------- Bank Op Date CLOSEDATE1
        dst_active_sheet[f'D{dst_row_num}'] = cash_transfer.Eigen_valuta        # -------------------------------------- Payment currency CURRENCY11
        dst_active_sheet[f'E{dst_row_num}'] = rabo_bank_op_type                 # -------------------------------------- Bank Op Type OPTYPE1
        dst_active_sheet[f'F{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Managed Fund Name LINKEDENTITY
        dst_active_sheet[f'G{dst_row_num}'] = 'Fund'                            # -------------------------------------- Entity Type XX_ENTITYCLASS
        dst_active_sheet[f'H{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Entity Name in Bank Op FUND5
        dst_active_sheet[f'I{dst_row_num}'] = 'Fund'                            # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
        dst_active_sheet[f'J{dst_row_num}'] = 'A Shares - LR'                   # -------------------------------------- Counterparty Fund.Fund FUND_CF
        dst_active_sheet[f'K{dst_row_num}'] = 'Rabobank EUR'                    # -------------------------------------- Bank Account BANKACCOUNTC1
        dst_active_sheet[f'L{dst_row_num}'] = 'Cash Pool BNP'                   # -------------------------------------- Bank Account ACCOUNTCODE2
        dst_active_sheet[f'M{dst_row_num}'] = amount                            # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
        dst_active_sheet[f'N{dst_row_num}'] = amount                            # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
        dst_active_sheet[f'O{dst_row_num}'] = amount                            # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
        dst_active_sheet[f'P{dst_row_num}'] = amount                            # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
        #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
        dst_active_sheet[f'Q{dst_row_num}'] = 'FALSE'                            # -------------------------------------- Draft DRAFT1
        dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
        dst_active_sheet[f'S{dst_row_num}'] = cash_transfer.Boek_datum          # -------------------------------------- CLOSEDATE22
        dst_active_sheet[f'T{dst_row_num}'] = 1                                 # -------------------------------------- INDEXOP22
        dst_active_sheet[f'U{dst_row_num}'] = ''                                # -------------------------------------- Fund Op Type OTYPE22
        dst_active_sheet[f'V{dst_row_num}'] = cash_transfer.Tekst               # -------------------------------------- Comment
        dst_row_num += 1

    wb.save(dest_file)

def migrate_managed_fee_bank_ops(dest_file,env='UAT'):
    """ Inputs:
            dst_file: the absolute path to the destination file the transformed data will be saved under 
        Extract the relevant data from the managed fund fees file and rearrange it to fit the bank ops import template 
    """
   # Read in the template file - this already has the starting balances
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/BankOp_Test.xlsx'
    wb = oxl.load_workbook(template_file)
    # dst_active_sheet = wb.create_sheet('Managed Fees', 0)
    dst_active_sheet = wb.active
    dst_active_sheet.title = 'Managed Fees'
    dst_row_num = 5 # Ignore headers and overwrite the starting balances

    # Read in managed fees data and format it 
    if env == 'UAT':
        managed_fees_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/15 Managed Fund Fees.xlsx'
        investor_bank_accounts_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/10 Link Bank Accounts to Fund, Company, Investor - TODO!.xlsx'
    else:
        managed_fees_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/15 DEV Managed Fund Fees.xlsx'
        investor_bank_accounts_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/10 DEV Link Bank Accounts to Fund, Company, Investor.xlsx'
    managed_fees = pd.read_excel(managed_fees_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Managed Fees') 
    managed_fees.drop(columns=['Fee','Standard','IQREGIONID1','DRAFT1'],inplace=True)
    # managed_fees['SETTLEMENTDATE1'] = managed_fees['SETTLEMENTDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day)) # - some nulls here
    managed_fees['REFERENCEDATE1'] = managed_fees['REFERENCEDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day))
    managed_fees['ACCOUNTINGDATE1'] = managed_fees['ACCOUNTINGDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day))
    managed_fees['EXPECTEDPAYMENTDATE1'] = managed_fees['EXPECTEDPAYMENTDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day))  
    # Format dates and drop all fees prior to 1st Jan 2021
    managed_fees = managed_fees[managed_fees['REFERENCEDATE1'] >= dt.date(2017,1,1)] #!!!! Check if this should be EXPECTEDPAYMENTDATE1, SETTLEMENTDATE1 or ACCOUNTINGDATE1
    
    
    investor_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Investor')
    investor_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True) 
    company_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Company')
    company_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True)
    fund_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Fund')
    fund_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True)

    for managed_fee in managed_fees.itertuples():
        # Get the main party bank accounts
        # main_bank_account = fund_bank_accounts.loc[fund_bank_accounts['FUND7'] == managed_fee.FUND2,'ACCOUNTCODE1']
        main_bank_account = 'Rabobank EUR'
        main_party = managed_fee.FUND2

        # Get the counterparty bank accounts
        if managed_fee.XX_COUNTERPARTYCLASS == 'Investor Account':
            investor_account = managed_fee.XX_COUNTERPARTY
            investor = investor_account.split(' in ')[0]
            try:
                cp_bank_account = investor_bank_accounts.loc[investor_bank_accounts['INVESTOR_NAME9'] == investor,'ACCOUNTCODE1'].iloc[0]
            except IndexError:
                print(main_party, investor)
            counterparty = investor_account
        elif managed_fee.XX_COUNTERPARTYCLASS == 'Company':
            cp_bank_account = company_bank_accounts.loc[company_bank_accounts['NAME6'] == managed_fee.XX_COUNTERPARTY,'ACCOUNTCODE1']
        elif managed_fee.XX_COUNTERPARTYCLASS == 'Fee':
            cp_bank_account = fund_bank_accounts.loc[fund_bank_accounts['FUND7'] == managed_fee.XX_COUNTERPARTY,'ACCOUNTCODE1']

        if managed_fee.DUEAMOUNT21 < 0:
            bank_op_type = 'Payment'
            amount = -managed_fee.DUEAMOUNT21
        else:
            bank_op_type = 'Receipt'
            amount = managed_fee.DUEAMOUNT21

        dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
        dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
        dst_active_sheet[f'C{dst_row_num}'] = managed_fee.REFERENCEDATE1        # -------------------------------------- Bank Op Date CLOSEDATE1
        dst_active_sheet[f'D{dst_row_num}'] = 'EUR'                             # -------------------------------------- Payment currency CURRENCY11
        dst_active_sheet[f'E{dst_row_num}'] = bank_op_type                      # -------------------------------------- Bank Op Type OPTYPE1
        dst_active_sheet[f'F{dst_row_num}'] = main_party                        # -------------------------------------- Managed Fund Name LINKEDENTITY
        dst_active_sheet[f'G{dst_row_num}'] = 'Fund'                            # -------------------------------------- Entity Type XX_ENTITYCLASS
        dst_active_sheet[f'H{dst_row_num}'] = main_party                        # -------------------------------------- Entity Name in Bank Op FUND5
        dst_active_sheet[f'I{dst_row_num}'] = managed_fee.XX_COUNTERPARTYCLASS  # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
        dst_active_sheet[f'J{dst_row_num}'] = ''                                # -------------------------------------- Counterparty Fund.Fund FUND_CF
        dst_active_sheet[f'K{dst_row_num}'] = main_bank_account                 # -------------------------------------- Bank Account BANKACCOUNTC1
        dst_active_sheet[f'L{dst_row_num}'] = cp_bank_account                   # -------------------------------------- Bank Account ACCOUNTCODE2
        dst_active_sheet[f'M{dst_row_num}'] = amount                            # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
        dst_active_sheet[f'N{dst_row_num}'] = amount                            # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
        dst_active_sheet[f'O{dst_row_num}'] = amount                            # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
        dst_active_sheet[f'P{dst_row_num}'] = amount                            # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
        #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
        dst_active_sheet[f'Q{dst_row_num}'] = 'TRUE'                            # -------------------------------------- Draft DRAFT1
        dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
        dst_active_sheet[f'S{dst_row_num}'] = managed_fee.REFERENCEDATE1        # -------------------------------------- CLOSEDATE22
        dst_active_sheet[f'T{dst_row_num}'] = managed_fee.INDEXOP1              # -------------------------------------- INDEXOP22
        dst_active_sheet[f'U{dst_row_num}'] = ''                                # -------------------------------------- Fund Op Type OTYPE22
        dst_active_sheet[f'V{dst_row_num}'] = ''                                # -------------------------------------- Comment
        dst_active_sheet[f'W{dst_row_num}'] = counterparty                      # -------------------------------------- Counterparty XX_COUNTERPARTY
        dst_row_num += 1

    wb.save(dest_file)

def migrate_fee_and_income_bank_ops(dest_file,env='UAT'):
    """ Inputs:
            dst_file: the absolute path to the destination file the transformed data will be saved under 
        Extract the relevant data from the managed fund fees file and rearrange it to fit the bank ops import template 
    """
    # Read in the template file - this already has the starting balances
    template_file = 'C:/Users/RajContractor/OneDrive - IT-Venture Ltd/Documents/Temp/BankOp_Test.xlsx'
    wb = oxl.load_workbook(template_file)
    # dst_active_sheet = wb.create_sheet('Fees and Incomes', 0)
    dst_active_sheet = wb.active
    dst_active_sheet.title = 'Fees and Incomes'
    dst_row_num = 5 # Ignore headers and overwrite the starting balances

    # Read in fees data and merge it together
    if env == 'UAT':
        fees_and_incomes = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/17 Fees and Incomes.xlsx'
        investor_bank_accounts_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/2 UAT Import Files/10 Link Bank Accounts to Fund, Company, Investor - TODO!.xlsx'
    else:
        fees_and_incomes = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/17 Fees and Incomes.xlsx'
        investor_bank_accounts_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/1 DEV Import Files/10 DEV Link Bank Accounts to Fund, Company, Investor.xlsx'
    fees = pd.read_excel(fees_and_incomes,index_col=None,usecols='C:W',skiprows=[0,2,3],sheet_name = 'FEES') 
    income_tax = pd.read_excel(fees_and_incomes,index_col=None,usecols='C:W',skiprows=[0,2,3],sheet_name = 'FEES - CURRENT INCOME TAX')
    incomes = pd.read_excel(fees_and_incomes,index_col=None,usecols='C:W',skiprows=[0,2,3],sheet_name = 'INCOMES')
    fees = fees.append(income_tax, ignore_index=True)
    fees = fees.append(incomes, ignore_index=True)

    fees.drop(columns=['FEESINVOICEDATE1','FEESINVOICENUMBER1','BANKACCOUNT1'],inplace=True)
    # replace EXP57 with 'Current Income Tax'
    fees['FTYPE1'] = fees['FTYPE1'].replace('EXP57','Current Income Tax')     
    fees['REFERENCEDATE1'] = fees['REFERENCEDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day))
    fees['EXPECTEDPAYMENTDATE1'] = fees['EXPECTEDPAYMENTDATE1'].apply(lambda x: dt.date(x.year,x.month,x.day))  
    # Drop all fees and incomes prior to 1st Jan 2021
    fees = fees[fees['REFERENCEDATE1'] >= dt.date(2017,1,1)] #!!!! Check if this should be EXPECTEDPAYMENTDATE1, SETTLEMENTDATE1 or ACCOUNTINGDATE1
    
    
    investor_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Investor')
    investor_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True) 
    company_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Company')
    company_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True)
    fund_bank_accounts = pd.read_excel(investor_bank_accounts_file,index_col=None,skiprows=[0,2,3],sheet_name = 'Link to Fund')
    fund_bank_accounts.drop(columns=['BankAccount','Standard'],inplace=True)
    missing_bank_accounts_companies = []
    missing_bank_accounts_funds = []

    for fee in fees.itertuples():
        # Get the main party bank accounts
        # main_bank_account = fund_bank_accounts.loc[fund_bank_accounts['FUND7'] == managed_fee.FUND2,'ACCOUNTCODE1']
        main_bank_account = 'Rabobank EUR'
        main_party = fee.FUND2

        # Get the counterparty bank accounts
        if fee.NAME_CA == fee.NAME_CA:
            # Company
            cp_bank_accounts = company_bank_accounts.loc[company_bank_accounts['NAME6'] == fee.NAME_CA,'ACCOUNTCODE1'].unique()
            if len(cp_bank_accounts) == 1:
                cp_bank_account = cp_bank_accounts.item()
            elif len(cp_bank_accounts) == 0:
                if fee.NAME_CA not in missing_bank_accounts_companies:
                    missing_bank_accounts_companies.append(fee.NAME_CA)
                cp_bank_account = None
            else:
                if 'Rabobank EUR' in cp_bank_accounts:
                    cp_bank_account = 'Rabobank EUR'
                else:
                    print(f"Multiple bank accounts found for Company: {fee.NAME_CA}")
                    for cp_bank_account in cp_bank_accounts:
                        print(f"\tBank Account: {cp_bank_account}")
            counterparty_type = 'Company'
            counterparty = fee.NAME_CA
        elif fee.FUND_CF == fee.FUND_CF:
            # Fund - there shouldn't be any
            cp_bank_account = fund_bank_accounts.loc[fund_bank_accounts['FUND7'] == fee.FUND_CF,'ACCOUNTCODE1'].unique()
            if len(cp_bank_accounts) == 1:
                cp_bank_account = cp_bank_accounts.item()
            elif len(cp_bank_accounts) == 0:
                if fee.FUND_CF not in missing_bank_accounts_funds:
                    missing_bank_accounts_funds.append(fee.FUND_CF)
                cp_bank_account = None
            else:
                if 'Rabobank EUR' in cp_bank_accounts:
                    cp_bank_account = 'Rabobank EUR'
                else:
                    print(f"Multiple bank accounts found for Fund: {fee.FUND_CF}")
                    for cp_bank_account in cp_bank_accounts:
                        print(f"\tBank Account: {cp_bank_account}")
            counterparty_type = 'Fund'
            counterparty = fee.FUND_CF

        if fee.DUEAMOUNT1 < 0:
            bank_op_type = 'Receipt'
            amount = -fee.DUEAMOUNT1
        else:
            bank_op_type = 'Payment'
            amount = fee.DUEAMOUNT1

        dst_active_sheet[f'A{dst_row_num}'] = ''                                # -------------------------------------- BankOperation - LEAVE BLAN
        dst_active_sheet[f'B{dst_row_num}'] = ''                                # -------------------------------------- PaymentAllocation - LEAVE BLANK
        dst_active_sheet[f'C{dst_row_num}'] = fee.REFERENCEDATE1                # -------------------------------------- Bank Op Date CLOSEDATE1
        dst_active_sheet[f'D{dst_row_num}'] = 'EUR'                             # -------------------------------------- Payment currency CURRENCY11
        dst_active_sheet[f'E{dst_row_num}'] = bank_op_type                      # -------------------------------------- Bank Op Type OPTYPE1
        dst_active_sheet[f'F{dst_row_num}'] = main_party                        # -------------------------------------- Managed Fund Name LINKEDENTITY
        dst_active_sheet[f'G{dst_row_num}'] = 'Fund'                            # -------------------------------------- Entity Type XX_ENTITYCLASS
        dst_active_sheet[f'H{dst_row_num}'] = main_party                        # -------------------------------------- Entity Name in Bank Op FUND5
        dst_active_sheet[f'I{dst_row_num}'] = counterparty_type                 # -------------------------------------- Counterparty Type XX_COUNTERPARTYCLASS
        dst_active_sheet[f'J{dst_row_num}'] = ''                                # -------------------------------------- Counterparty Fund.Fund FUND_CF
        dst_active_sheet[f'K{dst_row_num}'] = main_bank_account                 # -------------------------------------- Bank Account BANKACCOUNTC1
        dst_active_sheet[f'L{dst_row_num}'] = cp_bank_account                   # -------------------------------------- Bank Account ACCOUNTCODE2
        dst_active_sheet[f'M{dst_row_num}'] = amount                            # -------------------------------------- Amount (Bank) AMOUNT1 -- EUR
        dst_active_sheet[f'N{dst_row_num}'] = amount                            # -------------------------------------- Amount (Counterparty) AMOUNTC1 -- USD
        dst_active_sheet[f'O{dst_row_num}'] = amount                            # -------------------------------------- Amount (Payment) AMOUNT21 -- EUR
        dst_active_sheet[f'P{dst_row_num}'] = amount                            # -------------------------------------- Amount (Entity) AMOUNT31 -- EUR
        #(AMOUNTCB1 - VCBANKACCTOP.AMOUNTCB - Counterparty Bank Currency)          
        dst_active_sheet[f'Q{dst_row_num}'] = 'TRUE'                            # -------------------------------------- Draft DRAFT1
        dst_active_sheet[f'R{dst_row_num}'] = 'FALSE'                           # -------------------------------------- Locked XX_LOCKED
        dst_active_sheet[f'S{dst_row_num}'] = fee.REFERENCEDATE1        # -------------------------------------- CLOSEDATE22
        dst_active_sheet[f'T{dst_row_num}'] = fee.INDEXOP1              # -------------------------------------- INDEXOP22
        dst_active_sheet[f'U{dst_row_num}'] = ''                                # -------------------------------------- Fund Op Type OTYPE22
        dst_active_sheet[f'V{dst_row_num}'] = ''                                # -------------------------------------- Comment
        dst_active_sheet[f'W{dst_row_num}'] = counterparty                      # -------------------------------------- Counterparty XX_COUNTERPARTY
        dst_row_num += 1

    wb.save(dest_file)

    # Warn about any missing bank accounts
    
    if len(missing_bank_accounts_companies) > 0:
        print('\tNo bank account found for the following companies:')
        for company in missing_bank_accounts_companies:
            print('\t\t',company)
    if len(missing_bank_accounts_funds) > 0:
        print('\tNo bank account found for the following funds:')
        for fund in missing_bank_accounts_funds:
            print('\t\t',fund)
