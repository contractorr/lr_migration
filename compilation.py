##############################################################################################################
from datetime import date
import log 
logger = log.get_logger('root')
import glob
import os 
import mig_functions as mig
import openpyxl as oxl
import pandas as pd
from fuzzywuzzy import process

# Destination of key files
os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrate These')
#os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrate - All Files')
#os.chdir(r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Test')
input_files = glob.glob('[!~]*.xlsx')            
dest_directory = r'C:\Users\RajContractor\Documents\Python Files\Dev\LR Migration\Migrated'
log.update_handler(logger,'Compilation')

# File with the official fund names
fund_import_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/ITV Import Files/04 Funds.xlsx'
##############################################################################################################

# Define any functions we plan to use
def fix_fund_name(df_mapping):
    df_mapping['fund'] = df_mapping['Investee Fund'].copy()
    for row in df_mapping.itertuples():
        share = row.Share
        share = share.split(' ')[0] 
        fund_import = pd.read_excel(fund_import_file, index_col=None, header=2, usecols='D:Q',sheet_name=share)[2:] # First fund is the management fund 
        fund_import = fund_import['Fund'].to_list()
        fund_name_compare = mig.replace_fund_name(share,row.fund)
        fund_name = process.extractOne(fund_name_compare,fund_import)[0]
        df_mapping.loc[row.Index,'fund'] = fund_name
    
    return df_mapping

# Replace backslashes in dest_directory
dest_directory = dest_directory.replace('\\','/')
dest_file = dest_directory + f'/Compilation.xlsx' 
wb = oxl.Workbook()
ws_fees = wb.active
ws_fees.title = "Fees"
ws_roc = wb.create_sheet("Return of Capital")
ws_cg = wb.create_sheet("Capital Gains")
all_rows = pd.DataFrame()

dst_row_num_fees = 2
dst_row_num_roc = 2
dst_row_num_cg = 2

ws_fees[f'A1'] = 'Share'
ws_fees[f'B1'] = 'Investee Fund'
ws_fees[f'C1'] = 'Date'
ws_fees[f'D1'] = 'Description'
ws_fees[f'E1'] = 'Fee Amount'
ws_fees[f'F1'] = 'Return of Cap Amount'
ws_fees[f'G1'] = 'Capital Gains Amount'
ws_fees[f'H1'] = 'Mapping'
ws_roc[f'A1'] = 'Share'
ws_roc[f'B1'] = 'Investee Fund'
ws_roc[f'C1'] = 'Date'
ws_roc[f'D1'] = 'Description'
ws_roc[f'E1'] = 'Mapping'
ws_roc[f'F1'] = 'Question'
ws_roc[f'G1'] = 'Fee Amount'
ws_roc[f'H1'] = 'Return of Cap Amount'
ws_roc[f'I1'] = 'Capital Gains Amount'
ws_cg[f'A1'] = 'Share'
ws_cg[f'B1'] = 'Investee Fund'
ws_cg[f'C1'] = 'Date'
ws_cg[f'D1'] = 'Description'
ws_cg[f'E1'] = 'Mapping'
ws_cg[f'F1'] = 'Fund Op'
ws_cg[f'G1'] = 'Recallable/Non Recallable?'
ws_cg[f'H1'] = 'Question'
ws_cg[f'I1'] = 'Fee Amount'
ws_cg[f'J1'] = 'Return of Cap Amount'
ws_cg[f'K1'] = 'Capital Gains Amount'

for input_file in input_files:
    #if input_file == 'LRI_O shares@2021 Q1.xlsx':
    print(f'{input_file}-----------------------------------------------------------------')
    # Extract the file name and format it 
    if 'LRI_' in input_file:
        file_name = input_file.split("LRI_")[1].split("@")[0]
        file_name = file_name.replace('shares','Shares')
        prefix = file_name.split(' ')[0]
    else:
        file_name = 'Y Shares'   
        prefix = 'Y' 
    dst_rows = mig.compile_data(input_file, file_name)
    dst_rows.drop(list(dst_rows[dst_rows['fund_op_type'] == 'IF: Commitment'].index),inplace=True)
    all_rows = all_rows.append(dst_rows,ignore_index=True)


mapping_file = 'C:/Users/RajContractor/IT-Venture Ltd/Lion River - Documents/Import Files/Compilation - All Fees, Return of Capital and Capital Gains.xlsx'
mapping_file_df = pd.ExcelFile(mapping_file)
sheets = mapping_file_df.sheet_names[1:]
for sheet in sheets:
    contents = pd.read_excel(open(mapping_file,'rb'), sheet_name=sheet)
    contents = contents[(contents['Share'] == 'N Shares')|(contents['Share'] == 'O Shares')]
    if sheet == 'Fees':
        fees_mapping = fix_fund_name(contents)
    elif sheet == 'Return of Capital':
        roc_mapping = fix_fund_name(contents)
    elif sheet == 'Capital Gains':
        cap_mapping = fix_fund_name(contents)
all_rows['fees_fund_ccy'] = all_rows['fees_fund_ccy'].round(2) 
all_rows['roc_fund_ccy'] = all_rows['roc_fund_ccy'].round(2) 
all_rows['capital_gains_fund_ccy'] = all_rows['capital_gains_fund_ccy'].round(2) 
current_investee_fund = None



# Insert
for dst_row in all_rows.itertuples():
    if current_investee_fund == None or dst_row.fund != current_investee_fund:
        current_investee_fund = dst_row.fund
    elif dst_row.fund == current_investee_fund:
        current_investee_fund 
    if dst_row.fees_fund_ccy != 0:
        share = dst_row.investor
        share = share.split(' - ')[0]
        investee_fund = dst_row.fund_name 
        fee_amount =  dst_row.fees_fund_ccy 
        roc_amount = dst_row.roc_fund_ccy
        opp_date = dst_row.date
        description = dst_row.description
        cg_amount = dst_row.capital_gains_fund_ccy
        investor = dst_row.fund
        fund_op_type = dst_row.fund_op_type
        undrawn = dst_row.undrawn_fund_ccy
        # Find our mapping
        relevant_rows = fees_mapping[(fees_mapping['fund']==investee_fund)&(fees_mapping['Share']==share)&(fees_mapping['Description']==description)&(fees_mapping['Date']==opp_date)&(fees_mapping['Fee Amount']==fee_amount)&(fees_mapping['Return of Cap Amount']==roc_amount)&(fees_mapping['Capital Gains Amount']==cg_amount)]
        if len(relevant_rows) == 0:
            print(f'ROW NOT FOUND - FEE - {share} {investee_fund} {description} {opp_date} {fee_amount} {roc_amount} {cg_amount}')
            mapping = None
        else:
            relevant_rows.reset_index(inplace=True)
            mapping = relevant_rows.loc[0,'Mapping']
        # Insert the values in the destination row 
        ws_fees[f'A{dst_row_num_fees}'] = share
        ws_fees[f'B{dst_row_num_fees}'] = investee_fund
        ws_fees[f'C{dst_row_num_fees}'] = opp_date
        ws_fees[f'D{dst_row_num_fees}'] = description
        ws_fees[f'E{dst_row_num_fees}'] = fee_amount
        ws_fees[f'F{dst_row_num_fees}'] = roc_amount
        ws_fees[f'G{dst_row_num_fees}'] = cg_amount
        ws_fees[f'H{dst_row_num_fees}'] = mapping  
        dst_row_num_fees = dst_row_num_fees + 1
    elif dst_row.roc_fund_ccy != 0:
        share = dst_row.investor
        share = share.split(' - ')[0]
        investee_fund = dst_row.fund_name 
        fee_amount =  dst_row.fees_fund_ccy 
        roc_amount = dst_row.roc_fund_ccy
        opp_date = dst_row.date
        description = dst_row.description
        cg_amount = dst_row.capital_gains_fund_ccy
        investor = dst_row.fund
        fund_op_type = dst_row.fund_op_type
        undrawn = dst_row.undrawn_fund_ccy
        # Find our mapping
        relevant_rows = roc_mapping[(roc_mapping['fund']==investee_fund)&(roc_mapping['Share']==share)&(roc_mapping['Description']==description)&(roc_mapping['Date']==opp_date)&(roc_mapping['Fee Amount']==fee_amount)&(roc_mapping['Return of Cap Amount']==roc_amount)&(roc_mapping['Capital Gains Amount']==cg_amount)]
        relevant_rows.reset_index(inplace=True)
        if len(relevant_rows) == 0:
            print(f'ROW NOT FOUND - ROC - {share} {investee_fund} {description} {opp_date} {fee_amount} {roc_amount} {cg_amount}')
            mapping = None
            question = None
        else:
            mapping = relevant_rows.loc[0,'Mapping']  
            question = relevant_rows.loc[0,'Question']     
        # Insert the values in the destination row 
        ws_roc[f'A{dst_row_num_roc}'] = share
        ws_roc[f'B{dst_row_num_roc}'] = investee_fund
        ws_roc[f'C{dst_row_num_roc}'] = opp_date
        ws_roc[f'D{dst_row_num_roc}'] = description
        ws_roc[f'E{dst_row_num_roc}'] = mapping
        ws_roc[f'F{dst_row_num_roc}'] = question
        ws_roc[f'G{dst_row_num_roc}'] = fee_amount
        ws_roc[f'H{dst_row_num_roc}'] = roc_amount
        ws_roc[f'I{dst_row_num_roc}'] = cg_amount
        dst_row_num_roc = dst_row_num_roc + 1
    elif dst_row.capital_gains_fund_ccy != 0:
        share = dst_row.investor
        share = share.split(' - ')[0]
        investee_fund = dst_row.fund_name 
        fee_amount =  dst_row.fees_fund_ccy 
        roc_amount = dst_row.roc_fund_ccy
        opp_date = dst_row.date
        description = dst_row.description
        cg_amount = dst_row.capital_gains_fund_ccy
        investor = dst_row.fund
        fund_op_type = dst_row.fund_op_type
        undrawn = dst_row.undrawn_fund_ccy
        # Find our mapping
        relevant_rows = cap_mapping[(cap_mapping['fund']==investee_fund)&(cap_mapping['Share']==share)&(cap_mapping['Description']==description)&(cap_mapping['Date']==opp_date)&(cap_mapping['Fee Amount']==fee_amount)&(cap_mapping['Return of Cap Amount']==roc_amount)&(cap_mapping['Capital Gains Amount']==cg_amount)]
        relevant_rows.reset_index(inplace=True)
        if len(relevant_rows) == 0:
            print(f'ROW NOT FOUND - CAP - {share} {investee_fund} {description} {opp_date} {fee_amount} {roc_amount} {cg_amount}')
            mapping = None
            question = None
        else:
            mapping = relevant_rows.loc[0,'Mapping']  
            question = relevant_rows.loc[0,'Question']  
        # Insert the values in the destination row 
        ws_cg[f'A{dst_row_num_cg}'] = share
        ws_cg[f'B{dst_row_num_cg}'] = investee_fund
        ws_cg[f'C{dst_row_num_cg}'] = opp_date
        ws_cg[f'D{dst_row_num_cg}'] = description
        ws_cg[f'E{dst_row_num_cg}'] = mapping
        ws_cg[f'F{dst_row_num_cg}'] = 'IF Distribution'
        ws_cg[f'G{dst_row_num_cg}'] = '-'
        ws_cg[f'H{dst_row_num_cg}'] = question
        ws_cg[f'I{dst_row_num_cg}'] = fee_amount
        ws_cg[f'J{dst_row_num_cg}'] = roc_amount
        ws_cg[f'K{dst_row_num_cg}'] = cg_amount
        dst_row_num_cg = dst_row_num_cg + 1


wb.save(dest_file)

   


